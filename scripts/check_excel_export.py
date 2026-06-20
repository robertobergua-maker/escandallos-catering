import ast
import io
import re
import unicodedata
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def cargar_funciones_excel():
    source = Path("generador_fichas.py").read_text(encoding="utf-8")
    module = ast.parse(source)
    needed_assigns = {
        "PALABRAS_NO_ALIMENTARIAS",
        "PALABRAS_BLOQUEO_MATCHING",
        "FRASES_BLOQUEO_MATCHING",
        "INGREDIENTES_INCOHERENTES_ARROZ",
    }
    nodes = [
        node for node in module.body
        if (
            isinstance(node, ast.Assign)
            and any(isinstance(target, ast.Name) and target.id in needed_assigns for target in node.targets)
        )
        or isinstance(node, ast.FunctionDef) and node.name in {
            "normalizar_texto_busqueda",
            "parece_material_no_alimentario",
            "descripcion_generica_ingrediente",
            "contiene_bloqueo_matching",
            "bloqueo_pedido_explicitamente",
            "sugerir_ingredientes_similares",
            "inferir_unidad_medida",
            "receta_parece_arroz",
            "ingrediente_incoherente_en_arroz",
            "motivo_indica_presencia_explicita",
            "normalizar_unidad_ia",
            "convertir_cantidad_ia",
            "buscar_coincidencia_principal_inventario",
            "ajustar_ingredientes_por_raciones",
            "calcular_ajuste_raciones",
            "normalizar_respuesta_ingredientes_ia",
            "generar_excel",
        }
    ]
    compiled = compile(ast.Module(body=nodes, type_ignores=[]), "generador_fichas.py", "exec")
    namespace = {
        "openpyxl": openpyxl,
        "Font": Font,
        "Alignment": Alignment,
        "PatternFill": PatternFill,
        "get_column_letter": get_column_letter,
        "io": io,
        "pd": pd,
        "re": re,
        "unicodedata": unicodedata,
        "SequenceMatcher": SequenceMatcher,
        "inventario_df": pd.DataFrame(),
    }
    exec(compiled, namespace)
    return (
        namespace["ajustar_ingredientes_por_raciones"],
        namespace["calcular_ajuste_raciones"],
        namespace["normalizar_respuesta_ingredientes_ia"],
        namespace["parece_material_no_alimentario"],
        namespace["sugerir_ingredientes_similares"],
        namespace["generar_excel"],
    )


def comprobar_ajuste_raciones(ajustar_ingredientes_por_raciones, calcular_ajuste_raciones):
    ingredientes = [
        {"codigo": "ACE-01", "descripcion": "Aceite", "cantidad_bruta": 2.0, "merma": 5.0, "precio_unidad": 1.69},
        {"codigo": "HAR-01", "descripcion": "Harina", "cantidad_bruta": 4.0, "merma": 10.0, "precio_unidad": 1.1128},
    ]

    duplicados = ajustar_ingredientes_por_raciones(ingredientes, 20 / 10)
    reducidos = ajustar_ingredientes_por_raciones(ingredientes, 5 / 10)

    assert duplicados[0]["cantidad_bruta"] == 4.0
    assert duplicados[1]["cantidad_bruta"] == 8.0
    assert reducidos[0]["cantidad_bruta"] == 1.0
    assert reducidos[1]["cantidad_bruta"] == 2.0
    assert duplicados[0]["precio_unidad"] == ingredientes[0]["precio_unidad"]
    assert duplicados[1]["merma"] == ingredientes[1]["merma"]
    assert duplicados[0]["codigo"] == ingredientes[0]["codigo"]
    assert duplicados[1]["descripcion"] == ingredientes[1]["descripcion"]

    factor_12, ajustados_12 = calcular_ajuste_raciones(ingredientes, 6, 12)
    factor_3, ajustados_3 = calcular_ajuste_raciones(ingredientes, 6, 3)
    factor_18, ajustados_18 = calcular_ajuste_raciones(ingredientes, 6, 18)
    assert factor_12 == 2.0
    assert ajustados_12[0]["cantidad_bruta"] == 4.0
    assert factor_3 == 0.5
    assert ajustados_3[0]["cantidad_bruta"] == 1.0
    assert factor_18 == 3.0
    assert ajustados_18[0]["cantidad_bruta"] == 6.0


def comprobar_normalizacion_ia(normalizar_respuesta_ingredientes_ia):
    ingrediente = {"codigo": "ING-0001", "descripcion": "TOMATE PERA", "cantidad_bruta": 1.0, "merma": 0, "precio_unidad": 2.5}
    _, raciones, ingredientes = normalizar_respuesta_ingredientes_ia([ingrediente])
    assert raciones is None
    assert ingredientes[0]["descripcion"].upper() == "TOMATE PERA"
    assert ingredientes[0]["unidad_medida"] == "kg"

    _, raciones, ingredientes = normalizar_respuesta_ingredientes_ia({
        "raciones_base": 6,
        "ingredientes": [ingrediente],
    })
    assert raciones == 6.0
    assert ingredientes[0]["descripcion"].upper() == "TOMATE PERA"
    assert ingredientes[0]["unidad_medida"] == "kg"

    _, raciones, ingredientes = normalizar_respuesta_ingredientes_ia([
        {"codigo": "X", "descripcion": "PAPEL ENVOLVER polvorones", "cantidad_bruta": 1}
    ])
    assert raciones is None
    assert ingredientes == []

    nombre, raciones, ingredientes = normalizar_respuesta_ingredientes_ia({
        "nombre_receta": "Rollitos de pescado",
        "raciones_base": 4,
        "ingredientes": [
            {"nombre": "gambas", "cantidad": 400, "unidad": "gr"},
            {"nombre": "calamares", "cantidad": 200, "unidad": "gr."},
            {"nombre": "merluza", "cantidad": 200, "unidad": "gramos"},
            {"nombre": "champiñones", "cantidad": 200, "unidad": "g"},
            {"nombre": "ajo", "cantidad": 3, "unidad": "dientes"},
            {"nombre": "fideos chinos de arroz", "cantidad": 75, "unidad": "gr"},
        ],
    })
    assert nombre == "Rollitos de pescado"
    assert raciones == 4.0
    cantidades = [ing["cantidad_bruta"] / raciones for ing in ingredientes]
    unidades = [ing["unidad_medida"] for ing in ingredientes]
    assert cantidades == [0.1, 0.05, 0.05, 0.05, 0.75, 0.01875]
    assert unidades == ["kg", "kg", "kg", "kg", "ud", "kg"]


def comprobar_sugerencias(parece_material_no_alimentario, sugerir_ingredientes_similares):
    assert parece_material_no_alimentario("PAPEL ENVOLVER polvorones")
    inventario = pd.DataFrame([
        {"codigo": "TOM-01", "familia": "VERDURAS", "descripcion": "TOMATE PERA", "unidad_medida": "kg", "merma": 0, "precio_unidad": 2.5},
        {"codigo": "TOM-02", "familia": "VERDURAS", "descripcion": "TOMATE CHERRY", "unidad_medida": "kg", "merma": 0, "precio_unidad": 3.2},
        {"codigo": "PEP-01", "familia": "VERDURAS", "descripcion": "PEPINO", "unidad_medida": "kg", "merma": 0, "precio_unidad": 1.8},
        {"codigo": "MAT-01", "familia": "MATERIAL", "descripcion": "PAPEL ENVOLVER POLVORONES", "unidad_medida": "ud", "merma": 0, "precio_unidad": 0.1},
    ])
    sugerencias_tomate = sugerir_ingredientes_similares("tomate pera", inventario)
    sugerencias_pepino = sugerir_ingredientes_similares("pepino", inventario)
    assert len(sugerencias_tomate) >= 2
    assert any("TOMATE" in s["descripcion"] for s in sugerencias_tomate)
    assert any("PEPINO" in s["descripcion"] for s in sugerencias_pepino)
    assert not sugerir_ingredientes_similares("tomate", inventario)[0]["descripcion"].startswith("PAPEL")


def main():
    (
        ajustar_ingredientes_por_raciones,
        calcular_ajuste_raciones,
        normalizar_respuesta_ingredientes_ia,
        parece_material_no_alimentario,
        sugerir_ingredientes_similares,
        generar_excel,
    ) = cargar_funciones_excel()
    comprobar_ajuste_raciones(ajustar_ingredientes_por_raciones, calcular_ajuste_raciones)
    comprobar_normalizacion_ia(normalizar_respuesta_ingredientes_ia)
    comprobar_sugerencias(parece_material_no_alimentario, sugerir_ingredientes_similares)
    ingredientes = [
        {"codigo": "ACE-01", "descripcion": "Aceite", "cantidad_bruta": 2.0, "merma": 0.0, "precio_unidad": 1.69},
        {"codigo": "HAR-01", "descripcion": "Harina", "cantidad_bruta": 3.0, "merma": 5.0, "precio_unidad": 1.1128},
        {"codigo": "HUE-01", "descripcion": "Huevos", "cantidad_bruta": 12.0, "merma": 0.0, "precio_unidad": 0.25},
    ]
    workbook_bytes = generar_excel("Prueba", ingredientes, 10.0, 30.0, 10.0, 10.0, 25.0, 2.5)
    wb = openpyxl.load_workbook(workbook_bytes, data_only=False)
    ws = wb["Ficha Técnica"]

    expected = {
        "B2": 10.0,
        "E2": 25.0,
        "B3": 2.5,
        "F6": "=C6*(1-E6)",
        "H6": "=C6*G6",
        "F7": "=C7*(1-E7)",
        "H7": "=C7*G7",
        "H10": "=SUM(H6:H8)",
        "H11": "=H10*(10.0/100)",
        "H12": "=H10+H11",
        "H15": "=H12/(1-(30.0/100))",
        "H16": "=H15*(10.0/100)",
        "H17": "=H15+H16",
        "H18": "=IF($E$2>0,H17/$E$2,0)",
    }
    for cell, formula in expected.items():
        value = ws[cell].value
        if value != formula:
            raise AssertionError(f"{cell}: esperado {formula!r}, recibido {value!r}")

    if not wb.calculation.fullCalcOnLoad:
        raise AssertionError("El libro no fuerza recalculo al abrir.")

    print("Excel export formulas OK")


if __name__ == "__main__":
    main()
