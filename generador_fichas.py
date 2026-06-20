import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import io
import base64
import html
import json
import re
import unicodedata
from pathlib import Path
from difflib import SequenceMatcher
from openai import OpenAI
from supabase import create_client, Client
from calculos_jerarquia import (
    calcular_coste_factura,
    calcular_coste_menu,
    calcular_coste_por_racion,
    calcular_coste_presupuesto,
    calcular_coste_receta,
)

# Configuración de página obligatoria al inicio de Streamlit
st.set_page_config(
    page_title="Gestor de Fichas Técnicas Cloud",
    page_icon="👨‍🍳",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# =============================================================================
# 🔑 INICIALIZACIÓN DE CONEXIONES CLOUD (INVENTARIO Y CIA)
# =============================================================================
supabase_url = st.secrets.get("SUPABASE_URL", None)
supabase_key = st.secrets.get("SUPABASE_KEY", None)
api_key = st.secrets.get("OPENAI_API_KEY", None)

supabase_disponible = supabase_url is not None and supabase_key is not None

# Variable de cliente global para el inventario
supabase: Client = None
if supabase_disponible:
    try:
        supabase = create_client(supabase_url, supabase_key)
    except Exception as e:
        st.error(f"⚠️ Error al conectar con el inventario: {e}")
        supabase_disponible = False

BASE_DIR = Path(__file__).resolve().parent
LOGO_SAMIRARTE_PATH = BASE_DIR / "assets" / "logo_samirarte.png"
SESSION_FILE_PATH = BASE_DIR / ".supabase_session.json"


def logo_samirarte_existe():
    return LOGO_SAMIRARTE_PATH.exists()


def logo_samirarte_base64():
    if not logo_samirarte_existe():
        return ""
    try:
        return base64.b64encode(LOGO_SAMIRARTE_PATH.read_bytes()).decode("ascii")
    except Exception:
        return ""


def _obtener_campo_auth(objeto, campo, valor_por_defecto=None):
    if objeto is None:
        return valor_por_defecto
    if isinstance(objeto, dict):
        return objeto.get(campo, valor_por_defecto)
    return getattr(objeto, campo, valor_por_defecto)


def login_supabase(email, password):
    """
    Inicia sesion y guarda los datos basicos en session_state.
    """
    if not supabase_disponible or supabase is None:
        return False, "El inventario no está conectado correctamente."

    email_limpio = str(email or "").strip()
    if not email_limpio or not password:
        return False, "Introduce email y contraseña."

    try:
        respuesta = supabase.auth.sign_in_with_password({
            "email": email_limpio,
            "password": password,
        })
        usuario = _obtener_campo_auth(respuesta, "user")
        sesion = _obtener_campo_auth(respuesta, "session")
        user_id = _obtener_campo_auth(usuario, "id")
        user_email = _obtener_campo_auth(usuario, "email", email_limpio)
        access_token = _obtener_campo_auth(sesion, "access_token")
        refresh_token = _obtener_campo_auth(sesion, "refresh_token")

        if not user_id:
            return False, "No se pudo obtener el usuario autenticado."

        _guardar_sesion_supabase(usuario, sesion)
        asegurar_usuario_app_supabase()
        return True, "Sesión iniciada correctamente."
    except Exception as e:
        return False, f"No se pudo iniciar sesión: {e}"


def registrar_usuario_supabase(email, password):
    """
    Crea un usuario y guarda la sesion si el inventario la devuelve.
    """
    if not supabase_disponible or supabase is None:
        return False, "El inventario no está conectado correctamente."

    email_limpio = str(email or "").strip()
    password_limpio = str(password or "")
    if not email_limpio:
        return False, "Introduce un email para crear la cuenta."
    if len(password_limpio) < 6:
        return False, "La contraseña debe tener al menos 6 caracteres."

    try:
        respuesta = supabase.auth.sign_up({
            "email": email_limpio,
            "password": password_limpio,
        })
        usuario = _obtener_campo_auth(respuesta, "user")
        sesion = _obtener_campo_auth(respuesta, "session")
        identidades = _obtener_campo_auth(usuario, "identities")

        if identidades == []:
            return False, "Este email ya tiene una cuenta. Inicia sesión."

        user_id = _obtener_campo_auth(usuario, "id")
        user_email = _obtener_campo_auth(usuario, "email", email_limpio)

        if sesion and user_id:
            _guardar_sesion_supabase(usuario, sesion)
            asegurar_usuario_app_supabase()
            return True, "Cuenta creada e iniciada correctamente"

        return True, "Cuenta creada. Revisa tu correo para confirmar el registro antes de iniciar sesión."
    except Exception as e:
        mensaje_error = str(e)
        mensaje_error_lower = mensaje_error.lower()
        if (
            "already registered" in mensaje_error_lower
            or "already exists" in mensaje_error_lower
            or "user already" in mensaje_error_lower
            or "email exists" in mensaje_error_lower
            or "already been registered" in mensaje_error_lower
            or "user_already_exists" in mensaje_error_lower
        ):
            return False, "Este email ya tiene una cuenta. Inicia sesión."
        return False, f"No se pudo crear la cuenta: {mensaje_error}"


def cambiar_password_usuario(nueva_password):
    """
    Actualiza la contraseña del usuario conectado mediante Inventario/Auth.
    """
    if not supabase_disponible or supabase is None:
        return False, "El inventario no está conectado correctamente."

    if not obtener_usuario_actual():
        return False, "La sesión ha caducado. Vuelve a iniciar sesión para cambiar la contraseña."

    nueva_password_limpia = str(nueva_password or "")
    if not nueva_password_limpia:
        return False, "Introduce una nueva contraseña."
    if len(nueva_password_limpia) < 6:
        return False, "La contraseña debe tener al menos 6 caracteres."

    try:
        supabase.auth.update_user({"password": nueva_password_limpia})
        return True, "Contraseña actualizada correctamente."
    except Exception as e:
        mensaje_error = str(e)
        mensaje_error_lower = mensaje_error.lower()
        if (
            "session" in mensaje_error_lower
            or "jwt" in mensaje_error_lower
            or "expired" in mensaje_error_lower
            or "not authenticated" in mensaje_error_lower
            or "auth session missing" in mensaje_error_lower
        ):
            return False, "La sesión ha caducado. Vuelve a iniciar sesión para cambiar la contraseña."
        return False, f"No se pudo actualizar la contraseña: {mensaje_error}"


def logout_supabase():
    """
    Cierra la sesion y limpia los datos de autenticacion locales.
    """
    if supabase_disponible and supabase is not None:
        try:
            supabase.auth.sign_out()
        except Exception:
            pass

    _limpiar_sesion_autenticacion()
    _borrar_sesion_local()
    return True, "Sesión cerrada correctamente."


def obtener_usuario_actual():
    """
    Devuelve el usuario autenticado guardado en session_state, si existe.
    """
    user_id = st.session_state.get("user_id")
    user_email = st.session_state.get("user_email")
    if not user_id:
        return None
    return {
        "id": user_id,
        "email": user_email,
        "access_token": st.session_state.get("access_token"),
    }


def _leer_sesion_local():
    if not SESSION_FILE_PATH.exists():
        return None
    try:
        contenido = SESSION_FILE_PATH.read_text(encoding="utf-8")
        datos = json.loads(contenido)
        if not isinstance(datos, dict):
            return None
        return datos
    except Exception:
        return None


def _guardar_sesion_local(datos):
    try:
        SESSION_FILE_PATH.write_text(json.dumps(datos), encoding="utf-8")
    except Exception:
        pass


def _borrar_sesion_local():
    try:
        if SESSION_FILE_PATH.exists():
            SESSION_FILE_PATH.unlink()
    except Exception:
        pass


def _limpiar_sesion_autenticacion():
    st.session_state["user_id"] = None
    st.session_state["user_email"] = None
    st.session_state["access_token"] = None
    st.session_state["refresh_token"] = None
    st.session_state["usuario_app"] = None
    st.session_state["usuario_app_error"] = None
    st.session_state["usuario_app_user_id"] = None


def _cargar_sesion_local_en_session_state():
    datos = _leer_sesion_local()
    if not datos or not isinstance(datos, dict):
        return False
    user_id = datos.get("user_id")
    if not user_id:
        return False
    st.session_state["user_id"] = str(user_id)
    st.session_state["user_email"] = str(datos.get("user_email", ""))
    st.session_state["access_token"] = datos.get("access_token")
    st.session_state["refresh_token"] = datos.get("refresh_token")
    return True


def _guardar_sesion_supabase(usuario, sesion):
    if usuario is None or sesion is None:
        return
    user_id = _obtener_campo_auth(usuario, "id")
    user_email = _obtener_campo_auth(usuario, "email", "")
    access_token = _obtener_campo_auth(sesion, "access_token")
    refresh_token = _obtener_campo_auth(sesion, "refresh_token")
    if user_id:
        st.session_state["user_id"] = str(user_id)
        st.session_state["user_email"] = str(user_email or "")
        st.session_state["access_token"] = access_token
        st.session_state["refresh_token"] = refresh_token
        _guardar_sesion_local({
            "user_id": str(user_id),
            "user_email": str(user_email or ""),
            "access_token": access_token,
            "refresh_token": refresh_token,
        })


def restaurar_sesion_supabase():
    """
    Intenta restaurar la sesión activa desde los tokens guardados.
    """
    if obtener_usuario_actual():
        return True
    if not supabase_disponible or supabase is None:
        return False

    if not _cargar_sesion_local_en_session_state():
        return False

    refresh_token = st.session_state.get("refresh_token")
    if not refresh_token:
        return False

    try:
        respuesta = supabase.auth.refresh_session(refresh_token=refresh_token)
        usuario = _obtener_campo_auth(respuesta, "user")
        sesion = _obtener_campo_auth(respuesta, "session")
        if not usuario or not sesion:
            return False
        _guardar_sesion_supabase(usuario, sesion)
        asegurar_usuario_app_supabase()
        return True
    except Exception:
        return False


def obtener_user_id_actual():
    user_id = st.session_state.get("user_id")
    return str(user_id).strip() if user_id else None


def normalizar_rol_usuario(rol):
    rol_limpio = str(rol or "").strip().lower()
    return rol_limpio if rol_limpio in {"usuario", "admin"} else "usuario"


def cargar_usuario_app_supabase(user_id=None):
    """
    Lee el perfil interno de la app para resolver permisos de administracion.
    """
    user_id_limpio = str(user_id or obtener_user_id_actual() or "").strip()
    if not user_id_limpio:
        return False, "No hay usuario autenticado.", None
    if not supabase_disponible or supabase is None:
        return False, "El inventario no está conectado correctamente.", None

    try:
        respuesta = (
            supabase
            .table("usuarios_app")
            .select("user_id,email,nombre,rol,activo,created_at,updated_at")
            .eq("user_id", user_id_limpio)
            .limit(1)
            .execute()
        )
        filas = respuesta.data or []
        if not filas:
            return False, "El usuario no tiene perfil interno en usuarios_app.", None
        perfil = dict(filas[0])
        perfil["rol"] = normalizar_rol_usuario(perfil.get("rol"))
        perfil["activo"] = bool(perfil.get("activo", True))
        return True, "Perfil cargado.", perfil
    except Exception as e:
        return False, f"No se pudo leer usuarios_app: {e}", None


def asegurar_usuario_app_supabase():
    """
    Crea el perfil interno del usuario conectado si todavia no existe.
    """
    usuario = obtener_usuario_actual()
    if not usuario:
        return False, "No hay usuario autenticado.", None
    if not supabase_disponible or supabase is None:
        return False, "El inventario no está conectado correctamente.", None

    ok, mensaje, perfil = cargar_usuario_app_supabase(usuario.get("id"))
    if ok and perfil:
        st.session_state["usuario_app"] = perfil
        st.session_state["usuario_app_error"] = None
        st.session_state["usuario_app_user_id"] = usuario.get("id")
        return True, mensaje, perfil

    mensaje_lower = str(mensaje).lower()
    tabla_no_existe = "does not exist" in mensaje_lower or "schema cache" in mensaje_lower
    if not tabla_no_existe:
        try:
            datos_usuario = {
                "user_id": usuario.get("id"),
                "email": usuario.get("email") or "",
                "rol": "usuario",
                "activo": True,
            }
            supabase.table("usuarios_app").insert(datos_usuario).execute()
            ok_creado, mensaje_creado, perfil_creado = cargar_usuario_app_supabase(usuario.get("id"))
            if ok_creado and perfil_creado:
                st.session_state["usuario_app"] = perfil_creado
                st.session_state["usuario_app_error"] = None
                st.session_state["usuario_app_user_id"] = usuario.get("id")
                return True, "Perfil interno creado.", perfil_creado
        except Exception as e:
            mensaje = f"No se pudo crear el perfil interno: {e}"

    st.session_state["usuario_app"] = None
    st.session_state["usuario_app_error"] = mensaje
    st.session_state["usuario_app_user_id"] = usuario.get("id")
    return False, mensaje, None


def obtener_perfil_usuario_app():
    usuario = obtener_usuario_actual()
    if not usuario:
        return None
    if st.session_state.get("usuario_app_user_id") != usuario.get("id"):
        asegurar_usuario_app_supabase()
    return st.session_state.get("usuario_app")


def usuario_actual_es_admin():
    perfil = obtener_perfil_usuario_app()
    if not perfil:
        return False
    return bool(perfil.get("activo", True)) and normalizar_rol_usuario(perfil.get("rol")) == "admin"


def cargar_usuarios_app_supabase():
    if not supabase_disponible or supabase is None:
        return False, "El inventario no está conectado correctamente.", pd.DataFrame()
    if not usuario_actual_es_admin():
        return False, "Solo los administradores pueden consultar usuarios.", pd.DataFrame()

    columnas = ["user_id", "email", "nombre", "rol", "activo", "created_at", "updated_at"]
    try:
        respuesta = (
            supabase
            .table("usuarios_app")
            .select(",".join(columnas))
            .order("email")
            .execute()
        )
        df = pd.DataFrame(respuesta.data or [])
        if df.empty:
            return True, "No hay usuarios configurados.", pd.DataFrame(columns=columnas)
        for col in columnas:
            if col not in df.columns:
                df[col] = ""
        df["rol"] = df["rol"].apply(normalizar_rol_usuario)
        df["activo"] = df["activo"].fillna(True).astype(bool)
        return True, "Usuarios cargados.", df[columnas].copy()
    except Exception as e:
        return False, f"No se pudo cargar usuarios_app: {e}", pd.DataFrame(columns=columnas)


def valor_user_id_receta(receta):
    if receta is None:
        return None
    user_id = receta.get("user_id") if isinstance(receta, dict) else getattr(receta, "user_id", None)
    if user_id is None:
        return None
    try:
        if pd.isna(user_id):
            return None
    except TypeError:
        pass
    user_id = str(user_id).strip()
    return user_id or None


def receta_es_propia_o_antigua(receta):
    user_id_actual = obtener_user_id_actual()
    receta_user_id = valor_user_id_receta(receta)
    if user_id_actual:
        return receta_user_id in (user_id_actual, None)
    return receta_user_id is None


def receta_es_modificable(receta):
    user_id_actual = obtener_user_id_actual()
    receta_user_id = valor_user_id_receta(receta)
    if user_id_actual:
        return receta_user_id == user_id_actual
    return receta_user_id is None


def mensaje_receta_no_modificable():
    return "No puedes modificar una receta de otro usuario"


def valor_user_id_menu(menu):
    if menu is None:
        return None
    user_id = menu.get("user_id") if isinstance(menu, dict) else getattr(menu, "user_id", None)
    if user_id is None:
        return None
    try:
        if pd.isna(user_id):
            return None
    except TypeError:
        pass
    user_id = str(user_id).strip()
    return user_id or None


def menu_es_propio_o_antiguo(menu):
    user_id_actual = obtener_user_id_actual()
    menu_user_id = valor_user_id_menu(menu)
    if user_id_actual:
        return menu_user_id in (user_id_actual, None)
    return menu_user_id is None


def menu_es_modificable(menu):
    user_id_actual = obtener_user_id_actual()
    menu_user_id = valor_user_id_menu(menu)
    if user_id_actual:
        return menu_user_id == user_id_actual
    return menu_user_id is None


def mensaje_menu_no_modificable(menu):
    if obtener_user_id_actual() and valor_user_id_menu(menu) is None:
        return "Este menú antiguo no tiene propietario. Duplica el menú para guardarlo en tu cuenta."
    return "No puedes modificar un menú de otro usuario"


# Inicializar estado de sesión para el escandallo activo
if 'ingredientes' not in st.session_state:
    st.session_state['ingredientes'] = []

if 'raciones_base' not in st.session_state:
    st.session_state['raciones_base'] = 1.0

if 'raciones_deseadas' not in st.session_state:
    st.session_state['raciones_deseadas'] = 1.0

if 'factor_raciones' not in st.session_state:
    st.session_state['factor_raciones'] = 1.0

if 'ingredientes_base_raciones' not in st.session_state:
    st.session_state['ingredientes_base_raciones'] = st.session_state.get('ingredientes_originales_raciones', None)

if 'receta_nombre' not in st.session_state:
    st.session_state['receta_nombre'] = st.session_state.get('nombre_plato', "Mi Receta")

if 'receta_raciones_base' not in st.session_state:
    st.session_state['receta_raciones_base'] = float(st.session_state['raciones_base'])

if 'receta_raciones_deseadas' not in st.session_state:
    st.session_state['receta_raciones_deseadas'] = float(st.session_state['raciones_deseadas'])

if 'input_nombre_plato' not in st.session_state:
    st.session_state['input_nombre_plato'] = st.session_state['receta_nombre']

if 'input_raciones_base' not in st.session_state:
    st.session_state['input_raciones_base'] = float(st.session_state['receta_raciones_base'])

if 'input_raciones_deseadas' not in st.session_state:
    st.session_state['input_raciones_deseadas'] = float(st.session_state['receta_raciones_deseadas'])

if 'sincronizar_inputs_raciones' not in st.session_state:
    st.session_state['sincronizar_inputs_raciones'] = False

if 'sincronizar_campos_receta' not in st.session_state:
    st.session_state['sincronizar_campos_receta'] = False

if 'raciones_base_aplicadas' not in st.session_state:
    st.session_state['raciones_base_aplicadas'] = float(st.session_state['raciones_base'])

if 'raciones_deseadas_aplicadas' not in st.session_state:
    st.session_state['raciones_deseadas_aplicadas'] = float(st.session_state['raciones_deseadas'])

# Trigger para invalidar cache del inventario al editar el catalogo
if 'db_trigger' not in st.session_state:
    st.session_state['db_trigger'] = 0

if 'user_id' not in st.session_state:
    st.session_state['user_id'] = None
if 'user_email' not in st.session_state:
    st.session_state['user_email'] = None
if 'access_token' not in st.session_state:
    st.session_state['access_token'] = None
if 'refresh_token' not in st.session_state:
    st.session_state['refresh_token'] = None
if 'usuario_app' not in st.session_state:
    st.session_state['usuario_app'] = None
if 'usuario_app_error' not in st.session_state:
    st.session_state['usuario_app_error'] = None
if 'usuario_app_user_id' not in st.session_state:
    st.session_state['usuario_app_user_id'] = None

# Restaurar sesión guardada antes de renderizar la UI.
restaurar_sesion_supabase()

if 'nombre_plato' not in st.session_state:
    st.session_state['nombre_plato'] = st.session_state['receta_nombre']

if 'receta_categoria' not in st.session_state:
    st.session_state['receta_categoria'] = ""

if 'input_receta_categoria' not in st.session_state:
    st.session_state['input_receta_categoria'] = st.session_state['receta_categoria']

if 'receta_tipo_plato' not in st.session_state:
    st.session_state['receta_tipo_plato'] = ""

if 'input_receta_tipo_plato' not in st.session_state:
    st.session_state['input_receta_tipo_plato'] = st.session_state['receta_tipo_plato']

if 'receta_observaciones' not in st.session_state:
    st.session_state['receta_observaciones'] = ""

if 'input_receta_observaciones' not in st.session_state:
    st.session_state['input_receta_observaciones'] = st.session_state['receta_observaciones']

if 'receta_id_cargada' not in st.session_state:
    st.session_state['receta_id_cargada'] = None

if 'codigo_receta_cargada' not in st.session_state:
    st.session_state['codigo_receta_cargada'] = ""

if 'costes_indirectos_pct' not in st.session_state:
    st.session_state['costes_indirectos_pct'] = 10.0

if 'margen_beneficio_pct' not in st.session_state:
    st.session_state['margen_beneficio_pct'] = 30.0

if 'iva_pct' not in st.session_state:
    st.session_state['iva_pct'] = 10.0

if 'menu_actual' not in st.session_state:
    st.session_state['menu_actual'] = []
if 'menu_id' not in st.session_state:
    st.session_state['menu_id'] = None
if 'sincronizar_campos_menu' not in st.session_state:
    st.session_state['sincronizar_campos_menu'] = False
if 'factura_lineas_actual' not in st.session_state:
    st.session_state['factura_lineas_actual'] = []
if 'factura_id_en_edicion' not in st.session_state:
    st.session_state['factura_id_en_edicion'] = None
if 'factura_presupuesto_origen_id' not in st.session_state:
    st.session_state['factura_presupuesto_origen_id'] = None

# =============================================================================
# 📥 FUNCIÓN CACHÉ PARA CARGAR INVENTARIO
# =============================================================================
@st.cache_data(ttl=600)
def cargar_inventario_supabase(trigger):
    """
    Trae el listado de ingredientes del inventario de forma segura.
    Se invalida automáticamente al cambiar el 'trigger'.
    """
    cols_deseadas = [
        "codigo", "familia", "descripcion", "unidad_medida", "merma", "precio_unidad",
        "proveedor_precio", "formato_compra", "cantidad_formato_compra",
        "unidad_formato_compra", "precio_formato_compra", "fecha_precio",
        "url_precio", "observaciones_precio"
    ]
    if not supabase_disponible or supabase is None:
        return pd.DataFrame(columns=cols_deseadas)
    try:
        respuesta = (
            supabase
            .table("inventario")
            .select(",".join(cols_deseadas))
            .execute()
        )
        df = pd.DataFrame(respuesta.data)

        if df.empty:
            return pd.DataFrame(columns=cols_deseadas)

        df = df.loc[:, ~df.columns.duplicated()].copy()
        for col in cols_deseadas:
            if col not in df.columns:
                df[col] = 0.0 if col in ["merma", "precio_unidad", "cantidad_formato_compra", "precio_formato_compra"] else ""

        df = df[cols_deseadas].copy()
        df["codigo"] = df["codigo"].fillna("").astype(str).str.strip().str.upper()
        df["familia"] = df["familia"].fillna("").astype(str).str.strip()
        df["descripcion"] = df["descripcion"].fillna("").astype(str).str.strip()
        df["unidad_medida"] = df["unidad_medida"].fillna("kg").astype(str).str.strip().replace("", "kg")
        df["merma"] = pd.to_numeric(df["merma"], errors="coerce").fillna(0.0)
        df["precio_unidad"] = pd.to_numeric(df["precio_unidad"], errors="coerce").fillna(0.0)
        df["cantidad_formato_compra"] = pd.to_numeric(df["cantidad_formato_compra"], errors="coerce")
        df["precio_formato_compra"] = pd.to_numeric(df["precio_formato_compra"], errors="coerce")
        df = df.sort_values(by="descripcion")
        return df
    except Exception as e:
        st.error(f"Error al leer el inventario: {e}")
        return pd.DataFrame(columns=cols_deseadas)

# Cargar inventario actual para autocompletado y contexto IA
inventario_df = cargar_inventario_supabase(st.session_state['db_trigger'])

# Convertir inventario a diccionario seguro para búsquedas rápidas en milisegundos sin KeyError
inventario_dict = {}
if not inventario_df.empty:
    for _, row in inventario_df.iterrows():
        codigo_raw = row.get("codigo", None)
        if codigo_raw is not None and pd.notna(codigo_raw):
            codigo_str = str(codigo_raw).strip().upper()
            inventario_dict[codigo_str] = {
                "familia": row.get("familia", "VARIOS"),
                "descripcion": row.get("descripcion", ""),
                "unidad_medida": row.get("unidad_medida", "kg"),
                "merma": float(row.get("merma", 0.0)) if pd.notna(row.get("merma")) else 0.0,
                "precio_unidad": float(row.get("precio_unidad", 0.0)) if pd.notna(row.get("precio_unidad")) else 0.0,
                "proveedor_precio": row.get("proveedor_precio", ""),
                "formato_compra": row.get("formato_compra", ""),
                "cantidad_formato_compra": row.get("cantidad_formato_compra", None),
                "unidad_formato_compra": row.get("unidad_formato_compra", ""),
                "precio_formato_compra": row.get("precio_formato_compra", None),
                "fecha_precio": row.get("fecha_precio", ""),
                "url_precio": row.get("url_precio", ""),
                "observaciones_precio": row.get("observaciones_precio", "")
            }


def normalizar_texto_busqueda(texto):
    texto = "" if texto is None else str(texto)
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = re.sub(r"[^A-Za-z0-9 ]+", " ", texto.upper())
    return re.sub(r"\s+", " ", texto).strip()


PALABRAS_NO_ALIMENTARIAS = {
    "PAPEL", "ENVOLVER", "ENVOLTORIO", "CAJA", "BOLSA", "BANDEJA", "SERVILLETA",
    "VASO", "PLATO", "CUBIERTO", "GUANTE", "FILM", "ALUMINIO", "ETIQUETA",
    "PACKAGING", "ENVASE", "CUCHARA", "TENEDOR", "CUCHILLO", "MOLDE", "BLONDA", "TAPA"
}

PALABRAS_BLOQUEO_MATCHING = {
    "CARCASA", "HUESO", "HUESOS", "PIEL", "RECORTE", "RECORTES",
    "PAPEL", "ENVOLVER", "ENVASE", "BANDEJA", "CAJA", "ETIQUETA", "PACKAGING"
}

FRASES_BLOQUEO_MATCHING = {
    "SAL ESPECIAL HORNEAR"
}

INGREDIENTES_INCOHERENTES_ARROZ = {
    "MANGO", "VINAGRE", "ESPINACA", "ESPINACAS", "SETA", "SETAS",
    "ZANAHORIA", "SECRETO", "CARCASA", "HABAS"
}


def parece_material_no_alimentario(descripcion):
    palabras = set(normalizar_texto_busqueda(descripcion).split())
    return bool(palabras & PALABRAS_NO_ALIMENTARIAS)


def contiene_bloqueo_matching(descripcion):
    texto = normalizar_texto_busqueda(descripcion)
    palabras = set(texto.split())
    if palabras & PALABRAS_BLOQUEO_MATCHING:
        return True
    return any(frase in texto for frase in FRASES_BLOQUEO_MATCHING)


def bloqueo_pedido_explicitamente(descripcion_original, descripcion_inventario):
    original_norm = normalizar_texto_busqueda(descripcion_original)
    inventario_norm = normalizar_texto_busqueda(descripcion_inventario)
    palabras_bloqueadas = [
        palabra for palabra in PALABRAS_BLOQUEO_MATCHING
        if palabra in inventario_norm.split()
    ]
    frases_bloqueadas = [
        frase for frase in FRASES_BLOQUEO_MATCHING
        if frase in inventario_norm
    ]
    return all(palabra in original_norm.split() for palabra in palabras_bloqueadas) and all(
        frase in original_norm for frase in frases_bloqueadas
    )


def descripcion_generica_ingrediente(descripcion):
    texto = normalizar_texto_busqueda(descripcion)
    palabras_descartables = {
        "FRESCO", "FRESCA", "CONGELADO", "CONGELADA", "ECO", "ECOLOGICO", "ECOLOGICA",
        "BIO", "MARCA", "BOLSA", "LATA", "CAJA", "BANDEJA", "PAQUETE", "GARRAFA",
        "BOTELLA", "SACO", "MONODOSIS", "APROX", "SIN", "CON", "NATURAL", "PACK", "UNIDAD",
        "UD", "UDS", "GR", "G", "KG", "K", "ML", "CL", "L", "LITRO", "LITROS", "ESPECIAL",
        "HORNEAR", "FINDUS", "HS"
    }
    palabras = [p for p in texto.split() if not p.isdigit() and p not in palabras_descartables]
    return " ".join(palabras[:4]) or texto


def sugerir_ingredientes_similares(descripcion, inventario, limite=8, umbral=0.30):
    objetivo = descripcion_generica_ingrediente(descripcion)
    if not objetivo or inventario.empty or parece_material_no_alimentario(objetivo):
        return []

    sugerencias = []
    for _, row in inventario.iterrows():
        desc_bd = str(row.get("descripcion", ""))
        if parece_material_no_alimentario(desc_bd) and not parece_material_no_alimentario(objetivo):
            continue
        if contiene_bloqueo_matching(desc_bd) and not bloqueo_pedido_explicitamente(descripcion, desc_bd):
            continue
        desc_norm = normalizar_texto_busqueda(desc_bd)
        if not desc_norm:
            continue
        score = SequenceMatcher(None, objetivo, desc_norm).ratio()
        palabras_objetivo = set(objetivo.split())
        palabras_bd = set(desc_norm.split())
        if palabras_objetivo and palabras_bd:
            score = max(score, len(palabras_objetivo & palabras_bd) / len(palabras_objetivo | palabras_bd))
            palabra_principal = objetivo.split()[0] if objetivo.split() else ""
            if palabra_principal and palabra_principal in palabras_bd:
                score += 0.35
        if score >= umbral:
            sugerencias.append({
                "codigo": str(row.get("codigo", "")).strip().upper(),
                "familia": str(row.get("familia", "")).strip(),
                "descripcion": desc_bd,
                "unidad_medida": str(row.get("unidad_medida", "kg")).strip() or "kg",
                "merma": float(row.get("merma", 0.0)) if pd.notna(row.get("merma")) else 0.0,
                "precio_unidad": float(row.get("precio_unidad", 0.0)) if pd.notna(row.get("precio_unidad")) else 0.0,
                "proveedor_precio": str(row.get("proveedor_precio", "") or "").strip(),
                "formato_compra": str(row.get("formato_compra", "") or "").strip(),
                "cantidad_formato_compra": row.get("cantidad_formato_compra", None),
                "unidad_formato_compra": str(row.get("unidad_formato_compra", "") or "").strip(),
                "score": score
            })

    return sorted(sugerencias, key=lambda item: item["score"], reverse=True)[:limite]


def codigo_ingrediente_valido(codigo):
    codigo = "" if codigo is None else str(codigo).strip().upper()
    return codigo and codigo not in ["S/C", "SC", "SIN CODIGO", "SIN CÓDIGO"]


def inferir_unidad_medida(descripcion):
    texto = normalizar_texto_busqueda(descripcion)
    if any(p in texto.split() for p in ["AGUA", "CALDO", "VINO", "ZUMO", "LECHE", "NATA", "ACEITE", "VINAGRE"]):
        return "l"
    if any(p in texto.split() for p in ["HUEVO", "HUEVOS"]):
        return "ud"
    if "SOBRE" in texto or "SOBRES" in texto:
        return "sobre"
    return "kg"


def receta_parece_arroz(nombre_receta):
    texto = normalizar_texto_busqueda(nombre_receta)
    return bool({"PAELLA", "ARROZ", "RISOTTO"} & set(texto.split()))


def ingrediente_incoherente_en_arroz(nombre_ingrediente):
    palabras = set(normalizar_texto_busqueda(nombre_ingrediente).split())
    return bool(palabras & INGREDIENTES_INCOHERENTES_ARROZ)


def motivo_indica_presencia_explicita(motivo):
    texto = normalizar_texto_busqueda(motivo)
    return any(palabra in texto for palabra in ["APARECE", "EXPLICITO", "EXPLICITA", "TEXTO", "IMAGEN", "RECETA"])


def normalizar_unidad_ia(unidad, descripcion):
    unidad_norm = normalizar_texto_busqueda(unidad).lower()
    equivalencias = {
        "gr": "kg",
        "grs": "kg",
        "gramos": "kg",
        "gramo": "kg",
        "g": "kg",
        "kgs": "kg",
        "kilogramo": "kg",
        "kilogramos": "kg",
        "litro": "l",
        "litros": "l",
        "ml": "l",
        "mililitro": "l",
        "mililitros": "l",
        "unidad": "ud",
        "unidades": "ud",
        "pieza": "ud",
        "piezas": "ud",
        "diente": "ud",
        "dientes": "ud",
        "cucharada": "ud",
        "cucharadas": "ud",
        "cucharadita": "ud",
        "cucharaditas": "ud",
        "gota": "ud",
        "gotas": "ud",
        "hojas": "hoja",
    }
    return equivalencias.get(unidad_norm, unidad_norm or inferir_unidad_medida(descripcion))


def convertir_cantidad_ia(cantidad, unidad):
    cantidad_num = pd.to_numeric(cantidad, errors="coerce")
    if pd.isna(cantidad_num):
        return 0.0
    cantidad_num = float(cantidad_num)
    unidad_norm = normalizar_texto_busqueda(unidad).lower()
    if unidad_norm in ["g", "gr", "grs", "gramo", "gramos"]:
        return cantidad_num / 1000
    if unidad_norm in ["ml", "mililitro", "mililitros"]:
        return cantidad_num / 1000
    return cantidad_num


def buscar_coincidencia_principal_inventario(nombre_original, confianza):
    sugerencias = sugerir_ingredientes_similares(nombre_original, inventario_df, limite=6, umbral=0.35)
    if not sugerencias:
        return None, [], "sin_coincidencia"

    principal = sugerencias[0]
    segunda = sugerencias[1] if len(sugerencias) > 1 else None
    score_principal = float(principal.get("score", 0.0))
    score_segunda = float(segunda.get("score", 0.0)) if segunda else 0.0
    confianza_norm = normalizar_texto_busqueda(confianza)
    es_clara = score_principal >= 0.72 and (not segunda or score_principal - score_segunda >= 0.12)
    if confianza_norm == "BAJA":
        es_clara = False

    return principal, sugerencias, "clara" if es_clara else "dudosa"


def generar_codigo_ingrediente_nuevo(descripcion, existentes):
    base = normalizar_texto_busqueda(descripcion)
    prefijo = "".join(palabra[:3] for palabra in base.split()[:2])[:6] or "ING"
    prefijo = re.sub(r"[^A-Z0-9]", "", prefijo) or "ING"
    candidato = f"ING-{prefijo}"
    contador = 1
    while candidato in existentes:
        contador += 1
        candidato = f"ING-{prefijo}-{contador:02d}"
    existentes.add(candidato)
    return candidato


def preparar_fila_inventario_desde_ingrediente(ingrediente, codigo=None):
    codigo_final = codigo if codigo is not None else ingrediente.get("codigo", "")
    merma = pd.to_numeric(ingrediente.get("merma", 0.0), errors="coerce")
    precio = pd.to_numeric(ingrediente.get("precio_unidad", 0.0), errors="coerce")
    return {
        "codigo": str(codigo_final).strip().upper(),
        "familia": ingrediente.get("familia", "SIN CLASIFICAR"),
        "descripcion": str(ingrediente.get("descripcion", "")).strip() or "Ingrediente Nuevo",
        "unidad_medida": str(ingrediente.get("unidad_medida", "")).strip() or inferir_unidad_medida(ingrediente.get("descripcion", "")),
        "merma": 0.0 if pd.isna(merma) else float(merma),
        "precio_unidad": 0.0 if pd.isna(precio) else float(precio)
    }


RECETAS_COLUMNAS = [
    "id", "user_id", "codigo_receta", "nombre", "categoria", "tipo_plato",
    "raciones_base", "unidad_servicio", "descripcion", "elaboracion",
    "observaciones", "costes_indirectos_pct", "margen_beneficio_pct",
    "iva_pct", "coste_total", "precio_venta_sin_iva",
    "precio_venta_con_iva", "activa", "created_at", "updated_at"
]

RECETA_INGREDIENTES_COLUMNAS = [
    "id", "receta_id", "codigo_ingrediente", "descripcion_ingrediente",
    "cantidad_bruta", "unidad_medida", "merma", "cantidad_neta",
    "precio_unidad", "coste_total", "orden", "es_temporal",
    "created_at", "updated_at"
]

MENUS_COLUMNAS = [
    "id", "user_id", "nombre", "tipo_menu", "descripcion",
    "numero_comensales", "coste_total", "precio_total",
    "created_at", "updated_at"
]

MENU_RECETAS_COLUMNAS = [
    "id", "menu_id", "receta_id", "raciones", "orden", "seccion", "created_at"
]

CLIENTES_COLUMNAS = [
    "id", "user_id", "nombre", "tipo_cliente", "nif_cif", "email", "telefono",
    "direccion", "codigo_postal", "ciudad", "provincia", "pais",
    "observaciones", "created_at", "updated_at"
]

FACTURAS_COLUMNAS = [
    "id", "user_id", "cliente_id", "numero_factura", "tipo_documento", "estado",
    "fecha_emision", "fecha_vencimiento", "concepto", "base_imponible",
    "iva_pct", "iva_importe", "retencion_pct", "retencion_importe", "total",
    "metodo_pago", "estado_cobro", "notas", "created_at", "updated_at"
]

FACTURA_LINEAS_COLUMNAS = [
    "id", "factura_id", "origen_tipo", "origen_id", "descripcion", "cantidad",
    "unidad", "precio_unitario", "descuento_pct", "base_linea", "iva_pct",
    "iva_linea", "total_linea", "orden", "created_at"
]

FACTURAS_COLUMNAS_JERARQUIA = FACTURAS_COLUMNAS + [
    "presupuesto_id", "coste_total"
]

PRESUPUESTOS_COLUMNAS = [
    "id", "user_id", "cliente_id", "numero_presupuesto", "estado",
    "fecha_emision", "fecha_vencimiento", "concepto", "coste_total",
    "precio_total", "notas", "created_at", "updated_at"
]

PRESUPUESTO_MENUS_COLUMNAS = [
    "id", "presupuesto_id", "menu_id", "menu_nombre_snapshot",
    "cantidad_menu", "coste_total_menu", "coste_linea_menu_presupuesto",
    "precio_total_menu", "precio_linea_menu_presupuesto",
    "observaciones", "orden", "created_at"
]

FACTURA_MENUS_COLUMNAS = [
    "id", "factura_id", "menu_id", "menu_nombre_snapshot",
    "cantidad_menu", "coste_total_menu", "coste_linea_menu_factura",
    "precio_total_menu", "precio_linea_menu_factura",
    "observaciones", "orden", "created_at"
]

MENSAJE_CLIENTES_NO_ACTIVADO = (
    "El módulo de clientes todavía no está activado en Inventario. "
    "Ejecuta primero el SQL de facturación."
)

MENSAJE_FACTURACION_NO_ACTIVADA = (
    "El módulo de facturación todavía no está activado en Inventario. "
    "Ejecuta primero el SQL de facturación."
)


def numero_seguro(valor, defecto=0.0):
    numero = pd.to_numeric(valor, errors="coerce")
    return float(defecto) if pd.isna(numero) else float(numero)


def generar_codigo_receta():
    """
    Genera el siguiente codigo REC-0001, REC-0002... leyendo public.recetas.
    Si el inventario no esta disponible, devuelve un codigo inicial seguro.
    """
    if not supabase_disponible or supabase is None:
        return "REC-0001"

    try:
        respuesta = (
            supabase
            .table("recetas")
            .select("codigo_receta")
            .execute()
        )
        codigos_existentes = {
            str(fila.get("codigo_receta", "")).strip().upper()
            for fila in (respuesta.data or [])
            if fila.get("codigo_receta")
        }
        ultimo_numero = 0
        for codigo in codigos_existentes:
            coincidencia = re.fullmatch(r"REC-(\d+)", codigo)
            if coincidencia:
                ultimo_numero = max(ultimo_numero, int(coincidencia.group(1)))

        siguiente = ultimo_numero + 1
        candidato = f"REC-{siguiente:04d}"
        while candidato in codigos_existentes:
            siguiente += 1
            candidato = f"REC-{siguiente:04d}"
        return candidato
    except Exception:
        return "REC-0001"


def generar_nombre_copia_receta(nombre_original):
    """
    Genera un nombre de copia evitando duplicados sencillos.
    """
    nombre_base = str(nombre_original or "Receta").strip() or "Receta"
    if not supabase_disponible or supabase is None:
        return f"{nombre_base} copia"

    try:
        recetas_df = cargar_recetas_supabase(obtener_user_id_actual())
        nombres_existentes = {
            str(nombre).strip().lower()
            for nombre in recetas_df.get("nombre", [])
            if str(nombre).strip()
        }
        candidato = f"{nombre_base} copia"
        if candidato.lower() not in nombres_existentes:
            return candidato

        contador = 2
        while True:
            candidato = f"{nombre_base} copia {contador}"
            if candidato.lower() not in nombres_existentes:
                return candidato
            contador += 1
    except Exception:
        return f"{nombre_base} copia"


def generar_nombre_copia_menu(nombre_original):
    """
    Genera un nombre de copia de menu evitando duplicados sencillos.
    """
    nombre_base = str(nombre_original or "Menú").strip() or "Menú"
    if not supabase_disponible or supabase is None:
        return f"{nombre_base} copia"

    try:
        menus_df = cargar_menus_supabase(obtener_user_id_actual())
        nombres_existentes = {
            str(nombre).strip().lower()
            for nombre in menus_df.get("nombre", [])
            if str(nombre).strip()
        }
        candidato = f"{nombre_base} copia"
        if candidato.lower() not in nombres_existentes:
            return candidato

        contador = 2
        while True:
            candidato = f"{nombre_base} copia {contador}"
            if candidato.lower() not in nombres_existentes:
                return candidato
            contador += 1
    except Exception:
        return f"{nombre_base} copia"


def calcular_datos_ingrediente_receta(ing, orden=0):
    """
    Convierte un ingrediente de la app al formato de public.receta_ingredientes.
    """
    codigo = str(ing.get("codigo", "")).strip().upper()
    cantidad_bruta = numero_seguro(ing.get("cantidad_bruta", 0.0))
    merma = numero_seguro(ing.get("merma", 0.0))
    precio_unidad = numero_seguro(ing.get("precio_unidad", 0.0))
    cantidad_neta = cantidad_bruta * (1 - merma / 100)
    coste_total = cantidad_bruta * precio_unidad
    es_temporal = not codigo_ingrediente_valido(codigo)

    return {
        "codigo_ingrediente": None if es_temporal else codigo,
        "descripcion_ingrediente": str(
            ing.get("descripcion_ingrediente", ing.get("descripcion", ""))
        ).strip(),
        "cantidad_bruta": cantidad_bruta,
        "unidad_medida": str(ing.get("unidad_medida", "kg")).strip() or "kg",
        "merma": merma,
        "cantidad_neta": cantidad_neta,
        "precio_unidad": precio_unidad,
        "coste_total": coste_total,
        "orden": int(numero_seguro(orden, 0)),
        "es_temporal": es_temporal
    }


def validar_receta_para_guardar(ingredientes, raciones_base_receta):
    try:
        coste_ingredientes = calcular_coste_receta(ingredientes)
        calcular_coste_por_racion(coste_ingredientes, raciones_base_receta)
    except ValueError as exc:
        return False, str(exc)
    return True, ""


@st.cache_data(ttl=600)
def cargar_recetas_supabase(user_id_actual=None):
    """
    Lee public.recetas y devuelve un DataFrame seguro.
    """
    if not supabase_disponible or supabase is None:
        return pd.DataFrame(columns=RECETAS_COLUMNAS)

    try:
        respuesta = (
            supabase
            .table("recetas")
            .select(",".join(RECETAS_COLUMNAS))
            .order("created_at", desc=True)
            .execute()
        )
        df = pd.DataFrame(respuesta.data or [])
        if df.empty:
            return pd.DataFrame(columns=RECETAS_COLUMNAS)

        for col in RECETAS_COLUMNAS:
            if col not in df.columns:
                df[col] = None
        df = df[RECETAS_COLUMNAS].copy()
        user_id_actual = str(user_id_actual).strip() if user_id_actual else None
        user_ids = df["user_id"].apply(lambda valor: valor_user_id_receta({"user_id": valor}))
        if user_id_actual:
            return df[user_ids.isin([user_id_actual, None])].copy()
        return df[user_ids.isna()].copy()
    except Exception:
        return pd.DataFrame(columns=RECETAS_COLUMNAS)


def limpiar_cache_recetas_guardadas():
    """
    Invalida el listado cacheado de recetas guardadas si la funcion esta cacheada.
    """
    try:
        cargar_recetas_supabase.clear()
        return True
    except Exception:
        return False


def cargar_receta_detalle_supabase(receta_id):
    """
    Lee cabecera de public.recetas e ingredientes de public.receta_ingredientes.
    Devuelve una tupla: (cabecera_dict, ingredientes_df).
    """
    ingredientes_vacios = pd.DataFrame(columns=RECETA_INGREDIENTES_COLUMNAS)
    if not receta_id or not supabase_disponible or supabase is None:
        return {}, ingredientes_vacios

    try:
        respuesta_receta = (
            supabase
            .table("recetas")
            .select(",".join(RECETAS_COLUMNAS))
            .eq("id", receta_id)
            .limit(1)
            .execute()
        )
        cabecera = (respuesta_receta.data or [{}])[0] if respuesta_receta.data else {}
        if cabecera and not receta_es_propia_o_antigua(cabecera):
            return {}, ingredientes_vacios

        respuesta_ingredientes = (
            supabase
            .table("receta_ingredientes")
            .select(",".join(RECETA_INGREDIENTES_COLUMNAS))
            .eq("receta_id", receta_id)
            .order("orden")
            .execute()
        )
        ingredientes_df = pd.DataFrame(respuesta_ingredientes.data or [])
        if ingredientes_df.empty:
            return cabecera, ingredientes_vacios

        for col in RECETA_INGREDIENTES_COLUMNAS:
            if col not in ingredientes_df.columns:
                ingredientes_df[col] = None
        return cabecera, ingredientes_df[RECETA_INGREDIENTES_COLUMNAS].copy()
    except Exception:
        return {}, ingredientes_vacios


def guardar_receta_nueva_supabase(datos_receta, ingredientes):
    """
    Guarda una receta nueva y sus lineas de escandallo en el inventario.
    """
    if not supabase_disponible or supabase is None:
        return False, "El inventario no está conectado correctamente.", None
    datos_receta, ingredientes = preparar_receta_para_una_racion(
        datos_receta,
        ingredientes,
    )
    if datos_receta is None:
        return False, "La base de raciones debe ser mayor que 0.", None
    receta_valida, mensaje_validacion = validar_receta_para_guardar(ingredientes, 1.0)
    if not receta_valida:
        return False, mensaje_validacion, None

    try:
        user_id_actual = obtener_user_id_actual()
        datos_guardar = dict(datos_receta)
        datos_guardar["user_id"] = user_id_actual
        respuesta_receta = (
            supabase
            .table("recetas")
            .insert(datos_guardar)
            .execute()
        )
        receta_guardada = (respuesta_receta.data or [{}])[0]
        receta_id = receta_guardada.get("id")
        if not receta_id:
            return False, "No se pudo obtener el identificador de la receta guardada.", None

        lineas = []
        for orden, ing in enumerate(ingredientes, start=1):
            linea = calcular_datos_ingrediente_receta(ing, orden=orden)
            linea["receta_id"] = receta_id
            lineas.append(linea)

        if lineas:
            (
                supabase
                .table("receta_ingredientes")
                .insert(lineas)
                .execute()
            )

        limpiar_cache_recetas_guardadas()

        if user_id_actual:
            return True, "Receta guardada en tu cuenta", receta_guardada
        return True, "Receta guardada correctamente. Inicia sesión para guardar recetas en tu cuenta", receta_guardada
    except Exception as e:
        return False, f"Error al guardar la receta en el inventario: {e}", None


def actualizar_receta_supabase(receta_id, datos_receta, ingredientes):
    """
    Actualiza una receta existente y reemplaza solo sus lineas de escandallo.
    """
    if not receta_id:
        return False, "No hay una receta cargada para actualizar."
    if not supabase_disponible or supabase is None:
        return False, "El inventario no está conectado correctamente."
    datos_receta, ingredientes = preparar_receta_para_una_racion(
        datos_receta,
        ingredientes,
    )
    if datos_receta is None:
        return False, "La base de raciones debe ser mayor que 0."
    receta_valida, mensaje_validacion = validar_receta_para_guardar(ingredientes, 1.0)
    if not receta_valida:
        return False, mensaje_validacion

    try:
        cabecera_actual, _ = cargar_receta_detalle_supabase(receta_id)
        if not cabecera_actual:
            return False, "No se pudo comprobar la receta antes de actualizarla."
        if not receta_es_modificable(cabecera_actual):
            return False, mensaje_receta_no_modificable()

        datos_actualizar = dict(datos_receta)
        datos_actualizar["user_id"] = valor_user_id_receta(cabecera_actual)
        (
            supabase
            .table("recetas")
            .update(datos_actualizar)
            .eq("id", receta_id)
            .execute()
        )

        (
            supabase
            .table("receta_ingredientes")
            .delete()
            .eq("receta_id", receta_id)
            .execute()
        )

        lineas = []
        for orden, ing in enumerate(ingredientes, start=1):
            linea = calcular_datos_ingrediente_receta(ing, orden=orden)
            linea["receta_id"] = receta_id
            lineas.append(linea)

        if lineas:
            (
                supabase
                .table("receta_ingredientes")
                .insert(lineas)
                .execute()
            )

        limpiar_cache_recetas_guardadas()

        return True, "Receta actualizada correctamente."
    except Exception as e:
        return False, f"Error al actualizar la receta en el inventario: {e}"


def duplicar_receta_supabase(datos_receta, ingredientes):
    """
    Duplica una receta cargada como una receta nueva con codigo y nombre nuevos.
    """
    if not supabase_disponible or supabase is None:
        return False, "El inventario no está conectado correctamente.", None
    if not ingredientes:
        return False, "No hay ingredientes para duplicar.", None

    datos_copia = dict(datos_receta)
    datos_copia["user_id"] = obtener_user_id_actual()
    datos_copia["codigo_receta"] = generar_codigo_receta()
    datos_copia["nombre"] = generar_nombre_copia_receta(datos_receta.get("nombre", "Receta"))

    return guardar_receta_nueva_supabase(datos_copia, ingredientes)


def receta_esta_usada_en_menus(receta_id):
    """
    Comprueba si una receta aparece en alguna linea de menu sin modificar datos.
    """
    receta_id_limpio = str(receta_id or "").strip()
    if not receta_id_limpio or not supabase_disponible or supabase is None:
        return False

    try:
        respuesta = (
            supabase
            .table("menu_recetas")
            .select("id")
            .eq("receta_id", receta_id_limpio)
            .limit(1)
            .execute()
        )
        return bool(respuesta.data)
    except Exception:
        return False


def normalizar_ingredientes_a_una_racion(ingredientes, raciones_base):
    raciones_base_num = pd.to_numeric(raciones_base, errors="coerce")
    if pd.isna(raciones_base_num) or float(raciones_base_num) <= 0:
        return None

    divisor = float(raciones_base_num)
    ingredientes_normalizados = []
    for ing in ingredientes or []:
        ing_normalizado = dict(ing)
        cantidad_bruta = numero_seguro(ing_normalizado.get("cantidad_bruta", 0.0), 0.0)
        ing_normalizado["cantidad_bruta"] = cantidad_bruta / divisor
        ingredientes_normalizados.append(ing_normalizado)
    return ingredientes_normalizados


def preparar_receta_para_una_racion(datos_receta, ingredientes):
    """
    Devuelve cabecera e ingredientes listos para persistir, siempre referidos
    a una sola ración.
    """
    datos_normalizados = dict(datos_receta or {})
    raciones_base = numero_seguro(datos_normalizados.get("raciones_base", 1.0), 1.0)
    ingredientes_normalizados = normalizar_ingredientes_a_una_racion(
        ingredientes,
        raciones_base,
    )
    if ingredientes_normalizados is None:
        return None, None

    subtotal = sum(
        numero_seguro(ing.get("cantidad_bruta", 0.0), 0.0)
        * numero_seguro(ing.get("precio_unidad", 0.0), 0.0)
        for ing in ingredientes_normalizados
    )
    ci_pct = numero_seguro(datos_normalizados.get("costes_indirectos_pct", 0.0), 0.0)
    margen_pct = numero_seguro(datos_normalizados.get("margen_beneficio_pct", 0.0), 0.0)
    iva_pct = numero_seguro(datos_normalizados.get("iva_pct", 0.0), 0.0)
    coste_total = subtotal
    coste_para_precio = coste_total + (coste_total * ci_pct / 100)
    factor_margen = 1 - (margen_pct / 100)
    pvp_neto = coste_para_precio / factor_margen if factor_margen > 0 else 0.0
    pvp_final = pvp_neto + (pvp_neto * iva_pct / 100)

    datos_normalizados.update({
        "raciones_base": 1.0,
        "unidad_servicio": "ración",
        "coste_total": float(coste_total),
        "precio_venta_sin_iva": float(pvp_neto),
        "precio_venta_con_iva": float(pvp_final),
    })
    return datos_normalizados, ingredientes_normalizados


def crear_copia_receta_normalizada_supabase(datos_receta, ingredientes, raciones_base):
    """
    Crea una copia nueva de una receta antigua dimensionada a 1 racion.
    No modifica la receta original ni los menus que la usen.
    """
    ingredientes_normalizados = normalizar_ingredientes_a_una_racion(ingredientes, raciones_base)
    if ingredientes_normalizados is None:
        return False, "La base de raciones no es válida. No se puede normalizar esta receta.", None, None
    if not ingredientes_normalizados:
        return False, "No hay ingredientes para normalizar.", None, None

    ci_pct = numero_seguro(datos_receta.get("costes_indirectos_pct", 0.0), 0.0)
    margen_pct = numero_seguro(datos_receta.get("margen_beneficio_pct", 0.0), 0.0)
    iva_pct_receta = numero_seguro(datos_receta.get("iva_pct", 0.0), 0.0)
    subtotal = sum(
        numero_seguro(ing.get("cantidad_bruta", 0.0), 0.0) * numero_seguro(ing.get("precio_unidad", 0.0), 0.0)
        for ing in ingredientes_normalizados
    )
    coste_total = subtotal + (subtotal * ci_pct / 100)
    factor_margen = 1 - (margen_pct / 100)
    pvp_neto = coste_total / factor_margen if factor_margen > 0 else 0.0
    pvp_final = pvp_neto + (pvp_neto * iva_pct_receta / 100)

    nombre_original = str(datos_receta.get("nombre", "") or "Receta").strip() or "Receta"
    datos_copia = dict(datos_receta)
    datos_copia.update({
        "user_id": obtener_user_id_actual(),
        "codigo_receta": generar_codigo_receta(),
        "nombre": f"{nombre_original} - 1 ración",
        "raciones_base": 1.0,
        "unidad_servicio": "ración",
        "coste_total": float(coste_total),
        "precio_venta_sin_iva": float(pvp_neto),
        "precio_venta_con_iva": float(pvp_final),
        "activa": True
    })

    ok, mensaje, receta_guardada = guardar_receta_nueva_supabase(datos_copia, ingredientes_normalizados)
    return ok, mensaje, receta_guardada, ingredientes_normalizados


def eliminar_receta_supabase(receta_id):
    """
    Elimina una sola receta por id. Los ingredientes asociados dependen del cascade del inventario.
    """
    if not receta_id:
        return False, "No hay una receta seleccionada para eliminar."
    if not supabase_disponible or supabase is None:
        return False, "El inventario no está conectado correctamente."

    try:
        cabecera_actual, _ = cargar_receta_detalle_supabase(receta_id)
        if not cabecera_actual:
            return False, "No se pudo comprobar la receta antes de eliminarla."
        if not receta_es_modificable(cabecera_actual):
            return False, mensaje_receta_no_modificable()

        respuesta_uso_menu = (
            supabase
            .table("menu_recetas")
            .select("id")
            .eq("receta_id", receta_id)
            .limit(1)
            .execute()
        )
        if respuesta_uso_menu.data:
            return (
                False,
                "No se puede eliminar porque esta receta está usada en uno o varios menús. "
                "Primero elimínala del menú o duplica/modifica el menú."
            )

        (
            supabase
            .table("recetas")
            .delete()
            .eq("id", receta_id)
            .execute()
        )
        limpiar_cache_recetas_guardadas()
        return True, "Receta eliminada correctamente."
    except Exception as e:
        mensaje_error = str(e)
        mensaje_error_lower = mensaje_error.lower()
        if (
            "foreign key" in mensaje_error_lower
            or "violates foreign key" in mensaje_error_lower
            or "23503" in mensaje_error_lower
            or "menu_recetas" in mensaje_error_lower
        ):
            return (
                False,
                "No se puede eliminar porque esta receta está usada en uno o varios menús. "
                "Primero elimínala del menú o duplica/modifica el menú."
            )
        return False, f"Error al eliminar la receta en el inventario: {e}"


@st.cache_data(ttl=600)
def cargar_menus_supabase(user_id_actual=None):
    """
    Lee public.menus y devuelve un DataFrame seguro.
    """
    if not supabase_disponible or supabase is None:
        return pd.DataFrame(columns=MENUS_COLUMNAS)

    try:
        respuesta = (
            supabase
            .table("menus")
            .select(",".join(MENUS_COLUMNAS))
            .order("created_at", desc=True)
            .execute()
        )
        df = pd.DataFrame(respuesta.data or [])
        if df.empty:
            return pd.DataFrame(columns=MENUS_COLUMNAS)

        for col in MENUS_COLUMNAS:
            if col not in df.columns:
                df[col] = None
        df = df[MENUS_COLUMNAS].copy()
        user_id_actual = str(user_id_actual).strip() if user_id_actual else None
        user_ids = df["user_id"].apply(lambda valor: valor_user_id_menu({"user_id": valor}))
        if user_id_actual:
            return df[user_ids.isin([user_id_actual, None])].copy()
        return df[user_ids.isna()].copy()
    except Exception:
        return pd.DataFrame(columns=MENUS_COLUMNAS)


def cargar_menu_detalle_supabase(menu_id):
    """
    Lee cabecera de public.menus y lineas de public.menu_recetas.
    Enriquece las lineas con nombre, codigo y costes de public.recetas cuando existen.
    """
    lineas_vacias = pd.DataFrame(columns=MENU_RECETAS_COLUMNAS)
    if not menu_id or not supabase_disponible or supabase is None:
        return {}, lineas_vacias

    try:
        respuesta_menu = (
            supabase
            .table("menus")
            .select(",".join(MENUS_COLUMNAS))
            .eq("id", menu_id)
            .limit(1)
            .execute()
        )
        cabecera = (respuesta_menu.data or [{}])[0] if respuesta_menu.data else {}
        if cabecera and not menu_es_propio_o_antiguo(cabecera):
            return {}, lineas_vacias

        columnas_menu_recetas_nuevas = (
            "id,menu_id,receta_id,raciones_receta_en_menu,orden,"
            "seccion,observaciones,created_at"
        )
        try:
            respuesta_lineas = (
                supabase
                .table("menu_recetas")
                .select(columnas_menu_recetas_nuevas)
                .eq("menu_id", menu_id)
                .order("orden")
                .execute()
            )
        except Exception:
            respuesta_lineas = (
                supabase
                .table("menu_recetas")
                .select(",".join(MENU_RECETAS_COLUMNAS))
                .eq("menu_id", menu_id)
                .order("orden")
                .execute()
            )
        lineas_df = pd.DataFrame(respuesta_lineas.data or [])
        if lineas_df.empty:
            return cabecera, lineas_vacias

        if "raciones_receta_en_menu" in lineas_df.columns:
            lineas_df["raciones"] = lineas_df["raciones_receta_en_menu"]

        for col in MENU_RECETAS_COLUMNAS:
            if col not in lineas_df.columns:
                lineas_df[col] = None

        receta_ids = [
            str(receta_id)
            for receta_id in lineas_df["receta_id"].dropna().unique().tolist()
            if str(receta_id).strip()
        ]
        recetas_por_id = {}
        if receta_ids:
            respuesta_recetas = (
                supabase
                .table("recetas")
                .select("id,codigo_receta,nombre,raciones_base,coste_total,precio_venta_sin_iva,precio_venta_con_iva")
                .in_("id", receta_ids)
                .execute()
            )
            recetas_por_id = {
                str(receta.get("id")): receta
                for receta in (respuesta_recetas.data or [])
                if receta.get("id")
            }

        lineas_df["codigo_receta"] = lineas_df["receta_id"].apply(
            lambda receta_id: recetas_por_id.get(str(receta_id), {}).get("codigo_receta")
        )
        lineas_df["nombre_receta"] = lineas_df["receta_id"].apply(
            lambda receta_id: recetas_por_id.get(str(receta_id), {}).get("nombre")
        )
        lineas_df["coste_receta"] = lineas_df["receta_id"].apply(
            lambda receta_id: recetas_por_id.get(str(receta_id), {}).get("coste_total")
        )
        lineas_df["raciones_base"] = lineas_df["receta_id"].apply(
            lambda receta_id: recetas_por_id.get(str(receta_id), {}).get("raciones_base")
        )
        lineas_df["precio_receta_sin_iva"] = lineas_df["receta_id"].apply(
            lambda receta_id: recetas_por_id.get(str(receta_id), {}).get("precio_venta_sin_iva")
        )
        lineas_df["precio_receta_con_iva"] = lineas_df["receta_id"].apply(
            lambda receta_id: recetas_por_id.get(str(receta_id), {}).get("precio_venta_con_iva")
        )
        return cabecera, lineas_df.copy()
    except Exception:
        return {}, lineas_vacias


def error_tabla_clientes_no_activada(error):
    mensaje = str(error).lower()
    return (
        "clientes" in mensaje
        and (
            "does not exist" in mensaje
            or "schema cache" in mensaje
            or "42p01" in mensaje
            or "pgrst205" in mensaje
            or "could not find" in mensaje
        )
    )


def datos_cliente_desde_formulario(
    nombre, tipo_cliente, nif_cif, email, telefono, direccion,
    codigo_postal, ciudad, provincia, pais, observaciones
):
    return {
        "nombre": str(nombre or "").strip(),
        "tipo_cliente": str(tipo_cliente or "").strip() or None,
        "nif_cif": str(nif_cif or "").strip() or None,
        "email": str(email or "").strip() or None,
        "telefono": str(telefono or "").strip() or None,
        "direccion": str(direccion or "").strip() or None,
        "codigo_postal": str(codigo_postal or "").strip() or None,
        "ciudad": str(ciudad or "").strip() or None,
        "provincia": str(provincia or "").strip() or None,
        "pais": str(pais or "").strip() or "España",
        "observaciones": str(observaciones or "").strip() or None
    }


def cargar_clientes_supabase():
    """
    Lee los clientes del usuario conectado.
    """
    if not supabase_disponible or supabase is None:
        return False, "El inventario no está conectado correctamente.", pd.DataFrame(columns=CLIENTES_COLUMNAS)

    user_id_actual = obtener_user_id_actual()
    if not user_id_actual:
        return False, "Inicia sesión para guardar clientes en tu cuenta", pd.DataFrame(columns=CLIENTES_COLUMNAS)

    try:
        respuesta = (
            supabase
            .table("clientes")
            .select(",".join(CLIENTES_COLUMNAS))
            .eq("user_id", user_id_actual)
            .order("nombre")
            .execute()
        )
        df = pd.DataFrame(respuesta.data or [])
        if df.empty:
            return True, "", pd.DataFrame(columns=CLIENTES_COLUMNAS)
        for col in CLIENTES_COLUMNAS:
            if col not in df.columns:
                df[col] = None
        return True, "", df[CLIENTES_COLUMNAS].copy()
    except Exception as e:
        if error_tabla_clientes_no_activada(e):
            return False, MENSAJE_CLIENTES_NO_ACTIVADO, pd.DataFrame(columns=CLIENTES_COLUMNAS)
        return False, f"Error al cargar clientes: {e}", pd.DataFrame(columns=CLIENTES_COLUMNAS)


def guardar_cliente_supabase(cliente):
    """
    Guarda un cliente nuevo asociado al usuario conectado.
    """
    if not supabase_disponible or supabase is None:
        return False, "El inventario no está conectado correctamente.", None

    user_id_actual = obtener_user_id_actual()
    if not user_id_actual:
        return False, "Inicia sesión para guardar clientes en tu cuenta", None

    datos_guardar = dict(cliente)
    datos_guardar["user_id"] = user_id_actual
    if not datos_guardar.get("nombre"):
        return False, "Indica un nombre de cliente antes de guardar.", None

    try:
        respuesta = (
            supabase
            .table("clientes")
            .insert(datos_guardar)
            .execute()
        )
        cliente_guardado = (respuesta.data or [{}])[0]
        return True, "Cliente guardado correctamente.", cliente_guardado
    except Exception as e:
        if error_tabla_clientes_no_activada(e):
            return False, MENSAJE_CLIENTES_NO_ACTIVADO, None
        return False, f"Error al guardar cliente: {e}", None


def actualizar_cliente_supabase(cliente_id, cliente):
    """
    Actualiza un cliente del usuario conectado.
    """
    if not supabase_disponible or supabase is None:
        return False, "El inventario no está conectado correctamente."

    user_id_actual = obtener_user_id_actual()
    if not user_id_actual:
        return False, "Inicia sesión para guardar clientes en tu cuenta"

    cliente_id_limpio = str(cliente_id or "").strip()
    if not cliente_id_limpio:
        return False, "Selecciona un cliente antes de actualizar."
    if not cliente.get("nombre"):
        return False, "Indica un nombre de cliente antes de actualizar."

    try:
        (
            supabase
            .table("clientes")
            .update(dict(cliente))
            .eq("id", cliente_id_limpio)
            .eq("user_id", user_id_actual)
            .execute()
        )
        return True, "Cliente actualizado correctamente."
    except Exception as e:
        if error_tabla_clientes_no_activada(e):
            return False, MENSAJE_CLIENTES_NO_ACTIVADO
        return False, f"Error al actualizar cliente: {e}"


def eliminar_cliente_supabase(cliente_id):
    """
    Elimina un cliente del usuario conectado.
    """
    if not supabase_disponible or supabase is None:
        return False, "El inventario no está conectado correctamente."

    user_id_actual = obtener_user_id_actual()
    if not user_id_actual:
        return False, "Inicia sesión para guardar clientes en tu cuenta"

    cliente_id_limpio = str(cliente_id or "").strip()
    if not cliente_id_limpio:
        return False, "Selecciona un cliente antes de eliminar."

    try:
        (
            supabase
            .table("clientes")
            .delete()
            .eq("id", cliente_id_limpio)
            .eq("user_id", user_id_actual)
            .execute()
        )
        return True, "Cliente eliminado correctamente."
    except Exception as e:
        mensaje_error = str(e).lower()
        if error_tabla_clientes_no_activada(e):
            return False, MENSAJE_CLIENTES_NO_ACTIVADO
        if "foreign key" in mensaje_error or "violates foreign key" in mensaje_error or "23503" in mensaje_error:
            return False, "No se puede eliminar el cliente porque tiene facturas asociadas."
        return False, f"Error al eliminar cliente: {e}"


def error_tabla_facturacion_no_activada(error):
    mensaje = str(error).lower()
    return (
        (
            "facturas" in mensaje
            or "factura_lineas" in mensaje
            or "presupuestos" in mensaje
            or "presupuesto_menus" in mensaje
            or "factura_menus" in mensaje
        )
        and (
            "does not exist" in mensaje
            or "schema cache" in mensaje
            or "42p01" in mensaje
            or "pgrst205" in mensaje
            or "could not find" in mensaje
        )
    )


def linea_factura_vacia():
    return {
        "descripcion": "",
        "cantidad": 1.0,
        "unidad": "ud",
        "precio_unitario": 0.0,
        "descuento_pct": 0.0,
        "iva_pct": 21.0,
        "base_linea": 0.0,
        "iva_linea": 0.0,
        "total_linea": 0.0
    }


def calcular_linea_factura(linea):
    cantidad = numero_seguro(linea.get("cantidad", 1.0), 1.0)
    precio_unitario = numero_seguro(linea.get("precio_unitario", 0.0), 0.0)
    descuento_pct = numero_seguro(linea.get("descuento_pct", 0.0), 0.0)
    iva_pct = numero_seguro(linea.get("iva_pct", 21.0), 21.0)
    base_linea = cantidad * precio_unitario * (1 - descuento_pct / 100)
    iva_linea = base_linea * iva_pct / 100
    total_linea = base_linea + iva_linea
    linea_calculada = {
        "descripcion": str(linea.get("descripcion", "") or "").strip(),
        "cantidad": cantidad,
        "unidad": str(linea.get("unidad", "ud") or "ud").strip() or "ud",
        "precio_unitario": precio_unitario,
        "descuento_pct": descuento_pct,
        "iva_pct": iva_pct,
        "base_linea": round(base_linea, 2),
        "iva_linea": round(iva_linea, 2),
        "total_linea": round(total_linea, 2)
    }
    origen_tipo = str(linea.get("origen_tipo", "") or "").strip()
    if origen_tipo:
        linea_calculada["origen_tipo"] = origen_tipo
    if linea.get("origen_id") is not None and str(linea.get("origen_id") or "").strip():
        linea_calculada["origen_id"] = linea.get("origen_id")
    for campo_auxiliar in [
        "origen_receta_tipo",
        "origen_receta_id",
        "menu_origen_id",
        "menu_nombre_snapshot",
        "cantidad_menu",
        "coste_total_menu",
        "coste_linea_menu_documento",
        "precio_total_menu",
        "precio_linea_menu_documento",
        "observaciones",
    ]:
        if linea.get(campo_auxiliar) is not None and str(linea.get(campo_auxiliar) or "").strip():
            linea_calculada[campo_auxiliar] = linea.get(campo_auxiliar)
    return linea_calculada


def obtener_menu_id_desde_lineas_factura(lineas):
    for linea in lineas or []:
        origen_tipo = str(linea.get("origen_tipo", "") or "").strip()
        origen_id = str(linea.get("origen_id", "") or "").strip()
        menu_origen_id = str(linea.get("menu_origen_id", "") or "").strip()
        if origen_tipo == "menu" and origen_id:
            return origen_id
        if menu_origen_id:
            return menu_origen_id
    return None


def extraer_lineas_menu_documento(lineas):
    lineas_menu = []
    for indice, linea in enumerate(lineas or [], start=1):
        menu_id = str(
            linea.get("menu_origen_id", linea.get("origen_id", "")) or ""
        ).strip()
        if str(linea.get("origen_tipo", "") or "").strip() != "menu" or not menu_id:
            continue
        cantidad_menu = numero_seguro(
            linea.get("cantidad_menu", linea.get("cantidad", 0.0)),
            0.0,
        )
        coste_total_menu = numero_seguro(linea.get("coste_total_menu", 0.0), 0.0)
        if cantidad_menu <= 0:
            raise ValueError(f"La cantidad del menú {indice} debe ser mayor que 0.")
        if coste_total_menu < 0:
            raise ValueError(f"El coste del menú {indice} no puede ser negativo.")
        precio_total_menu = linea.get("precio_total_menu")
        if precio_total_menu is None:
            precio_total_menu = linea.get("precio_unitario")
        precio_total_menu = numero_seguro(precio_total_menu, 0.0)
        if precio_total_menu < 0:
            raise ValueError(f"El precio del menú {indice} no puede ser negativo.")
        lineas_menu.append({
            "menu_id": menu_id,
            "menu_nombre_snapshot": str(
                linea.get("menu_nombre_snapshot", linea.get("descripcion", "")) or ""
            ).strip(),
            "cantidad_menu": cantidad_menu,
            "coste_total_menu": coste_total_menu,
            "coste_linea_menu_documento": coste_total_menu * cantidad_menu,
            "precio_total_menu": precio_total_menu,
            "precio_linea_menu_documento": precio_total_menu * cantidad_menu,
            "observaciones": str(linea.get("observaciones", "") or "").strip() or None,
            "orden": indice,
        })
    return lineas_menu


def validar_documento_con_menus(factura, lineas):
    if not str(factura.get("cliente_id", "") or "").strip():
        return False, "Selecciona un cliente antes de guardar el documento."
    try:
        lineas_menu = extraer_lineas_menu_documento(lineas)
        if not lineas_menu:
            return False, "Añade al menos un menú antes de guardar el documento."
        tipo = str(factura.get("tipo_documento", "factura") or "factura")
        if tipo == "presupuesto":
            calcular_coste_presupuesto(lineas_menu)
        else:
            calcular_coste_factura(lineas_menu)
    except ValueError as exc:
        return False, str(exc)
    return True, ""


def obtener_nombre_menu_por_id(menu_id):
    menu_id_limpio = str(menu_id or "").strip()
    if not menu_id_limpio:
        return None
    menus_df = cargar_menus_supabase(obtener_user_id_actual())
    if menus_df.empty:
        return None
    coincidencias = menus_df[menus_df["id"].astype(str) == menu_id_limpio]
    if coincidencias.empty:
        return None
    return str(coincidencias.iloc[0].get("nombre") or "").strip() or None


def calcular_totales_factura(lineas, iva_pct=21, retencion_pct=0):
    lineas_calculadas = [calcular_linea_factura(linea) for linea in (lineas or [])]
    base_imponible = sum(numero_seguro(linea.get("base_linea", 0.0), 0.0) for linea in lineas_calculadas)
    iva_importe = sum(numero_seguro(linea.get("iva_linea", 0.0), 0.0) for linea in lineas_calculadas)
    retencion_pct_num = numero_seguro(retencion_pct, 0.0)
    retencion_importe = base_imponible * retencion_pct_num / 100
    total = base_imponible + iva_importe - retencion_importe
    iva_pct_num = numero_seguro(iva_pct, 21.0)
    return {
        "lineas": lineas_calculadas,
        "base_imponible": round(base_imponible, 2),
        "iva_pct": iva_pct_num,
        "iva_importe": round(iva_importe, 2),
        "retencion_pct": retencion_pct_num,
        "retencion_importe": round(retencion_importe, 2),
        "total": round(total, 2)
    }


def formatear_importe_euros(valor):
    importe = numero_seguro(valor, 0.0)
    return f"{importe:,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".")


def generar_pdf_factura(factura, lineas, cliente):
    """
    Genera un documento descargable. Usa PDF si reportlab esta disponible;
    si no, devuelve HTML imprimible sin requerir dependencias nuevas.
    """
    factura = factura or {}
    cliente = cliente or {}
    lineas_validas = [
        calcular_linea_factura(linea)
        for linea in (lineas or [])
        if str(linea.get("descripcion", "") or "").strip()
    ]
    totales = calcular_totales_factura(
        lineas_validas,
        factura.get("iva_pct", 21),
        factura.get("retencion_pct", 0)
    )
    tipo_documento = str(factura.get("tipo_documento") or "factura").strip() or "factura"
    tipo_visible = tipo_documento.replace("_", " ").title()
    numero = str(factura.get("numero_factura") or "sin_numero").strip() or "sin_numero"
    extension_base = "presupuesto" if tipo_documento == "presupuesto" else "factura"
    numero_archivo = re.sub(r"[^A-Za-z0-9_-]+", "_", numero).strip("_") or "sin_numero"
    menu_id_asociado = obtener_menu_id_desde_lineas_factura(lineas_validas)
    nombre_menu_asociado = str(factura.get("menu_nombre_asociado") or "").strip()
    if menu_id_asociado and not nombre_menu_asociado:
        nombre_menu_asociado = obtener_nombre_menu_por_id(menu_id_asociado) or ""

    datos_cliente = [
        cliente.get("nombre"),
        cliente.get("nif_cif"),
        cliente.get("email"),
        cliente.get("telefono"),
        cliente.get("direccion"),
        " ".join(
            str(valor or "").strip()
            for valor in [cliente.get("codigo_postal"), cliente.get("ciudad")]
            if str(valor or "").strip()
        ),
        cliente.get("provincia"),
        cliente.get("pais"),
    ]
    datos_cliente = [str(valor).strip() for valor in datos_cliente if str(valor or "").strip()]

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        estilos = getSampleStyleSheet()
        elementos = []
        if logo_samirarte_existe():
            elementos.extend([Image(str(LOGO_SAMIRARTE_PATH), width=120, height=60), Spacer(1, 8)])
        elementos.extend([
            Paragraph("Samirarte", estilos["Title"]),
            Paragraph(f"{tipo_visible} {html.escape(numero)}", estilos["Heading2"]),
            Paragraph(f"Fecha emisión: {factura.get('fecha_emision') or ''}", estilos["Normal"]),
            Paragraph(f"Fecha vencimiento: {factura.get('fecha_vencimiento') or ''}", estilos["Normal"]),
            Spacer(1, 12),
            Paragraph("Cliente", estilos["Heading3"]),
        ])
        for dato in datos_cliente:
            elementos.append(Paragraph(html.escape(dato), estilos["Normal"]))
        if factura.get("concepto"):
            elementos.extend([Spacer(1, 8), Paragraph(f"Concepto: {html.escape(str(factura.get('concepto')))}", estilos["Normal"])])
        if nombre_menu_asociado:
            elementos.append(Paragraph(f"Documento asociado a menú: {html.escape(nombre_menu_asociado)}", estilos["Normal"]))

        tabla = [["Descripción", "Cant.", "Ud.", "Precio", "Dto.", "IVA", "Base", "Total"]]
        for linea in lineas_validas:
            tabla.append([
                Paragraph(html.escape(linea["descripcion"]), estilos["BodyText"]),
                f"{linea['cantidad']:.3f}",
                html.escape(linea["unidad"]),
                formatear_importe_euros(linea["precio_unitario"]),
                f"{linea['descuento_pct']:.2f}%",
                f"{linea['iva_pct']:.2f}%",
                formatear_importe_euros(linea["base_linea"]),
                formatear_importe_euros(linea["total_linea"]),
            ])
        tabla_pdf = Table(tabla, repeatRows=1, colWidths=[150, 45, 35, 58, 42, 42, 58, 58])
        tabla_pdf.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.white),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        elementos.extend([Spacer(1, 14), tabla_pdf, Spacer(1, 14)])
        elementos.extend([
            Paragraph(f"Base imponible: {formatear_importe_euros(totales['base_imponible'])}", estilos["Normal"]),
            Paragraph(f"IVA: {formatear_importe_euros(totales['iva_importe'])}", estilos["Normal"]),
        ])
        if numero_seguro(totales.get("retencion_importe"), 0.0):
            elementos.append(Paragraph(f"Retención: {formatear_importe_euros(totales['retencion_importe'])}", estilos["Normal"]))
        elementos.append(Paragraph(f"Total: {formatear_importe_euros(totales['total'])}", estilos["Heading3"]))
        if factura.get("metodo_pago"):
            elementos.append(Paragraph(f"Método de pago: {html.escape(str(factura.get('metodo_pago')))}", estilos["Normal"]))
        if factura.get("notas"):
            elementos.append(Paragraph(f"Notas: {html.escape(str(factura.get('notas')))}", estilos["Normal"]))
        elementos.extend([Spacer(1, 12), Paragraph("Documento generado por Samirarte. Módulo de facturación interna.", estilos["Italic"])])
        doc.build(elementos)
        return {
            "data": buffer.getvalue(),
            "file_name": f"{extension_base}_{numero_archivo}.pdf",
            "mime": "application/pdf",
            "label": "Descargar PDF",
            "formato": "pdf"
        }
    except Exception:
        pass

    escapar = lambda valor: html.escape(str(valor or ""))
    filas_html = "\n".join(
        f"""
        <tr>
          <td>{escapar(linea['descripcion'])}</td>
          <td class="num">{linea['cantidad']:.3f}</td>
          <td>{escapar(linea['unidad'])}</td>
          <td class="num">{formatear_importe_euros(linea['precio_unitario'])}</td>
          <td class="num">{linea['descuento_pct']:.2f}%</td>
          <td class="num">{linea['iva_pct']:.2f}%</td>
          <td class="num">{formatear_importe_euros(linea['base_linea'])}</td>
          <td class="num">{formatear_importe_euros(linea['iva_linea'])}</td>
          <td class="num">{formatear_importe_euros(linea['total_linea'])}</td>
        </tr>
        """
        for linea in lineas_validas
    )
    cliente_html = "".join(f"<div>{escapar(dato)}</div>" for dato in datos_cliente)
    logo_b64 = logo_samirarte_base64()
    logo_html = ""
    if logo_b64:
        logo_html = f'<img class="logo" src="data:image/png;base64,{logo_b64}" alt="Samirarte">'
    retencion_html = ""
    if numero_seguro(totales.get("retencion_importe"), 0.0):
        retencion_html = f"""
        <div class="total-row">
          <span>Retención ({totales['retencion_pct']:.2f}%)</span>
          <strong>{formatear_importe_euros(totales['retencion_importe'])}</strong>
        </div>
        """
    menu_asociado_html = ""
    if nombre_menu_asociado:
        menu_asociado_html = f"""
        <section>
          <div class="label">Documento asociado a menú</div>
          <div>{escapar(nombre_menu_asociado)}</div>
        </section>
        """
    html_documento = f"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <title>{escapar(tipo_visible)} {escapar(numero)}</title>
  <style>
    body {{ font-family: Arial, Helvetica, sans-serif; color: #111; margin: 36px; }}
    header {{ border-bottom: 2px solid #111; padding-bottom: 16px; margin-bottom: 24px; }}
    .logo {{ display: block; max-width: 180px; height: auto; margin-bottom: 10px; }}
    h1 {{ margin: 0; font-size: 28px; }}
    h2 {{ margin: 8px 0 0; font-size: 20px; font-weight: normal; }}
    .grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 28px; margin-bottom: 24px; }}
    .label {{ font-weight: bold; margin-bottom: 6px; }}
    table {{ width: 100%; border-collapse: collapse; margin-top: 18px; }}
    th, td {{ border: 1px solid #111; padding: 7px; font-size: 12px; vertical-align: top; }}
    th {{ text-align: left; }}
    .num {{ text-align: right; white-space: nowrap; }}
    .totals {{ margin-left: auto; margin-top: 18px; width: 310px; }}
    .total-row {{ display: flex; justify-content: space-between; border-bottom: 1px solid #999; padding: 6px 0; }}
    .grand {{ font-size: 18px; border-bottom: 2px solid #111; }}
    footer {{ margin-top: 36px; font-size: 11px; color: #333; }}
  </style>
</head>
<body>
  <header>
    {logo_html}
    <h1>Samirarte</h1>
    <h2>{escapar(tipo_visible)} {escapar(numero)}</h2>
  </header>
  <section class="grid">
    <div>
      <div class="label">Cliente</div>
      {cliente_html}
    </div>
    <div>
      <div><strong>Fecha emisión:</strong> {escapar(factura.get('fecha_emision'))}</div>
      <div><strong>Fecha vencimiento:</strong> {escapar(factura.get('fecha_vencimiento'))}</div>
      <div><strong>Método de pago:</strong> {escapar(factura.get('metodo_pago'))}</div>
    </div>
  </section>
  <section>
    <div class="label">Concepto</div>
    <div>{escapar(factura.get('concepto'))}</div>
  </section>
  {menu_asociado_html}
  <table>
    <thead>
      <tr>
        <th>Descripción</th><th>Cant.</th><th>Ud.</th><th>Precio</th>
        <th>Dto.</th><th>IVA</th><th>Base</th><th>IVA línea</th><th>Total</th>
      </tr>
    </thead>
    <tbody>{filas_html}</tbody>
  </table>
  <section class="totals">
    <div class="total-row"><span>Base imponible</span><strong>{formatear_importe_euros(totales['base_imponible'])}</strong></div>
    <div class="total-row"><span>IVA</span><strong>{formatear_importe_euros(totales['iva_importe'])}</strong></div>
    {retencion_html}
    <div class="total-row grand"><span>Total</span><strong>{formatear_importe_euros(totales['total'])}</strong></div>
  </section>
  <section>
    <div class="label">Notas</div>
    <div>{escapar(factura.get('notas'))}</div>
  </section>
  <footer>Documento generado por Samirarte. Módulo de facturación interna.</footer>
</body>
</html>"""
    return {
        "data": html_documento.encode("utf-8"),
        "file_name": f"{extension_base}_{numero_archivo}.html",
        "mime": "text/html",
        "label": "Descargar documento",
        "formato": "html"
    }


def generar_numero_factura(tipo_documento):
    prefijos = {
        "presupuesto": "PRE",
        "factura": "FAC",
        "factura_rectificativa": "REC"
    }
    prefijo = prefijos.get(str(tipo_documento or "").strip(), "DOC")
    anio = pd.Timestamp.today().year
    if not supabase_disponible or supabase is None or not obtener_user_id_actual():
        return f"{prefijo}-{anio}-0001"

    try:
        (
            supabase
            .table("factura_lineas")
            .select("id")
            .limit(1)
            .execute()
        )
        respuesta = (
            supabase
            .table("facturas")
            .select("numero_factura")
            .eq("user_id", obtener_user_id_actual())
            .ilike("numero_factura", f"{prefijo}-{anio}-%")
            .execute()
        )
        ultimo = 0
        for fila in respuesta.data or []:
            numero = str(fila.get("numero_factura", "") or "")
            coincidencia = re.search(r"(\d+)$", numero)
            if coincidencia:
                ultimo = max(ultimo, int(coincidencia.group(1)))
        return f"{prefijo}-{anio}-{ultimo + 1:04d}"
    except Exception:
        return f"{prefijo}-{anio}-0001"


def preparar_relaciones_menu_documento(lineas, tipo_documento, documento_id):
    relaciones = []
    for linea in extraer_lineas_menu_documento(lineas):
        relacion = {
            "menu_id": linea["menu_id"],
            "menu_nombre_snapshot": linea["menu_nombre_snapshot"],
            "cantidad_menu": linea["cantidad_menu"],
            "coste_total_menu": linea["coste_total_menu"],
            "precio_total_menu": linea["precio_total_menu"],
            "observaciones": linea["observaciones"],
            "orden": linea["orden"],
        }
        if tipo_documento == "presupuesto":
            relacion.update({
                "presupuesto_id": documento_id,
                "coste_linea_menu_presupuesto": linea["coste_linea_menu_documento"],
                "precio_linea_menu_presupuesto": linea["precio_linea_menu_documento"],
            })
        else:
            relacion.update({
                "factura_id": documento_id,
                "coste_linea_menu_factura": linea["coste_linea_menu_documento"],
                "precio_linea_menu_factura": linea["precio_linea_menu_documento"],
            })
        relaciones.append(relacion)
    return relaciones


def guardar_presupuesto_supabase(presupuesto, lineas):
    user_id_actual = obtener_user_id_actual()
    documento_valido, mensaje_validacion = validar_documento_con_menus(
        presupuesto,
        lineas,
    )
    if not documento_valido:
        return False, mensaje_validacion, None

    lineas_menu = extraer_lineas_menu_documento(lineas)
    coste_total = calcular_coste_presupuesto(lineas_menu)
    precio_total = sum(
        numero_seguro(linea.get("precio_linea_menu_documento"), 0.0)
        for linea in lineas_menu
    )
    datos = {
        "user_id": user_id_actual,
        "cliente_id": presupuesto.get("cliente_id"),
        "numero_presupuesto": presupuesto.get("numero_factura"),
        "estado": presupuesto.get("estado", "borrador"),
        "fecha_emision": presupuesto.get("fecha_emision"),
        "fecha_vencimiento": presupuesto.get("fecha_vencimiento"),
        "concepto": presupuesto.get("concepto"),
        "coste_total": coste_total,
        "precio_total": precio_total,
        "notas": presupuesto.get("notas"),
    }
    respuesta = supabase.table("presupuestos").insert(datos).execute()
    guardado = (respuesta.data or [{}])[0]
    presupuesto_id = guardado.get("id")
    if not presupuesto_id:
        return False, "No se pudo obtener el identificador del presupuesto.", None
    relaciones = preparar_relaciones_menu_documento(
        lineas,
        "presupuesto",
        presupuesto_id,
    )
    supabase.table("presupuesto_menus").insert(relaciones).execute()
    guardado["numero_factura"] = guardado.get("numero_presupuesto")
    guardado["tipo_documento"] = "presupuesto"
    guardado["total"] = guardado.get("precio_total", precio_total)
    guardado["_origen_almacen"] = "presupuestos"
    return True, "Presupuesto guardado correctamente.", guardado


def guardar_relaciones_factura_supabase(factura_id, lineas):
    relaciones = preparar_relaciones_menu_documento(lineas, "factura", factura_id)
    if relaciones:
        supabase.table("factura_menus").insert(relaciones).execute()


def cargar_facturas_supabase():
    if not supabase_disponible or supabase is None:
        return False, "El inventario no está conectado correctamente.", pd.DataFrame(columns=FACTURAS_COLUMNAS)

    user_id_actual = obtener_user_id_actual()
    if not user_id_actual:
        return False, "Inicia sesión para guardar clientes en tu cuenta", pd.DataFrame(columns=FACTURAS_COLUMNAS)

    try:
        try:
            respuesta = (
                supabase
                .table("facturas")
                .select(",".join(FACTURAS_COLUMNAS_JERARQUIA))
                .eq("user_id", user_id_actual)
                .order("created_at", desc=True)
                .limit(25)
                .execute()
            )
        except Exception:
            respuesta = (
                supabase
                .table("facturas")
                .select(",".join(FACTURAS_COLUMNAS))
                .eq("user_id", user_id_actual)
                .order("created_at", desc=True)
                .limit(25)
                .execute()
            )
        df = pd.DataFrame(respuesta.data or [])
        if not df.empty:
            df["_origen_almacen"] = "facturas"
        for col in FACTURAS_COLUMNAS:
            if col not in df.columns:
                df[col] = None
        columnas_salida = FACTURAS_COLUMNAS_JERARQUIA + ["_origen_almacen"]
        if "_origen_almacen" not in df.columns:
            df["_origen_almacen"] = None

        try:
            respuesta_presupuestos = (
                supabase
                .table("presupuestos")
                .select(",".join(PRESUPUESTOS_COLUMNAS))
                .eq("user_id", user_id_actual)
                .order("created_at", desc=True)
                .limit(25)
                .execute()
            )
            presupuestos_df = pd.DataFrame(respuesta_presupuestos.data or [])
            if not presupuestos_df.empty:
                presupuestos_df = presupuestos_df.rename(columns={
                    "numero_presupuesto": "numero_factura",
                    "precio_total": "total",
                })
                presupuestos_df["tipo_documento"] = "presupuesto"
                presupuestos_df["base_imponible"] = presupuestos_df["total"]
                presupuestos_df["iva_pct"] = 0.0
                presupuestos_df["iva_importe"] = 0.0
                presupuestos_df["retencion_pct"] = 0.0
                presupuestos_df["retencion_importe"] = 0.0
                presupuestos_df["metodo_pago"] = None
                presupuestos_df["estado_cobro"] = None
                presupuestos_df["_origen_almacen"] = "presupuestos"
                for col in columnas_salida:
                    if col not in presupuestos_df.columns:
                        presupuestos_df[col] = None
                df = pd.concat(
                    [df[columnas_salida], presupuestos_df[columnas_salida]],
                    ignore_index=True,
                )
        except Exception:
            pass

        if df.empty:
            return True, "", pd.DataFrame(columns=columnas_salida)
        df = df.sort_values("created_at", ascending=False, na_position="last")
        return True, "", df[columnas_salida].copy()
    except Exception as e:
        if error_tabla_facturacion_no_activada(e):
            return False, MENSAJE_FACTURACION_NO_ACTIVADA, pd.DataFrame(columns=FACTURAS_COLUMNAS)
        return False, f"Error al cargar documentos: {e}", pd.DataFrame(columns=FACTURAS_COLUMNAS)


def cargar_factura_detalle_supabase(factura_id):
    if not supabase_disponible or supabase is None:
        return False, "El inventario no está conectado correctamente.", None, pd.DataFrame(columns=FACTURA_LINEAS_COLUMNAS)

    factura_id_limpio = str(factura_id or "").strip()
    if not factura_id_limpio:
        return False, "Selecciona un documento.", None, pd.DataFrame(columns=FACTURA_LINEAS_COLUMNAS)

    user_id_actual = obtener_user_id_actual()
    if not user_id_actual:
        return False, "Inicia sesión para cargar documentos.", None, pd.DataFrame(columns=FACTURA_LINEAS_COLUMNAS)

    try:
        try:
            respuesta_factura = (
                supabase
                .table("facturas")
                .select(",".join(FACTURAS_COLUMNAS_JERARQUIA))
                .eq("id", factura_id_limpio)
                .limit(1)
                .execute()
            )
        except Exception:
            respuesta_factura = (
                supabase
                .table("facturas")
                .select(",".join(FACTURAS_COLUMNAS))
                .eq("id", factura_id_limpio)
                .limit(1)
                .execute()
            )
        factura = (respuesta_factura.data or [None])[0]
        if not factura:
            try:
                respuesta_presupuesto = (
                    supabase
                    .table("presupuestos")
                    .select(",".join(PRESUPUESTOS_COLUMNAS))
                    .eq("id", factura_id_limpio)
                    .limit(1)
                    .execute()
                )
                presupuesto = (respuesta_presupuesto.data or [None])[0]
            except Exception:
                presupuesto = None
            if not presupuesto:
                return False, "No se encontró el documento seleccionado.", None, pd.DataFrame(columns=FACTURA_LINEAS_COLUMNAS)

            presupuesto_user_id = str(presupuesto.get("user_id") or "").strip()
            if presupuesto_user_id and presupuesto_user_id != user_id_actual:
                return False, "No puedes editar documentos de otro usuario.", None, pd.DataFrame(columns=FACTURA_LINEAS_COLUMNAS)
            respuesta_menus = (
                supabase
                .table("presupuesto_menus")
                .select(",".join(PRESUPUESTO_MENUS_COLUMNAS))
                .eq("presupuesto_id", factura_id_limpio)
                .order("orden")
                .execute()
            )
            lineas = []
            for relacion in respuesta_menus.data or []:
                lineas.append(calcular_linea_factura({
                    "descripcion": f"Menú: {relacion.get('menu_nombre_snapshot') or relacion.get('menu_id')}",
                    "cantidad": relacion.get("cantidad_menu", 1),
                    "cantidad_menu": relacion.get("cantidad_menu", 1),
                    "unidad": "menú",
                    "precio_unitario": relacion.get("precio_total_menu", 0),
                    "iva_pct": 0,
                    "origen_tipo": "menu",
                    "origen_id": relacion.get("menu_id"),
                    "menu_origen_id": relacion.get("menu_id"),
                    "menu_nombre_snapshot": relacion.get("menu_nombre_snapshot"),
                    "coste_total_menu": relacion.get("coste_total_menu", 0),
                    "coste_linea_menu_documento": relacion.get(
                        "coste_linea_menu_presupuesto",
                        0,
                    ),
                    "precio_total_menu": relacion.get("precio_total_menu", 0),
                    "precio_linea_menu_documento": relacion.get(
                        "precio_linea_menu_presupuesto",
                        0,
                    ),
                    "observaciones": relacion.get("observaciones"),
                }))
            presupuesto = dict(presupuesto)
            presupuesto.update({
                "numero_factura": presupuesto.get("numero_presupuesto"),
                "tipo_documento": "presupuesto",
                "base_imponible": presupuesto.get("precio_total", 0),
                "iva_pct": 0,
                "iva_importe": 0,
                "retencion_pct": 0,
                "retencion_importe": 0,
                "total": presupuesto.get("precio_total", 0),
                "_origen_almacen": "presupuestos",
            })
            return (
                True,
                "Presupuesto cargado correctamente.",
                presupuesto,
                pd.DataFrame(lineas),
            )

        factura_user_id = str(factura.get("user_id") or "").strip()
        if factura_user_id and factura_user_id != user_id_actual:
            return False, "No puedes editar documentos de otro usuario.", None, pd.DataFrame(columns=FACTURA_LINEAS_COLUMNAS)

        respuesta_lineas = (
            supabase
            .table("factura_lineas")
            .select(",".join(FACTURA_LINEAS_COLUMNAS))
            .eq("factura_id", factura_id_limpio)
            .order("orden", desc=False)
            .execute()
        )
        lineas_df = pd.DataFrame(respuesta_lineas.data or [])
        try:
            respuesta_menus_factura = (
                supabase
                .table("factura_menus")
                .select(",".join(FACTURA_MENUS_COLUMNAS))
                .eq("factura_id", factura_id_limpio)
                .order("orden")
                .execute()
            )
            relaciones_factura = respuesta_menus_factura.data or []
        except Exception:
            relaciones_factura = []
        if relaciones_factura and not lineas_df.empty:
            relaciones_por_menu = {
                str(relacion.get("menu_id")): relacion
                for relacion in relaciones_factura
            }
            for indice, linea in lineas_df.iterrows():
                relacion = relaciones_por_menu.get(str(linea.get("origen_id")))
                if not relacion:
                    continue
                lineas_df.at[indice, "cantidad_menu"] = relacion.get("cantidad_menu", 1)
                lineas_df.at[indice, "menu_origen_id"] = relacion.get("menu_id")
                lineas_df.at[indice, "menu_nombre_snapshot"] = relacion.get("menu_nombre_snapshot")
                lineas_df.at[indice, "coste_total_menu"] = relacion.get("coste_total_menu", 0)
                lineas_df.at[indice, "coste_linea_menu_documento"] = relacion.get(
                    "coste_linea_menu_factura",
                    0,
                )
                lineas_df.at[indice, "precio_total_menu"] = relacion.get("precio_total_menu", 0)
                lineas_df.at[indice, "precio_linea_menu_documento"] = relacion.get(
                    "precio_linea_menu_factura",
                    0,
                )
                lineas_df.at[indice, "observaciones"] = relacion.get("observaciones")
        elif relaciones_factura:
            lineas_df = pd.DataFrame([
                calcular_linea_factura({
                    "descripcion": f"Menú: {relacion.get('menu_nombre_snapshot') or relacion.get('menu_id')}",
                    "cantidad": relacion.get("cantidad_menu", 1),
                    "cantidad_menu": relacion.get("cantidad_menu", 1),
                    "unidad": "menú",
                    "precio_unitario": relacion.get("precio_total_menu", 0),
                    "iva_pct": 21,
                    "origen_tipo": "menu",
                    "origen_id": relacion.get("menu_id"),
                    "menu_origen_id": relacion.get("menu_id"),
                    "menu_nombre_snapshot": relacion.get("menu_nombre_snapshot"),
                    "coste_total_menu": relacion.get("coste_total_menu", 0),
                    "coste_linea_menu_documento": relacion.get(
                        "coste_linea_menu_factura",
                        0,
                    ),
                    "precio_total_menu": relacion.get("precio_total_menu", 0),
                    "precio_linea_menu_documento": relacion.get(
                        "precio_linea_menu_factura",
                        0,
                    ),
                    "observaciones": relacion.get("observaciones"),
                })
                for relacion in relaciones_factura
            ])
        if lineas_df.empty:
            lineas_df = pd.DataFrame(columns=FACTURA_LINEAS_COLUMNAS)
        for col in FACTURA_LINEAS_COLUMNAS:
            if col not in lineas_df.columns:
                lineas_df[col] = None
        mensaje = "Documento cargado correctamente."
        if not factura_user_id:
            mensaje = (
                "Documento antiguo sin usuario asignado. Puedes revisarlo, "
                "pero duplica el documento para guardarlo en tu cuenta."
            )
        return True, mensaje, factura, lineas_df.copy()
    except Exception as e:
        if error_tabla_facturacion_no_activada(e):
            return False, MENSAJE_FACTURACION_NO_ACTIVADA, None, pd.DataFrame(columns=FACTURA_LINEAS_COLUMNAS)
        return False, f"Error al cargar el documento: {e}", None, pd.DataFrame(columns=FACTURA_LINEAS_COLUMNAS)


def guardar_factura_supabase(factura, lineas):
    if not supabase_disponible or supabase is None:
        return False, "El inventario no está conectado correctamente.", None

    user_id_actual = obtener_user_id_actual()
    if not user_id_actual:
        return False, "Inicia sesión para guardar clientes en tu cuenta", None
    documento_valido, mensaje_validacion = validar_documento_con_menus(factura, lineas)
    if not documento_valido:
        return False, mensaje_validacion, None

    if str(factura.get("tipo_documento") or "") == "presupuesto":
        try:
            return guardar_presupuesto_supabase(factura, lineas)
        except Exception as error_presupuestos:
            if not error_tabla_facturacion_no_activada(error_presupuestos):
                return False, f"Error al guardar presupuesto: {error_presupuestos}", None

    lineas_validas = [
        calcular_linea_factura(linea)
        for linea in (lineas or [])
        if str(linea.get("descripcion", "") or "").strip()
    ]
    if not lineas_validas:
        return False, "Añade al menos una línea antes de guardar el documento.", None

    totales = calcular_totales_factura(
        lineas_validas,
        factura.get("iva_pct", 21),
        factura.get("retencion_pct", 0)
    )
    columnas_factura_escritura = [
        columna
        for columna in FACTURAS_COLUMNAS_JERARQUIA
        if columna not in {"id", "created_at", "updated_at"}
    ]
    datos_factura = {
        columna: factura.get(columna)
        for columna in columnas_factura_escritura
        if columna in factura
    }
    datos_factura["user_id"] = user_id_actual
    datos_factura["base_imponible"] = totales["base_imponible"]
    datos_factura["iva_importe"] = totales["iva_importe"]
    datos_factura["retencion_importe"] = totales["retencion_importe"]
    datos_factura["total"] = totales["total"]
    datos_factura["coste_total"] = calcular_coste_factura(
        extraer_lineas_menu_documento(lineas_validas)
    )

    try:
        try:
            respuesta_factura = (
                supabase
                .table("facturas")
                .insert(datos_factura)
                .execute()
            )
        except Exception as error_columnas_jerarquia:
            if not any(
                campo in str(error_columnas_jerarquia).lower()
                for campo in ("presupuesto_id", "coste_total", "schema cache", "pgrst204")
            ):
                raise
            datos_factura.pop("presupuesto_id", None)
            datos_factura.pop("coste_total", None)
            respuesta_factura = (
                supabase
                .table("facturas")
                .insert(datos_factura)
                .execute()
            )
        factura_guardada = (respuesta_factura.data or [{}])[0]
        factura_id = factura_guardada.get("id")
        if not factura_id:
            return False, "No se pudo obtener el identificador del documento guardado.", None

        lineas_guardar = []
        for orden, linea in enumerate(lineas_validas, start=1):
            lineas_guardar.append({
                "factura_id": factura_id,
                "origen_tipo": linea.get("origen_tipo", "manual") or "manual",
                "origen_id": linea.get("origen_id"),
                "descripcion": linea["descripcion"],
                "cantidad": linea["cantidad"],
                "unidad": linea["unidad"],
                "precio_unitario": linea["precio_unitario"],
                "descuento_pct": linea["descuento_pct"],
                "base_linea": linea["base_linea"],
                "iva_pct": linea["iva_pct"],
                "iva_linea": linea["iva_linea"],
                "total_linea": linea["total_linea"],
                "orden": orden
            })

        (
            supabase
            .table("factura_lineas")
            .insert(lineas_guardar)
            .execute()
        )
        try:
            guardar_relaciones_factura_supabase(factura_id, lineas_validas)
        except Exception:
            pass
        return True, "Documento guardado correctamente.", factura_guardada
    except Exception as e:
        if error_tabla_facturacion_no_activada(e):
            return False, MENSAJE_FACTURACION_NO_ACTIVADA, None
        return False, f"Error al guardar documento: {e}", None


def actualizar_presupuesto_supabase(presupuesto_id, presupuesto, lineas):
    documento_valido, mensaje_validacion = validar_documento_con_menus(
        presupuesto,
        lineas,
    )
    if not documento_valido:
        return False, mensaje_validacion, None
    lineas_menu = extraer_lineas_menu_documento(lineas)
    coste_total = calcular_coste_presupuesto(lineas_menu)
    precio_total = sum(
        numero_seguro(linea.get("precio_linea_menu_documento"), 0.0)
        for linea in lineas_menu
    )
    datos = {
        "cliente_id": presupuesto.get("cliente_id"),
        "numero_presupuesto": presupuesto.get("numero_factura"),
        "estado": presupuesto.get("estado", "borrador"),
        "fecha_emision": presupuesto.get("fecha_emision"),
        "fecha_vencimiento": presupuesto.get("fecha_vencimiento"),
        "concepto": presupuesto.get("concepto"),
        "coste_total": coste_total,
        "precio_total": precio_total,
        "notas": presupuesto.get("notas"),
    }
    respuesta = (
        supabase
        .table("presupuestos")
        .update(datos)
        .eq("id", presupuesto_id)
        .eq("user_id", obtener_user_id_actual())
        .execute()
    )
    supabase.table("presupuesto_menus").delete().eq(
        "presupuesto_id",
        presupuesto_id,
    ).execute()
    relaciones = preparar_relaciones_menu_documento(
        lineas,
        "presupuesto",
        presupuesto_id,
    )
    supabase.table("presupuesto_menus").insert(relaciones).execute()
    actualizado = (respuesta.data or [{}])[0]
    actualizado["numero_factura"] = actualizado.get("numero_presupuesto")
    return True, "Presupuesto actualizado correctamente.", actualizado


def actualizar_factura_supabase(factura_id, factura, lineas):
    if not supabase_disponible or supabase is None:
        return False, "El inventario no está conectado correctamente.", None

    user_id_actual = obtener_user_id_actual()
    if not user_id_actual:
        return False, "Inicia sesión para actualizar documentos.", None

    ok_detalle, mensaje_detalle, factura_existente, _ = cargar_factura_detalle_supabase(factura_id)
    if not ok_detalle:
        return False, mensaje_detalle, None

    factura_user_id = str((factura_existente or {}).get("user_id") or "").strip()
    if not factura_user_id:
        return False, "Este documento antiguo debe duplicarse antes de guardarse en tu cuenta.", None
    if factura_user_id != user_id_actual:
        return False, "No puedes editar documentos de otro usuario.", None
    if (factura_existente or {}).get("_origen_almacen") == "presupuestos":
        try:
            return actualizar_presupuesto_supabase(factura_id, factura, lineas)
        except Exception as exc:
            return False, f"Error al actualizar presupuesto: {exc}", None

    documento_valido, mensaje_validacion = validar_documento_con_menus(factura, lineas)
    if not documento_valido:
        return False, mensaje_validacion, None

    lineas_validas = [
        calcular_linea_factura(linea)
        for linea in (lineas or [])
        if str(linea.get("descripcion", "") or "").strip()
    ]
    if not lineas_validas:
        return False, "Añade al menos una línea antes de actualizar el documento.", None

    totales = calcular_totales_factura(
        lineas_validas,
        factura.get("iva_pct", 21),
        factura.get("retencion_pct", 0)
    )
    columnas_factura_escritura = [
        columna
        for columna in FACTURAS_COLUMNAS_JERARQUIA
        if columna not in {"id", "created_at", "updated_at"}
    ]
    datos_factura = {
        columna: factura.get(columna)
        for columna in columnas_factura_escritura
        if columna in factura
    }
    datos_factura["user_id"] = user_id_actual
    datos_factura["base_imponible"] = totales["base_imponible"]
    datos_factura["iva_importe"] = totales["iva_importe"]
    datos_factura["retencion_importe"] = totales["retencion_importe"]
    datos_factura["total"] = totales["total"]
    datos_factura["coste_total"] = calcular_coste_factura(
        extraer_lineas_menu_documento(lineas_validas)
    )

    try:
        try:
            respuesta_factura = (
                supabase
                .table("facturas")
                .update(datos_factura)
                .eq("id", factura_id)
                .eq("user_id", user_id_actual)
                .execute()
            )
        except Exception as error_columnas_jerarquia:
            if not any(
                campo in str(error_columnas_jerarquia).lower()
                for campo in ("presupuesto_id", "coste_total", "schema cache", "pgrst204")
            ):
                raise
            datos_factura.pop("presupuesto_id", None)
            datos_factura.pop("coste_total", None)
            respuesta_factura = (
                supabase
                .table("facturas")
                .update(datos_factura)
                .eq("id", factura_id)
                .eq("user_id", user_id_actual)
                .execute()
            )
        factura_actualizada = (respuesta_factura.data or [{}])[0]

        (
            supabase
            .table("factura_lineas")
            .delete()
            .eq("factura_id", factura_id)
            .execute()
        )

        lineas_guardar = []
        for orden, linea in enumerate(lineas_validas, start=1):
            lineas_guardar.append({
                "factura_id": factura_id,
                "origen_tipo": linea.get("origen_tipo", "manual") or "manual",
                "origen_id": linea.get("origen_id"),
                "descripcion": linea["descripcion"],
                "cantidad": linea["cantidad"],
                "unidad": linea["unidad"],
                "precio_unitario": linea["precio_unitario"],
                "descuento_pct": linea["descuento_pct"],
                "base_linea": linea["base_linea"],
                "iva_pct": linea["iva_pct"],
                "iva_linea": linea["iva_linea"],
                "total_linea": linea["total_linea"],
                "orden": orden
            })

        (
            supabase
            .table("factura_lineas")
            .insert(lineas_guardar)
            .execute()
        )
        try:
            supabase.table("factura_menus").delete().eq(
                "factura_id",
                factura_id,
            ).execute()
            guardar_relaciones_factura_supabase(factura_id, lineas_validas)
        except Exception:
            pass
        return True, "Documento actualizado correctamente.", factura_actualizada
    except Exception as e:
        if error_tabla_facturacion_no_activada(e):
            return False, MENSAJE_FACTURACION_NO_ACTIVADA, None
        return False, f"Error al actualizar documento: {e}", None


def duplicar_factura_supabase(factura_id):
    ok_detalle, mensaje_detalle, factura, lineas_df = cargar_factura_detalle_supabase(factura_id)
    if not ok_detalle:
        return False, mensaje_detalle, None

    user_id_actual = obtener_user_id_actual()
    if not user_id_actual:
        return False, "Inicia sesión para duplicar documentos.", None

    nueva_factura = dict(factura)
    for campo in ["id", "created_at", "updated_at", "base_imponible", "iva_importe", "retencion_importe", "total"]:
        nueva_factura.pop(campo, None)
    nueva_factura["user_id"] = user_id_actual
    nueva_factura["estado"] = "borrador"
    nueva_factura["estado_cobro"] = "pendiente"
    nueva_factura["numero_factura"] = generar_numero_factura(nueva_factura.get("tipo_documento"))

    lineas = []
    for _, linea in lineas_df.iterrows():
        linea_dict = linea.to_dict()
        for campo in ["id", "factura_id", "created_at"]:
            linea_dict.pop(campo, None)
        lineas.append(calcular_linea_factura(linea_dict))

    return guardar_factura_supabase(nueva_factura, lineas)


def fecha_factura_para_widget(valor):
    fecha = pd.to_datetime(valor, errors="coerce")
    if pd.isna(fecha):
        return pd.Timestamp.today().date()
    return fecha.date()


def cargar_factura_en_sesion(factura, lineas_df):
    tipo = str(factura.get("tipo_documento") or "factura").strip() or "factura"
    st.session_state["factura_id_en_edicion"] = str(factura.get("id") or "").strip() or None
    st.session_state["factura_presupuesto_origen_id"] = factura.get("presupuesto_id")
    st.session_state["factura_cliente_id"] = str(factura.get("cliente_id") or "").strip()
    st.session_state["factura_tipo_documento"] = tipo
    st.session_state[f"factura_numero_actual_{tipo}"] = str(factura.get("numero_factura") or "").strip()
    st.session_state["factura_fecha_emision"] = fecha_factura_para_widget(factura.get("fecha_emision"))
    st.session_state["factura_fecha_vencimiento"] = fecha_factura_para_widget(factura.get("fecha_vencimiento"))
    st.session_state["factura_estado"] = str(factura.get("estado") or "borrador").strip() or "borrador"
    st.session_state["factura_metodo_pago"] = str(factura.get("metodo_pago") or "")
    st.session_state["factura_estado_cobro"] = str(factura.get("estado_cobro") or "pendiente").strip() or "pendiente"
    st.session_state["factura_retencion_pct"] = numero_seguro(factura.get("retencion_pct"), 0.0)
    st.session_state["factura_concepto"] = str(factura.get("concepto") or "")
    st.session_state["factura_notas"] = str(factura.get("notas") or "")

    lineas = []
    if lineas_df is not None and not lineas_df.empty:
        for _, linea in lineas_df.iterrows():
            lineas.append(calcular_linea_factura(linea.to_dict()))
    st.session_state["factura_lineas_actual"] = lineas or [linea_factura_vacia()]


def calcular_totales_menu(lineas_menu):
    """
    Calcula el menu usando las raciones independientes de cada receta.
    """
    if lineas_menu is None:
        return {"coste_total": 0.0, "precio_total": None}

    if isinstance(lineas_menu, pd.DataFrame):
        lineas = lineas_menu.to_dict(orient="records")
    else:
        lineas = list(lineas_menu)

    if not lineas:
        return {"coste_total": 0.0, "precio_total": None}

    coste_total = calcular_coste_menu(lineas)
    precio_total = 0.0
    hay_precio = False

    for linea in lineas:
        factor = factor_linea_menu(linea)
        precio_receta = linea.get(
            "precio_receta_con_iva",
            linea.get(
                "precio_receta",
                linea.get("precio_venta_con_iva", linea.get("precio_total", None))
            )
        )

        if precio_receta is not None and not pd.isna(precio_receta):
            precio_total += numero_seguro(precio_receta, 0.0) * factor
            hay_precio = True

    return {
        "coste_total": coste_total,
        "precio_total": precio_total if hay_precio else None
    }


def crear_lineas_factura_desde_menu(
    cabecera_menu,
    lineas_menu_df,
    desglosar=False,
    cantidad_menu=1.0,
):
    """
    Convierte un menu guardado en lineas editables de documento de facturacion.
    No modifica el menu ni sus recetas.
    """
    cabecera_menu = cabecera_menu or {}
    lineas_menu = lineas_menu_df.to_dict(orient="records") if isinstance(lineas_menu_df, pd.DataFrame) else list(lineas_menu_df or [])
    menu_id = str(cabecera_menu.get("id", "") or "").strip()
    nombre_menu = str(cabecera_menu.get("nombre", "") or "Menú").strip()
    cantidad_menu = numero_seguro(cantidad_menu, 0.0)
    if cantidad_menu <= 0:
        raise ValueError("La cantidad del menú debe ser mayor que 0.")
    totales_menu = calcular_totales_menu(lineas_menu)
    coste_total_menu = numero_seguro(
        cabecera_menu.get("coste_total", totales_menu["coste_total"]),
        totales_menu["coste_total"],
    )
    aviso_precio = False

    if not desglosar:
        precio_total = cabecera_menu.get("precio_total", None)
        hay_precio_total = precio_total is not None and not pd.isna(precio_total)
        if hay_precio_total:
            precio_unitario = numero_seguro(precio_total, 0.0)
        else:
            precio_unitario = numero_seguro(
                coste_total_menu,
                0.0
            )
            aviso_precio = True
        return [
            calcular_linea_factura({
                "descripcion": f"Menú: {nombre_menu}",
                "cantidad": cantidad_menu,
                "cantidad_menu": cantidad_menu,
                "unidad": "menú",
                "precio_unitario": precio_unitario,
                "descuento_pct": 0.0,
                "iva_pct": 21.0,
                "origen_tipo": "menu",
                "origen_id": menu_id,
                "menu_origen_id": menu_id,
                "menu_nombre_snapshot": nombre_menu,
                "coste_total_menu": coste_total_menu,
                "coste_linea_menu_documento": coste_total_menu * cantidad_menu,
                "precio_total_menu": precio_unitario,
                "precio_linea_menu_documento": precio_unitario * cantidad_menu,
            })
        ], aviso_precio

    lineas_factura = []
    for linea in normalizar_lineas_menu(lineas_menu):
        raciones = numero_seguro(linea.get("raciones", 1.0), 1.0)
        divisor = raciones if raciones > 0 else 1.0
        precio_total_linea = linea.get("precio_total_linea", None)
        hay_precio_linea = precio_total_linea is not None and not pd.isna(precio_total_linea)
        if hay_precio_linea:
            precio_unitario = numero_seguro(precio_total_linea, 0.0) / divisor
        else:
            precio_unitario = numero_seguro(linea.get("coste_total_linea", 0.0), 0.0) / divisor
            aviso_precio = True
        lineas_factura.append(calcular_linea_factura({
            "descripcion": str(linea.get("nombre_receta", "") or "Receta del menú"),
            "cantidad": raciones,
            "unidad": "ración",
            "precio_unitario": precio_unitario,
            "descuento_pct": 0.0,
            "iva_pct": 21.0,
            "origen_tipo": "menu",
            "origen_id": menu_id,
            "origen_receta_tipo": "menu_receta",
            "origen_receta_id": linea.get("receta_id"),
            "menu_origen_id": menu_id
        }))

    return lineas_factura, aviso_precio


SECCIONES_MENU = ["", "Aperitivo", "Entrante", "Principal", "Postre", "Bebida", "Otro"]


def factor_linea_menu(linea):
    """
    Devuelve el multiplicador entre la receta completa y las raciones
    asignadas a esta linea del menu.
    """
    raciones = numero_seguro(
        linea.get("raciones_receta_en_menu", linea.get("raciones", 0.0)),
        0.0,
    )
    raciones_base = numero_seguro(
        linea.get("raciones_base_receta", linea.get("raciones_base", 0.0)),
        0.0,
    )
    if raciones <= 0 or raciones_base <= 0:
        raise ValueError("Las raciones de la receta y su base deben ser mayores que 0.")
    return raciones / raciones_base


def recalcular_linea_menu(linea):
    """
    Recalcula factor, coste y precio de una linea de menu sin modificar recetas.
    """
    linea_recalculada = dict(linea)
    raciones = numero_seguro(
        linea_recalculada.get(
            "raciones_receta_en_menu",
            linea_recalculada.get("raciones", 0.0),
        ),
        0.0,
    )
    raciones_base = numero_seguro(
        linea_recalculada.get(
            "raciones_base_receta",
            linea_recalculada.get("raciones_base", 0.0),
        ),
        0.0,
    )
    factor = factor_linea_menu(linea_recalculada)
    coste_receta = numero_seguro(linea_recalculada.get("coste_receta", 0.0), 0.0)
    precio_receta = linea_recalculada.get("precio_receta", None)
    hay_precio = precio_receta is not None and not pd.isna(precio_receta)

    coste_por_racion_receta = calcular_coste_por_racion(coste_receta, raciones_base)
    linea_recalculada["raciones_receta_en_menu"] = float(raciones)
    linea_recalculada["raciones"] = float(raciones)
    linea_recalculada["raciones_base_receta"] = float(raciones_base)
    linea_recalculada["raciones_base"] = float(raciones_base)
    linea_recalculada["coste_receta"] = float(coste_receta)
    linea_recalculada["coste_por_racion_receta"] = float(coste_por_racion_receta)
    linea_recalculada["coste_total_linea"] = float(coste_por_racion_receta * raciones)
    linea_recalculada["precio_total_linea"] = (
        float(numero_seguro(precio_receta, 0.0) * factor)
        if hay_precio else None
    )
    return linea_recalculada


def normalizar_lineas_menu(lineas_menu):
    """
    Ordena y recalcula todas las lineas del menu actual.
    """
    lineas = []
    for orden, linea in enumerate(lineas_menu or [], start=1):
        linea_normalizada = dict(linea)
        linea_normalizada["orden"] = int(numero_seguro(linea_normalizada.get("orden", orden), orden))
        linea_normalizada["seccion"] = str(linea_normalizada.get("seccion", "") or "").strip()
        lineas.append(recalcular_linea_menu(linea_normalizada))

    lineas = sorted(lineas, key=lambda item: int(numero_seguro(item.get("orden", 0), 0)))
    for orden, linea in enumerate(lineas, start=1):
        linea["orden"] = orden
    return lineas


def crear_linea_menu_desde_receta(receta_id, receta, raciones, orden, seccion=""):
    """
    Construye una linea de menu desde una receta guardada.
    """
    raciones_base = numero_seguro(receta.get("raciones_base", 1.0), 1.0)
    coste_receta = numero_seguro(receta.get("coste_total", 0.0), 0.0)
    precio_receta = receta.get("precio_venta_sin_iva", None)
    if precio_receta is None or pd.isna(precio_receta):
        precio_receta = receta.get("precio_venta_con_iva", None)

    hay_precio = precio_receta is not None and not pd.isna(precio_receta)
    linea = {
        "receta_id": str(receta_id),
        "codigo_receta": str(receta.get("codigo_receta", "") or ""),
        "nombre_receta": str(receta.get("nombre", "") or "Receta sin nombre"),
        "raciones_receta_en_menu": float(raciones),
        "raciones": float(raciones),
        "raciones_base_receta": float(raciones_base),
        "raciones_base": float(raciones_base),
        "coste_receta": float(coste_receta),
        "precio_receta": float(numero_seguro(precio_receta, 0.0)) if hay_precio else None,
        "orden": int(orden),
        "seccion": str(seccion or "").strip()
    }
    return recalcular_linea_menu(linea)


def preparar_lineas_menu_para_supabase(lineas_menu, columnas_nuevas=True):
    """
    Convierte lineas de menu al formato de public.menu_recetas.
    """
    if lineas_menu is None:
        return []

    if isinstance(lineas_menu, pd.DataFrame):
        lineas = lineas_menu.to_dict(orient="records")
    else:
        lineas = list(lineas_menu)

    lineas_preparadas = []
    for orden, linea in enumerate(lineas, start=1):
        receta_id = linea.get("receta_id", linea.get("id_receta", linea.get("id", None)))
        if not receta_id:
            continue
        raciones_receta = numero_seguro(
            linea.get("raciones_receta_en_menu", linea.get("raciones", 0.0)),
            0.0,
        )
        if raciones_receta <= 0:
            raise ValueError("Cada receta del menú debe tener raciones mayores que 0.")
        linea_preparada = {
            "receta_id": receta_id,
            "orden": int(numero_seguro(linea.get("orden", orden), orden)),
            "seccion": str(linea.get("seccion", "") or "").strip() or None
        }
        if columnas_nuevas:
            linea_preparada["raciones_receta_en_menu"] = raciones_receta
            linea_preparada["observaciones"] = (
                str(linea.get("observaciones", "") or "").strip() or None
            )
        else:
            linea_preparada["raciones"] = raciones_receta
        lineas_preparadas.append(linea_preparada)
    return lineas_preparadas


def insertar_lineas_menu_supabase(menu_id, lineas_menu):
    lineas_nuevas = preparar_lineas_menu_para_supabase(lineas_menu, columnas_nuevas=True)
    for linea in lineas_nuevas:
        linea["menu_id"] = menu_id
    try:
        supabase.table("menu_recetas").insert(lineas_nuevas).execute()
        return
    except Exception as error_columnas_nuevas:
        mensaje = str(error_columnas_nuevas).lower()
        if not any(
            texto in mensaje
            for texto in (
                "raciones_receta_en_menu",
                "observaciones",
                "schema cache",
                "pgrst204",
            )
        ):
            raise

    lineas_legacy = preparar_lineas_menu_para_supabase(lineas_menu, columnas_nuevas=False)
    for linea in lineas_legacy:
        linea["menu_id"] = menu_id
    supabase.table("menu_recetas").insert(lineas_legacy).execute()


def guardar_menu_supabase(datos_menu, lineas_menu):
    """
    Crea un menu nuevo y sus recetas asociadas. No modifica recetas.
    """
    if not supabase_disponible or supabase is None:
        return False, "El inventario no está conectado correctamente.", None

    try:
        lineas_menu = normalizar_lineas_menu(lineas_menu)
        lineas_preparadas = preparar_lineas_menu_para_supabase(
            lineas_menu,
            columnas_nuevas=True,
        )
        totales = calcular_totales_menu(lineas_menu)
    except ValueError as exc:
        return False, str(exc), None
    if not lineas_preparadas:
        return False, "Añade al menos una receta al menú antes de guardar.", None

    try:
        user_id_actual = obtener_user_id_actual()
        datos_guardar = dict(datos_menu)
        datos_guardar["user_id"] = user_id_actual
        if datos_guardar.get("coste_total") is None:
            datos_guardar["coste_total"] = totales["coste_total"]
        datos_guardar["coste_total"] = numero_seguro(datos_guardar.get("coste_total", 0.0))
        if datos_guardar.get("precio_total") is None:
            datos_guardar.pop("precio_total", None)
        if "precio_total" not in datos_guardar and totales["precio_total"] is not None:
            datos_guardar["precio_total"] = totales["precio_total"]

        respuesta_menu = (
            supabase
            .table("menus")
            .insert(datos_guardar)
            .execute()
        )
        menu_guardado = (respuesta_menu.data or [{}])[0]
        menu_id = menu_guardado.get("id")
        if not menu_id:
            return False, "No se pudo obtener el identificador del menú guardado.", None

        insertar_lineas_menu_supabase(menu_id, lineas_menu)

        try:
            cargar_menus_supabase.clear()
        except Exception:
            pass

        if user_id_actual:
            return True, "Menú guardado en tu cuenta", menu_guardado
        return True, "Menú guardado correctamente. Inicia sesión para guardar menús en tu cuenta", menu_guardado
    except Exception as e:
        return False, f"Error al guardar el menú en el inventario: {e}", None


def actualizar_menu_supabase(menu_id, datos_menu, lineas_menu):
    """
    Actualiza un menu y reemplaza solo sus lineas en public.menu_recetas.
    """
    if not menu_id:
        return False, "No hay un menú seleccionado para actualizar."
    if not supabase_disponible or supabase is None:
        return False, "El inventario no está conectado correctamente."

    try:
        lineas_menu = normalizar_lineas_menu(lineas_menu)
        lineas_preparadas = preparar_lineas_menu_para_supabase(
            lineas_menu,
            columnas_nuevas=True,
        )
        totales = calcular_totales_menu(lineas_menu)
    except ValueError as exc:
        return False, str(exc)
    if not lineas_preparadas:
        return False, "Añade al menos una receta al menú antes de actualizar."

    try:
        cabecera_actual, _ = cargar_menu_detalle_supabase(menu_id)
        if not cabecera_actual:
            return False, "No se pudo comprobar el menú antes de actualizarlo."
        if not menu_es_modificable(cabecera_actual):
            return False, mensaje_menu_no_modificable(cabecera_actual)

        datos_actualizar = dict(datos_menu)
        datos_actualizar["user_id"] = valor_user_id_menu(cabecera_actual)
        if datos_actualizar.get("coste_total") is None:
            datos_actualizar["coste_total"] = totales["coste_total"]
        datos_actualizar["coste_total"] = numero_seguro(datos_actualizar.get("coste_total", 0.0))
        if datos_actualizar.get("precio_total") is None:
            datos_actualizar.pop("precio_total", None)
        if "precio_total" not in datos_actualizar and totales["precio_total"] is not None:
            datos_actualizar["precio_total"] = totales["precio_total"]

        (
            supabase
            .table("menus")
            .update(datos_actualizar)
            .eq("id", menu_id)
            .execute()
        )

        (
            supabase
            .table("menu_recetas")
            .delete()
            .eq("menu_id", menu_id)
            .execute()
        )

        insertar_lineas_menu_supabase(menu_id, lineas_menu)

        try:
            cargar_menus_supabase.clear()
        except Exception:
            pass

        return True, "Menú actualizado correctamente."
    except Exception as e:
        return False, f"Error al actualizar el menú en el inventario: {e}"


def preparar_ingredientes_receta_para_sesion(ingredientes_df):
    """
    Adapta public.receta_ingredientes al formato interno de st.session_state.
    """
    if ingredientes_df is None or ingredientes_df.empty:
        return []

    ingredientes = []
    for _, fila in ingredientes_df.iterrows():
        codigo = str(fila.get("codigo_ingrediente", "") or "").strip().upper()
        ingredientes.append({
            "codigo": codigo if codigo else "S/C",
            "descripcion": str(fila.get("descripcion_ingrediente", "") or "").strip(),
            "cantidad_bruta": numero_seguro(fila.get("cantidad_bruta", 0.0)),
            "unidad_medida": str(fila.get("unidad_medida", "kg") or "kg").strip() or "kg",
            "merma": numero_seguro(fila.get("merma", 0.0)),
            "precio_unidad": numero_seguro(fila.get("precio_unidad", 0.0))
        })
    return ingredientes


# =============================================================================
# 🧠 PROCESAMIENTO INTELIGENTE CON CIA
# =============================================================================
def procesar_con_openai(texto_plano=None, bytes_imagen=None, mime_type=None):
    """
    Envia la informacion a CIA pasandole el inventario actual como contexto.
    """
    if not api_key:
        st.error("❌ Error: No se ha encontrado la clave de CIA en los secrets.")
        return []

    try:
        client = OpenAI(api_key=api_key)
        
        # Mandar un subconjunto ligero del catálogo real a la IA para optimizar la ventana de contexto
        catalogo_reducido = []
        for codigo, data in list(inventario_dict.items()):
            catalogo_reducido.append({
                "c": codigo,
                "d": data["descripcion"],
                "u": data.get("unidad_medida", "kg"),
                "m": data["merma"],
                "p": data["precio_unidad"]
            })

        bd_contexto = json.dumps(catalogo_reducido, ensure_ascii=False)

        prompt_sistema = f'''Eres un experto en cocina y escandallos. Trabaja SIEMPRE en dos fases:
FASE 1: extrae la receta culinaria real a partir del texto o imagen.
FASE 2: solo como apoyo mental, consulta el inventario para entender equivalencias posibles, pero NO devuelvas productos comerciales ni codigos.

Dispones del catálogo completo de ingredientes reales de nuestra cocina en formato JSON de referencia:
{bd_contexto}

INSTRUCCIONES CRÍTICAS:
1. Devuelve SOLO ingredientes que aparezcan claramente en la receta, imagen o texto, o que sean imprescindibles para la preparacion. No inventes ingredientes para completar.
2. Si no estas seguro de un ingrediente, omitelo o incluyelo con "confianza": "baja" y un motivo claro. No lo conviertas en producto de inventario.
3. Devuelve ingredientes genericos de cocina: pollo, arroz, judia verde, garrofon, tomate, aceite de oliva, sal. NO devuelvas productos comerciales como GARROFON 5K, GARROFON FINDUS, VERDURA PAELLA, ACEITE OLIVA 1L o SAL ESPECIAL HORNEAR HS.
4. No devuelvas codigos de inventario, marcas, formatos, proveedores, envases ni productos preparados del catalogo.
5. Descarta envases y utensilios: PAPEL, ENVOLVER, ENVOLTORIO, BANDEJA, CAJA, ETIQUETA, PACKAGING, SERVILLETA, PLATO, VASO, CUCHARA, TENEDOR, CUCHILLO, MOLDE.
6. No uses automaticamente carcasa, hueso, piel, recorte ni sal especial hornear salvo que el texto lo pida expresamente.
7. Si la receta parece una paella o arroz, no añadas mango, vinagre, espinaca, setas, zanahoria, secreto, carcasa o habas salvo que aparezcan explicitamente en el texto o imagen.
8. Si detectas cantidades, conserva con precisión la cantidad y su unidad original. Convierte gramos (g, gr, gr.) a kg dividiendo entre 1000 y mililitros (ml) a litros dividiendo entre 1000. Ejemplos: 400 gr = 0.4 kg; 75 gr = 0.075 kg. Para dientes, cucharadas, cucharaditas y gotas usa unidad "ud". Si no hay cantidad, usa 0.
9. Si detectas el nombre del plato o receta, devuelvelo como "nombre_receta" con el nombre limpio y legible.
10. Si detectas raciones de receta en expresiones como "6 porciones", "6 raciones", "para 6 personas", "serves 6" o "6 servings", devuelve ese numero como "raciones_base".
11. La app guardara recetas nuevas siempre a 1 racion. No dividas cantidades aunque detectes otra racion base; ese dato sera solo un aviso para el usuario.
12. Ignora titulos de columnas de cabecera de Excel, importes totales o subtotales de facturas.

REQUISITO EXCLUSIVO DE RESPUESTA: Devuelve ÚNICAMENTE JSON puro sin bloques de código markdown de tipo ```json y sin explicaciones adicionales.
Devuelve preferentemente este objeto:
{{
  "nombre_receta": "Paella valenciana",
  "raciones_base": 6,
  "ingredientes": [
    {{
      "nombre": "pollo",
      "cantidad": 0.5,
      "unidad": "kg",
      "confianza": "alta",
      "motivo": "aparece en la receta"
    }}
  ]
}}
Mantén compatibilidad si no puedes usar objeto y devuelve una lista generica:
[
  {{"nombre": "pollo", "cantidad": 0.5, "unidad": "kg", "confianza": "alta", "motivo": "aparece en la receta"}}
]'''

        contenido_usuario = []
        if texto_plano:
            contenido_usuario.append({
                "type": "text",
                "text": f"Analiza esta receta o lista. Primero extrae solo ingredientes culinarios genericos reales; despues deja que la app haga el cruce con inventario:\n\n{texto_plano}"
            })
        elif bytes_imagen:
            base64_image = base64.b64encode(bytes_imagen).decode('utf-8')
            contenido_usuario.append({
                "type": "text",
                "text": "Analiza esta imagen (puede ser una receta, tique, albarán o factura) y extrae solo ingredientes culinarios genericos reales. No devuelvas codigos ni productos comerciales del inventario."
            })
            contenido_usuario.append({
                "type": "image_url",
                "image_url": {
                    "url": f"data:{mime_type};base64,{base64_image}"
                }
            })

        respuesta = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": prompt_sistema},
                {"role": "user", "content": contenido_usuario}
            ],
            temperature=0.1
        )

        json_texto = respuesta.choices[0].message.content.strip()
        if json_texto.startswith("```"):
            json_texto = json_texto.split("\n", 1)[1].rsplit("\n", 1)[0].strip()
            if json_texto.startswith("json"):
                json_texto = json_texto[4:].strip()

        return json.loads(json_texto)

    except Exception as e:
        st.error(f"⚠️ Error al conectar con CIA: {str(e)}")
        return []


def normalizar_respuesta_ingredientes_ia(respuesta_ia):
    """
    Acepta el formato antiguo y el formato culinario generico con matching posterior.
    """
    nombre_receta = ""
    if isinstance(respuesta_ia, list):
        ingredientes = respuesta_ia
        raciones_base = None
    elif isinstance(respuesta_ia, dict):
        ingredientes = respuesta_ia.get("ingredientes", [])
        nombre_receta = str(respuesta_ia.get("nombre_receta", "") or "").strip()
        raciones_base = pd.to_numeric(respuesta_ia.get("raciones_base"), errors="coerce")
        raciones_base = None if pd.isna(raciones_base) or float(raciones_base) <= 0 else float(raciones_base)
    else:
        return "", None, []

    ingredientes_limpios = []
    ingredientes_vistos = set()
    es_receta_arroz = receta_parece_arroz(nombre_receta)
    for ing in ingredientes if isinstance(ingredientes, list) else []:
        if not isinstance(ing, dict):
            continue
        nombre_original = str(
            ing.get("nombre")
            or ing.get("ingrediente")
            or ing.get("descripcion")
            or ""
        ).strip()
        descripcion_generica = descripcion_generica_ingrediente(nombre_original)
        if not descripcion_generica or parece_material_no_alimentario(descripcion_generica):
            continue
        confianza = str(ing.get("confianza", "media") or "media").strip().lower()
        motivo = str(ing.get("motivo", "") or "").strip()
        if es_receta_arroz and ingrediente_incoherente_en_arroz(descripcion_generica) and not motivo_indica_presencia_explicita(motivo):
            continue
        clave_generica = normalizar_texto_busqueda(descripcion_generica)
        if not clave_generica or clave_generica in ingredientes_vistos:
            continue
        ingredientes_vistos.add(clave_generica)

        unidad_original = ing.get("unidad", ing.get("unidad_medida", ""))
        unidad_medida = normalizar_unidad_ia(unidad_original, descripcion_generica)
        cantidad_bruta = convertir_cantidad_ia(
            ing.get("cantidad", ing.get("cantidad_bruta", 0.0)),
            unidad_original or unidad_medida
        )

        coincidencia, alternativas, estado_match = buscar_coincidencia_principal_inventario(descripcion_generica, confianza)
        requiere_revision = estado_match != "clara"
        ing_limpio = {
            "codigo": "S/C",
            "descripcion": descripcion_generica.title(),
            "unidad_medida": unidad_medida,
            "cantidad_bruta": cantidad_bruta,
            "merma": 0.0,
            "precio_unidad": 0.0,
            "ia_nombre_original": nombre_original,
            "ia_confianza": confianza or "media",
            "ia_motivo": motivo or "detectado por IA",
            "ia_requiere_revision": requiere_revision,
            "ia_estado_match": estado_match,
            "ia_coincidencia_elegida": coincidencia.get("codigo") if coincidencia else "",
            "ia_alternativas": [alt.get("codigo") for alt in alternativas]
        }
        if coincidencia and not requiere_revision:
            ing_limpio.update({
                "codigo": coincidencia["codigo"],
                "familia": coincidencia.get("familia", "SIN CLASIFICAR"),
                "descripcion": coincidencia["descripcion"],
                "unidad_medida": coincidencia.get("unidad_medida", unidad_medida) or unidad_medida,
                "merma": coincidencia.get("merma", 0.0),
                "precio_unidad": coincidencia.get("precio_unidad", 0.0)
            })
        ingredientes_limpios.append(ing_limpio)

    return nombre_receta, raciones_base, ingredientes_limpios


def incorporar_ingredientes_ia(respuesta_ia):
    nombre_receta_detectado, raciones_base_detectadas, nuevos = normalizar_respuesta_ingredientes_ia(respuesta_ia)
    if not nuevos:
        return False

    if nombre_receta_detectado:
        st.session_state['receta_nombre'] = nombre_receta_detectado
        st.session_state['sincronizar_campos_receta'] = True

    # Si la IA detectó un número de raciones en la captura, normalizamos
    # automáticamente las cantidades a 1 ración (todo se guarda referido a 1 ración).
    # Solo normalizamos automáticamente si las raciones detectadas son un valor
    # razonable (entero >= 1). Evitamos normalizar frente a detecciones erróneas
    # que pudieran producir factores extremos (p.ej. 0.01 -> x100).
    if (
        raciones_base_detectadas is not None
        and isinstance(raciones_base_detectadas, (int, float))
        and float(raciones_base_detectadas) >= 1.0
        and float(raciones_base_detectadas) <= 10000.0
    ):
        factor_normalizar = 1.0 / float(raciones_base_detectadas)
        ingredientes_normalizados = ajustar_ingredientes_por_raciones(nuevos, factor_normalizar)
        st.session_state['ingredientes'].extend(ingredientes_normalizados)
        # Guardar copia de las cantidades originales detectadas por la IA
        st.session_state['ingredientes_base_raciones'] = [dict(ing) for ing in nuevos]
        st.session_state['factor_raciones'] = float(factor_normalizar)
        # Actualizar raciones base de la receta a 1 (ahora todo está normalizado por ración)
        st.session_state['receta_raciones_base'] = 1.0
        st.session_state['input_raciones_base'] = 1.0
        st.session_state['raciones_base'] = 1.0
        st.session_state['raciones_base_aplicadas'] = 1.0
        raciones_detectadas_texto = f"{raciones_base_detectadas:g}"
        st.session_state["aviso_raciones_ia"] = (
            f"La receta detectada estaba pensada para {raciones_detectadas_texto} raciones; "
            "las cantidades se han normalizado a 1 ración automáticamente."
        )
    else:
        st.session_state['ingredientes'].extend(nuevos)
        st.session_state['ingredientes_base_raciones'] = [dict(ing) for ing in st.session_state['ingredientes']]
        st.session_state['factor_raciones'] = 1.0
        # Si la IA devolvió un valor de raciones sospechoso (por ejemplo <1 o no numérico),
        # avisamos al usuario para que ajuste manualmente.
        if raciones_base_detectadas is not None:
            try:
                val = float(raciones_base_detectadas)
                if val < 1.0 or val > 10000.0:
                    st.session_state["aviso_raciones_ia"] = (
                        f"La detección de raciones ({raciones_base_detectadas}) parece errónea; "
                        "no se ha normalizado automáticamente. Ajusta las cantidades manualmente."
                    )
            except Exception:
                st.session_state["aviso_raciones_ia"] = (
                    "La detección de raciones no es numérica; no se ha normalizado automáticamente."
                )

    return True


def marcar_receta_modificada_manualmente():
    st.session_state['ingredientes_base_raciones'] = None
    st.session_state['factor_raciones'] = 1.0


def normalizar_raciones_desde_campo_receta():
    """
    Aplica el valor introducido en 'Raciones base receta' como divisor y deja
    inmediatamente la receta activa expresada para una ración.
    """
    raciones = numero_seguro(st.session_state.get("input_raciones_base_receta", 1.0), 1.0)
    if raciones <= 0:
        st.session_state["input_raciones_base_receta"] = 1.0
        return

    if abs(raciones - 1.0) > 1e-9 and st.session_state.get("ingredientes"):
        ingredientes_normalizados = normalizar_ingredientes_a_una_racion(
            st.session_state["ingredientes"],
            raciones,
        )
        if ingredientes_normalizados is not None:
            st.session_state["ingredientes"] = ingredientes_normalizados
            st.session_state["aviso_raciones_ia"] = (
                f"Las cantidades se han dividido entre {raciones:g} y la receta "
                "ha quedado referida a 1 ración."
            )
            marcar_receta_modificada_manualmente()

    st.session_state["receta_raciones_base"] = 1.0
    st.session_state["raciones_base"] = 1.0
    st.session_state["raciones_deseadas"] = 1.0
    st.session_state["receta_raciones_deseadas"] = 1.0
    st.session_state["raciones_base_aplicadas"] = 1.0
    st.session_state["raciones_deseadas_aplicadas"] = 1.0
    st.session_state["input_raciones_base_receta"] = 1.0


def sincronizar_inputs_raciones():
    if "input_raciones_base_receta_pendiente" in st.session_state:
        st.session_state["input_raciones_base_receta"] = float(
            st.session_state.pop("input_raciones_base_receta_pendiente")
        )

    if st.session_state.get('sincronizar_campos_receta', False):
        st.session_state['input_nombre_plato'] = st.session_state.get('receta_nombre', "Mi Receta")
        st.session_state['input_receta_categoria'] = st.session_state.get('receta_categoria', "")
        st.session_state['input_receta_tipo_plato'] = st.session_state.get('receta_tipo_plato', "")
        st.session_state['input_receta_observaciones'] = st.session_state.get('receta_observaciones', "")
        st.session_state['sincronizar_campos_receta'] = False

    if st.session_state.get('sincronizar_inputs_raciones', False):
        raciones_base = float(st.session_state.get('receta_raciones_base', st.session_state['raciones_base']))
        raciones_deseadas = float(st.session_state.get('receta_raciones_deseadas', st.session_state['raciones_deseadas']))
        st.session_state['input_raciones_base'] = raciones_base
        st.session_state['input_raciones_deseadas'] = raciones_deseadas
        st.session_state['raciones_base'] = raciones_base
        st.session_state['raciones_deseadas'] = raciones_deseadas
        st.session_state['raciones_base_aplicadas'] = raciones_base
        st.session_state['raciones_deseadas_aplicadas'] = raciones_deseadas
        st.session_state['sincronizar_inputs_raciones'] = False


def sincronizar_widgets_menu():
    if st.session_state.get('sincronizar_campos_menu', False):
        if 'menu_nombre_pendiente' in st.session_state:
            st.session_state['menu_nombre'] = st.session_state.pop('menu_nombre_pendiente')
        if 'menu_tipo_pendiente' in st.session_state:
            st.session_state['menu_tipo'] = st.session_state.pop('menu_tipo_pendiente')
        if 'menu_numero_comensales_pendiente' in st.session_state:
            st.session_state['menu_numero_comensales'] = st.session_state.pop('menu_numero_comensales_pendiente')
        st.session_state['sincronizar_campos_menu'] = False

# =============================================================================
# 📊 GENERADOR DE EXCEL CON FÓRMULAS CONSOLIDADAS
# =============================================================================
def ajustar_ingredientes_por_raciones(ingredientes, factor):
    """
    Devuelve una copia de ingredientes ajustando solo la cantidad bruta.
    """
    ingredientes_ajustados = []
    for ing in ingredientes:
        ing_ajustado = dict(ing)
        cantidad = pd.to_numeric(ing_ajustado.get('cantidad_bruta', 0.0), errors='coerce')
        cantidad = 0.0 if pd.isna(cantidad) else float(cantidad)
        ing_ajustado['cantidad_bruta'] = cantidad * factor
        ingredientes_ajustados.append(ing_ajustado)
    return ingredientes_ajustados


def calcular_ajuste_raciones(ingredientes_base, raciones_base, raciones_deseadas):
    raciones_base_num = pd.to_numeric(raciones_base, errors='coerce')
    raciones_deseadas_num = pd.to_numeric(raciones_deseadas, errors='coerce')
    if pd.isna(raciones_base_num) or pd.isna(raciones_deseadas_num) or float(raciones_base_num) <= 0 or float(raciones_deseadas_num) <= 0:
        raise ValueError("Las raciones base y deseadas deben ser mayores que 0.")
    factor = float(raciones_deseadas_num) / float(raciones_base_num)
    return factor, ajustar_ingredientes_por_raciones(ingredientes_base, factor)


def raciones_han_cambiado(raciones_base, raciones_deseadas, raciones_base_aplicadas, raciones_deseadas_aplicadas):
    return (
        float(raciones_base) != float(raciones_base_aplicadas)
        or float(raciones_deseadas) != float(raciones_deseadas_aplicadas)
    )


def generar_excel(
    nombre_plato,
    ingredientes,
    costes_indirectos_pct,
    margen_beneficio_pct,
    iva_pct,
    raciones_base=1.0,
    raciones_deseadas=1.0,
    factor_raciones=1.0
):
    wb = openpyxl.Workbook()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    ws = wb.active
    ws.title = "Ficha Técnica"

    # Título Principal Estilizado en Azul Marino
    ws.merge_cells('A1:H1')
    titulo_cell = ws['A1']
    titulo_cell.value = f"FICHA TÉCNICA: {nombre_plato.upper()}"
    titulo_cell.font = Font(bold=True, size=14, color="FFFFFF")
    titulo_cell.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    # Datos de raciones usados para escalar la receta
    ws['A2'] = "Raciones base:"
    ws['B2'] = float(raciones_base)
    ws['D2'] = "Raciones calculadas / deseadas:"
    ws['E2'] = float(raciones_deseadas)
    ws['A3'] = "Factor de ajuste:"
    ws['B3'] = float(factor_raciones)
    for cell_ref in ['A2', 'D2', 'A3']:
        ws[cell_ref].font = Font(bold=True)
    ws['B2'].number_format = '#,##0.00'
    ws['E2'].number_format = '#,##0.00'
    ws['B3'].number_format = '#,##0.0000'

    # Cabeceras alineadas con la nueva columna Código
    headers = ['Código', 'Ingrediente', 'Cantidad Bruta', 'Unidad', '% Merma', 'Cantidad Neta', 'Precio Unidad (€)', 'Coste Total (€)']
    header_row = 5
    start_row = 6
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=header_row, column=col_num, value=header)
    col_idx = {header: idx for idx, header in enumerate(headers, 1)}
    col_letras = {header: get_column_letter(idx) for header, idx in col_idx.items()}
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[header_row].height = 25

    # Insertar registros y configurar fórmulas dinámicas
    for i, ing in enumerate(ingredientes):
        row = start_row + i
        ws.cell(row=row, column=col_idx['Código'], value=ing.get('codigo', 'S/C'))
        ws.cell(row=row, column=col_idx['Ingrediente'], value=ing.get('descripcion', ''))
        ws.cell(row=row, column=col_idx['Cantidad Bruta'], value=float(ing.get('cantidad_bruta', 0.0))).number_format = '#,##0.000'
        ws.cell(row=row, column=col_idx['Unidad'], value=ing.get('unidad_medida', 'kg'))
        ws.cell(row=row, column=col_idx['% Merma'], value=float(ing.get('merma', 0.0))/100).number_format = '0.00%'
        ws.cell(
            row=row,
            column=col_idx['Cantidad Neta'],
            value=f"={col_letras['Cantidad Bruta']}{row}*(1-{col_letras['% Merma']}{row})",
        ).number_format = '#,##0.000'
        ws.cell(row=row, column=col_idx['Precio Unidad (€)'], value=float(ing.get('precio_unidad', 0.0))).number_format = '#,##0.00 €'
        ws.cell(
            row=row,
            column=col_idx['Coste Total (€)'],
            value=f"={col_letras['Cantidad Bruta']}{row}*{col_letras['Precio Unidad (€)']}{row}",
        ).number_format = '#,##0.00 €'

    last_row = start_row + len(ingredientes) - 1 if ingredientes else start_row
    res = last_row + 2
    total_col = col_idx['Coste Total (€)']
    total_col_letter = col_letras['Coste Total (€)']
    
    # Cálculos Financieros del Escandallo
    ws.cell(row=res, column=6, value="Subtotal Ingredientes:").font = Font(bold=True)
    ws.cell(row=res, column=total_col, value=f"=SUM({total_col_letter}{start_row}:{total_col_letter}{last_row})").number_format = '#,##0.00 €'

    ci = costes_indirectos_pct if costes_indirectos_pct is not None else 10.0
    mb = margen_beneficio_pct if margen_beneficio_pct is not None else 30.0
    iva = iva_pct if iva_pct is not None else 10.0

    ws.cell(row=res+1, column=6, value=f"Costes Indirectos ({ci}%):")
    ws.cell(row=res+1, column=total_col, value=f"={total_col_letter}{res}*({ci}/100)").number_format = '#,##0.00 €'

    ws.cell(row=res+2, column=6, value="COSTE TOTAL DEL PLATO:").font = Font(bold=True)
    ws.cell(row=res+2, column=total_col, value=f"={total_col_letter}{res}+{total_col_letter}{res+1}").number_format = '#,##0.00 €'

    ws.cell(row=res+4, column=6, value=f"Margen Deseado ({mb}%):").font = Font(bold=True)
    
    ws.cell(row=res+5, column=6, value="PRECIO VENTA (SIN IVA):").font = Font(bold=True)
    pvp_sin_iva = ws.cell(row=res+5, column=total_col, value=f"={total_col_letter}{res+2}/(1-({mb}/100))")
    pvp_sin_iva.number_format = '#,##0.00 €'
    pvp_sin_iva.font = Font(bold=True)

    ws.cell(row=res+6, column=6, value=f"IVA Aplicado ({iva}%):")
    iva_calc = ws.cell(row=res+6, column=total_col, value=f"={total_col_letter}{res+5}*({iva}/100)")
    iva_calc.number_format = '#,##0.00 €'

    ws.cell(row=res+7, column=6, value="PRECIO DE VENTA TOTAL (PVP):").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=res+7, column=6).fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
    
    pvp_total = ws.cell(row=res+7, column=total_col, value=f"={total_col_letter}{res+5}+{total_col_letter}{res+6}")
    pvp_total.number_format = '#,##0.00 €'
    pvp_total.font = Font(bold=True)
    pvp_total.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    ws.cell(row=res+8, column=6, value="PRECIO POR RACIÓN:").font = Font(bold=True)
    precio_racion = ws.cell(row=res+8, column=total_col, value=f"=IF($E$2>0,{total_col_letter}{res+7}/$E$2,0)")
    precio_racion.number_format = '#,##0.00 €'
    precio_racion.font = Font(bold=True)
    precio_racion.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Ajuste automático del ancho de las celdas para que no salgan truncadas (###)
    for col in ws.columns:
        max_len = 0
        col_idx = col[0].column
        col_letter = get_column_letter(col_idx)
        for cell in col:
            if cell.row == 1:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 20)

    virtual_workbook = io.BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)
    return virtual_workbook


# --- INTERFAZ GRÁFICA DE STREAMLIT ---
st.markdown(
    """
    <style>
    :root {
        --samirarte-navy: #14264a;
        --samirarte-muted: #60708d;
        --samirarte-border: #e4e9f2;
        --samirarte-bg: #f7f9fc;
        --samirarte-red: #ff3b3b;
    }
    html, body, [data-testid="stAppViewContainer"] {
        background: var(--samirarte-bg);
        color: var(--samirarte-navy);
    }
    .block-container {
        padding-top: 0;
        padding-bottom: 0.8rem;
        padding-left: 1.35rem;
        padding-right: 1.35rem;
        max-width: 100%;
    }
    hr {
        margin: 0.55rem 0 !important;
    }
    h1 {
        margin-bottom: 0.12rem;
        color: var(--samirarte-navy);
        font-weight: 800;
    }
    h2, h3, h4 {
        margin-top: 0.28rem;
        margin-bottom: 0.22rem;
        color: var(--samirarte-navy);
    }
    h5, h6 {
        margin-top: 0.18rem;
        margin-bottom: 0.16rem;
        color: var(--samirarte-navy);
    }
    p {
        margin-bottom: 0.35rem;
    }
    div[data-testid="stVerticalBlock"] {
        gap: 0.28rem;
    }
    .samirarte-topbar {
        display: flex;
        align-items: center;
        justify-content: space-between;
        min-height: 42px;
        padding: 0 0 0.22rem;
        border-bottom: 1px solid var(--samirarte-border);
        background: rgba(247, 249, 252, 0.92);
    }
    .samirarte-brand {
        display: flex;
        align-items: center;
        gap: 0.8rem;
        min-width: 160px;
    }
    .samirarte-brand img {
        height: 34px;
        object-fit: contain;
    }
    .samirarte-brand-fallback {
        font-weight: 800;
        letter-spacing: 0.02em;
        color: var(--samirarte-navy);
    }
    .samirarte-user {
        color: var(--samirarte-navy);
        font-weight: 700;
        font-size: 0.88rem;
    }
    .samirarte-hero {
        display: flex;
        align-items: center;
        gap: 0.65rem;
        margin: 0.55rem 0 0.45rem;
    }
    .samirarte-hero-icon {
        font-size: 1.72rem;
        line-height: 1;
    }
    .samirarte-hero h1 {
        font-size: clamp(1.55rem, 1.85vw, 2.05rem);
        line-height: 1.02;
        margin: 0;
    }
    .samirarte-hero p {
        color: var(--samirarte-muted);
        margin: 0.08rem 0 0;
        font-size: 0.92rem;
    }
    .samirarte-form-card {
        border: 1px solid var(--samirarte-border);
        border-radius: 10px;
        background: #ffffff;
        padding: 0.55rem 0.82rem 0.64rem;
        box-shadow: 0 10px 24px rgba(20, 38, 74, 0.04);
    }
    .samirarte-section-title {
        color: var(--samirarte-navy);
        font-size: 1.15rem;
        font-weight: 800;
        margin: 0.25rem 0 0.25rem;
    }
    div[data-testid="stMetric"] {
        background: #ffffff;
        border: 1px solid var(--samirarte-border);
        border-radius: 10px;
        padding: 0.5rem 0.78rem;
        min-height: 68px;
        box-shadow: 0 10px 24px rgba(20, 38, 74, 0.04);
    }
    div[data-testid="stMetric"] label {
        white-space: normal;
        color: var(--samirarte-muted);
        font-weight: 700;
    }
    div[data-testid="stMetricValue"] {
        color: var(--samirarte-navy);
        font-size: 1.18rem;
        font-weight: 800;
    }
    div[data-testid="stDataFrame"] {
        border: 1px solid var(--samirarte-border);
        border-radius: 10px;
        overflow: hidden;
        box-shadow: 0 10px 24px rgba(20, 38, 74, 0.04);
        font-size: 1.03rem;
    }
    div[data-testid="stDataFrame"] [role="gridcell"],
    div[data-testid="stDataFrame"] [role="columnheader"] {
        font-size: 1.03rem;
    }
    div[data-testid="stTextInput"] input,
    div[data-testid="stNumberInput"] input,
    div[data-baseweb="select"] {
        min-height: 2.05rem;
    }
    div[data-testid="stTextInput"],
    div[data-testid="stNumberInput"],
    div[data-testid="stSelectbox"] {
        margin-bottom: 0;
    }
    div[data-testid="stTextInput"] input,
    div[data-testid="stNumberInput"] input,
    textarea,
    div[data-baseweb="select"] > div {
        border-color: var(--samirarte-border);
        border-radius: 7px;
    }
    div[data-testid="stTabs"] [data-baseweb="tab-list"] {
        justify-content: center;
        gap: 1.45rem;
        border-bottom: 1px solid var(--samirarte-border);
        background: rgba(247, 249, 252, 0.95);
        position: sticky;
        top: 0;
        z-index: 99;
        padding-top: 0.05rem;
        min-height: 2.35rem;
    }
    div[data-testid="stTabs"] button[role="tab"] {
        color: var(--samirarte-navy);
        font-weight: 700;
        padding: 0.42rem 0.18rem;
    }
    div[data-testid="stTabs"] button[aria-selected="true"] {
        color: var(--samirarte-red);
        border-bottom-color: var(--samirarte-red);
    }
    div[data-testid="stExpander"] details {
        border-radius: 8px;
        border-color: var(--samirarte-border);
        background: #ffffff;
    }
    div[data-testid="stButton"] button,
    div[data-testid="stDownloadButton"] button,
    div[data-testid="stFormSubmitButton"] button {
        border-radius: 7px;
        font-weight: 800;
        min-height: 2.12rem;
        padding-top: 0.28rem;
        padding-bottom: 0.28rem;
    }
    div[data-testid="stAlert"] {
        padding: 0.45rem 0.75rem;
    }
    div[data-testid="stAlert"] p {
        margin-bottom: 0;
    }
    @media (max-width: 720px) {
        .block-container {
            padding-left: 0.85rem;
            padding-right: 0.85rem;
        }
        .samirarte-topbar {
            min-height: 38px;
        }
        .samirarte-brand {
            min-width: 0;
        }
        .samirarte-brand img {
            height: 30px;
        }
        .samirarte-user {
            font-size: 0.8rem;
        }
        .samirarte-hero {
            align-items: flex-start;
            gap: 0.5rem;
            margin: 0.42rem 0 0.35rem;
        }
        .samirarte-hero-icon {
            font-size: 1.45rem;
        }
        .samirarte-hero h1 {
            font-size: 1.42rem;
            line-height: 1.08;
        }
        .samirarte-hero p {
            font-size: 0.88rem;
        }
        div[data-testid="stTabs"] [data-baseweb="tab-list"] {
            gap: 0.75rem;
            overflow-x: auto;
            justify-content: flex-start;
        }
        div[data-testid="stTabs"] button[role="tab"] {
            padding: 0.38rem 0.1rem;
        }
    }
    </style>
    """,
    unsafe_allow_html=True
)

usuario_actual = obtener_usuario_actual()
if not usuario_actual:
    restaurar_sesion_supabase()
    usuario_actual = obtener_usuario_actual()
logo_html = (
    f'<img src="data:image/png;base64,{logo_samirarte_base64()}" alt="Samirarte">'
    if logo_samirarte_existe()
    else '<span class="samirarte-brand-fallback">SAMIRARTE</span>'
)
usuario_topbar = (usuario_actual or {}).get("email") or "Usuario"
st.markdown(
    f"""
    <div class="samirarte-topbar">
        <div class="samirarte-brand">{logo_html}</div>
        <div class="samirarte-user">Usuario · {html.escape(str(usuario_topbar))}</div>
    </div>
    <div class="samirarte-hero">
        <div class="samirarte-hero-icon">👨‍🍳</div>
        <div>
            <h1>Gestor de Escandallos Inteligente Samirarte</h1>
            <p>Gestiona tus recetas y controla costes de forma eficiente</p>
        </div>
    </div>
    """,
    unsafe_allow_html=True
)

# Panel lateral indicador de conexiones activas
with st.sidebar:
    if logo_samirarte_existe():
        st.image(str(LOGO_SAMIRARTE_PATH), width=160)
    else:
        st.caption("Logo no disponible")

    st.header("⚙️ Estado de Conexiones")
    if supabase_disponible:
        st.success("⚡ Inventario: Conectado")
    else:
        st.warning("⚠️ Inventario: Desconectado. Configura las credenciales en los secrets.")
        
    if api_key:
        st.success("🤖 CIA: Activo")
    else:
        st.warning("⚠️ CIA: Desconectado. Configura la clave de IA en los secrets.")

    st.divider()
    st.subheader("Acceso")
    usuario_actual = obtener_usuario_actual()
    auth_feedback = st.session_state.pop("auth_feedback", None)
    if auth_feedback:
        tipo_feedback, mensaje_feedback = auth_feedback
        if tipo_feedback == "success":
            st.success(mensaje_feedback)
        elif tipo_feedback == "warning":
            st.warning(mensaje_feedback)
        else:
            st.error(mensaje_feedback)

    if usuario_actual:
        st.success(f"Sesión iniciada: {usuario_actual.get('email') or 'usuario'}")
        perfil_actual = obtener_perfil_usuario_app()
        if perfil_actual:
            st.caption(f"Rol: {perfil_actual.get('rol', 'usuario')}")
        elif st.session_state.get("usuario_app_error"):
            st.caption(st.session_state.get("usuario_app_error"))
        if st.button("Recargar permisos", use_container_width=True):
            asegurar_usuario_app_supabase()
            st.rerun()
        if st.button("Cerrar sesión", use_container_width=True):
            ok_logout, mensaje_logout = logout_supabase()
            if ok_logout:
                st.session_state["auth_feedback"] = ("success", mensaje_logout)
                st.rerun()
            else:
                st.error(mensaje_logout)

        st.markdown("#### Cambiar contraseña")
        with st.form("form_cambiar_password", clear_on_submit=True):
            nueva_password = st.text_input("Nueva contraseña", type="password")
            repetir_password = st.text_input("Repetir nueva contraseña", type="password")
            actualizar_password = st.form_submit_button("Actualizar contraseña", use_container_width=True)

            if actualizar_password:
                if not obtener_usuario_actual():
                    st.error("La sesión ha caducado. Vuelve a iniciar sesión para cambiar la contraseña.")
                elif not nueva_password:
                    st.error("Introduce una nueva contraseña.")
                elif len(str(nueva_password)) < 6:
                    st.error("La contraseña debe tener al menos 6 caracteres.")
                elif nueva_password != repetir_password:
                    st.error("Las contraseñas no coinciden.")
                else:
                    ok_password, mensaje_password = cambiar_password_usuario(nueva_password)
                    if ok_password:
                        st.session_state["auth_feedback"] = ("success", mensaje_password)
                        st.rerun()
                    else:
                        st.error(mensaje_password)
    else:
        with st.form("form_acceso_supabase"):
            acceso_email = st.text_input("Email")
            acceso_password = st.text_input("Contraseña", type="password")
            login_col, registro_col = st.columns(2)
            with login_col:
                iniciar_sesion = st.form_submit_button("Iniciar sesión", use_container_width=True)
            with registro_col:
                crear_cuenta = st.form_submit_button("Crear cuenta", use_container_width=True)
            if iniciar_sesion:
                ok_login, mensaje_login = login_supabase(acceso_email, acceso_password)
                if ok_login:
                    st.session_state["auth_feedback"] = ("success", mensaje_login)
                    st.rerun()
                else:
                    st.error(mensaje_login)
            elif crear_cuenta:
                ok_registro, mensaje_registro = registrar_usuario_supabase(acceso_email, acceso_password)
                if ok_registro:
                    if obtener_usuario_actual():
                        st.session_state["auth_feedback"] = ("success", mensaje_registro)
                        st.rerun()
                    else:
                        st.success(mensaje_registro)
                else:
                    st.error(mensaje_registro)

    st.divider()
    st.info("💡 Consejo: el inventario común se gestiona desde Administración, disponible solo para usuarios admin.")

    ingredientes_vinculados = [
        (idx, ing, str(ing.get("codigo", "")).strip().upper())
        for idx, ing in enumerate(st.session_state['ingredientes'])
        if str(ing.get("codigo", "")).strip().upper() in inventario_dict
    ]
    if ingredientes_vinculados:
        st.markdown("#### 🧾 Ficha editable del ingrediente seleccionado")
        opciones_ficha = [f"{idx}|{codigo}" for idx, _, codigo in ingredientes_vinculados]
        etiquetas_ficha = {
            f"{idx}|{codigo}": f"{idx + 1}. {codigo} · {inventario_dict[codigo].get('descripcion', '')}"
            for idx, _, codigo in ingredientes_vinculados
        }
        seleccion_ficha = st.selectbox(
            "Ingrediente vinculado a BBDD",
            opciones_ficha,
            format_func=lambda opcion: etiquetas_ficha.get(opcion, opcion),
            key="selector_ficha_ingrediente_activo"
        )
        idx_ficha = int(seleccion_ficha.split("|", 1)[0])
        codigo_ficha = seleccion_ficha.split("|", 1)[1]
        ficha_bd = inventario_dict[codigo_ficha]

        with st.form(f"ficha_inventario_seleccionada_{codigo_ficha}_{idx_ficha}"):
            st.caption(f"Código: {codigo_ficha}")
            descripcion_ficha = st.text_input("Descripción", value=str(ficha_bd.get("descripcion", "")))
            familia_ficha = st.text_input("Familia", value=str(ficha_bd.get("familia", "SIN CLASIFICAR")))
            unidades_validas = ["kg", "l", "ud", "sobre", "botella", "lata", "paquete", "caja", "bandeja", "hoja"]
            unidad_actual = str(ficha_bd.get("unidad_medida", "kg")).strip() or "kg"
            unidad_ficha = st.selectbox(
                "Unidad base",
                unidades_validas,
                index=unidades_validas.index(unidad_actual) if unidad_actual in unidades_validas else 0
            )
            merma_ficha = st.number_input(
                "% Merma",
                min_value=0.0,
                max_value=100.0,
                step=0.01,
                value=float(ficha_bd.get("merma", 0.0))
            )
            precio_ficha = st.number_input(
                "Precio Unidad (€)",
                min_value=0.0,
                step=0.01,
                value=float(ficha_bd.get("precio_unidad", 0.0))
            )

            st.text_input("Proveedor precio", value=str(ficha_bd.get("proveedor_precio", "")), disabled=True)
            st.text_input("Formato compra", value=str(ficha_bd.get("formato_compra", "")), disabled=True)
            st.text_input("Fecha precio", value=str(ficha_bd.get("fecha_precio", "")), disabled=True)
            st.text_input("Cantidad formato", value="" if pd.isna(ficha_bd.get("cantidad_formato_compra", None)) else str(ficha_bd.get("cantidad_formato_compra", "")), disabled=True)
            st.text_input("Unidad formato", value=str(ficha_bd.get("unidad_formato_compra", "")), disabled=True)
            st.text_input("Precio formato", value="" if pd.isna(ficha_bd.get("precio_formato_compra", None)) else str(ficha_bd.get("precio_formato_compra", "")), disabled=True)
            st.text_input("URL precio", value=str(ficha_bd.get("url_precio", "")), disabled=True)
            st.text_area("Observaciones precio", value=str(ficha_bd.get("observaciones_precio", "")), disabled=True, height=80)

            if st.form_submit_button("Guardar cambios en BBDD"):
                if not supabase_disponible:
                    st.error("El inventario no está conectado correctamente.")
                else:
                    datos_ficha = {
                        "codigo": codigo_ficha,
                        "familia": familia_ficha,
                        "descripcion": descripcion_ficha,
                        "unidad_medida": unidad_ficha,
                        "merma": float(merma_ficha),
                        "precio_unidad": float(precio_ficha)
                    }
                    try:
                        supabase.table("inventario").upsert(datos_ficha).execute()
                        st.session_state['ingredientes'][idx_ficha].update({
                            "descripcion": descripcion_ficha,
                            "unidad_medida": unidad_ficha,
                            "merma": float(merma_ficha),
                            "precio_unidad": float(precio_ficha)
                        })
                        st.session_state['db_trigger'] += 1
                        marcar_receta_modificada_manualmente()
                        st.success(f"Ficha {codigo_ficha} actualizada.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error al guardar la ficha en el inventario: {e}")



ADMIN_TABLAS_BBDD = {
    "inventario": {
        "label": "Inventario",
        "pk": "codigo",
        "order": "descripcion",
        "columns": [
            "codigo", "familia", "descripcion", "unidad_medida", "merma", "precio_unidad",
            "proveedor_precio", "formato_compra", "cantidad_formato_compra",
            "unidad_formato_compra", "precio_formato_compra", "fecha_precio",
            "url_precio", "observaciones_precio"
        ],
        "generated_pk": False,
    },
    "recetas": {
        "label": "Recetas",
        "pk": "id",
        "order": "nombre",
        "columns": [
            "id", "user_id", "codigo_receta", "nombre", "categoria", "tipo_plato",
            "raciones_base", "unidad_servicio", "descripcion", "elaboracion",
            "observaciones", "costes_indirectos_pct", "margen_beneficio_pct",
            "iva_pct", "coste_total", "precio_venta_sin_iva", "precio_venta_con_iva",
            "activa", "created_at", "updated_at"
        ],
        "generated_pk": True,
    },
    "receta_ingredientes": {
        "label": "Ingredientes de recetas",
        "pk": "id",
        "order": "orden",
        "columns": [
            "id", "receta_id", "codigo_ingrediente", "descripcion_ingrediente",
            "cantidad_bruta", "unidad_medida", "merma", "cantidad_neta",
            "precio_unidad", "coste_total", "orden", "es_temporal",
            "created_at", "updated_at"
        ],
        "generated_pk": True,
    },
    "menus": {
        "label": "Menús",
        "pk": "id",
        "order": "nombre",
        "columns": [
            "id", "user_id", "nombre", "tipo_menu", "descripcion",
            "numero_comensales", "pax_referencia_menu", "coste_total",
            "precio_total", "created_at", "updated_at"
        ],
        "generated_pk": True,
    },
    "menu_recetas": {
        "label": "Recetas en menús",
        "pk": "id",
        "order": "orden",
        "columns": [
            "id", "menu_id", "receta_id", "raciones",
            "raciones_receta_en_menu", "orden", "seccion",
            "observaciones", "created_at"
        ],
        "generated_pk": True,
    },
    "clientes": {
        "label": "Clientes",
        "pk": "id",
        "order": "nombre",
        "columns": [
            "id", "user_id", "nombre", "tipo_cliente", "nif_cif", "email",
            "telefono", "direccion", "codigo_postal", "ciudad", "provincia",
            "pais", "observaciones", "created_at", "updated_at"
        ],
        "generated_pk": True,
    },
    "facturas": {
        "label": "Facturas y presupuestos",
        "pk": "id",
        "order": "fecha_emision",
        "columns": [
            "id", "user_id", "cliente_id", "numero_factura", "tipo_documento",
            "estado", "fecha_emision", "fecha_vencimiento", "concepto",
            "base_imponible", "iva_pct", "iva_importe", "retencion_pct",
            "retencion_importe", "total", "metodo_pago", "estado_cobro",
            "presupuesto_id", "coste_total", "notas", "created_at", "updated_at"
        ],
        "generated_pk": True,
    },
    "presupuestos": {
        "label": "Presupuestos",
        "pk": "id",
        "order": "fecha_emision",
        "columns": PRESUPUESTOS_COLUMNAS,
        "generated_pk": True,
    },
    "presupuesto_menus": {
        "label": "Menús de presupuestos",
        "pk": "id",
        "order": "orden",
        "columns": PRESUPUESTO_MENUS_COLUMNAS,
        "generated_pk": True,
    },
    "factura_menus": {
        "label": "Menús de facturas",
        "pk": "id",
        "order": "orden",
        "columns": FACTURA_MENUS_COLUMNAS,
        "generated_pk": True,
    },
    "factura_lineas": {
        "label": "Líneas de facturas",
        "pk": "id",
        "order": "orden",
        "columns": [
            "id", "factura_id", "origen_tipo", "origen_id", "descripcion",
            "cantidad", "unidad", "precio_unitario", "descuento_pct",
            "base_linea", "iva_pct", "iva_linea", "total_linea", "orden", "created_at"
        ],
        "generated_pk": True,
    },
    "usuarios_app": {
        "label": "Usuarios de la app",
        "pk": "user_id",
        "order": "email",
        "columns": ["user_id", "email", "nombre", "rol", "activo", "created_at", "updated_at"],
        "generated_pk": False,
        "fixed_rows": True,
    },
}

ADMIN_NUMERIC_COLUMNS = {
    "merma", "precio_unidad", "cantidad_formato_compra", "precio_formato_compra",
    "raciones_base", "costes_indirectos_pct", "margen_beneficio_pct", "iva_pct",
    "coste_total", "precio_venta_sin_iva", "precio_venta_con_iva", "cantidad_bruta",
    "cantidad_neta", "raciones", "raciones_receta_en_menu",
    "numero_comensales", "pax_referencia_menu", "precio_total",
    "base_imponible", "iva_importe", "retencion_pct", "retencion_importe", "total",
    "cantidad", "precio_unitario", "descuento_pct", "base_linea", "iva_linea",
    "total_linea", "cantidad_menu", "coste_total_menu",
    "coste_linea_menu_presupuesto", "precio_linea_menu_presupuesto",
    "coste_linea_menu_factura", "precio_linea_menu_factura", "orden"
}

ADMIN_BOOLEAN_COLUMNS = {"activa", "es_temporal", "activo"}
ADMIN_READONLY_COLUMNS = {"id", "created_at", "updated_at"}


def normalizar_valor_admin(valor, columna):
    if valor is None:
        return None
    try:
        if pd.isna(valor):
            return None
    except TypeError:
        pass
    if columna in ADMIN_BOOLEAN_COLUMNS:
        return bool(valor)
    if columna in ADMIN_NUMERIC_COLUMNS:
        numero = pd.to_numeric(valor, errors="coerce")
        if pd.isna(numero):
            return None
        if columna in {"orden", "numero_comensales"}:
            return int(numero)
        return float(numero)
    texto = str(valor).strip()
    return texto or None


def fila_admin_para_guardar(fila, config, incluir_pk=True):
    datos = {}
    pk = config["pk"]
    for columna in config["columns"]:
        if columna in {"created_at", "updated_at"}:
            continue
        if columna == pk and config.get("generated_pk") and not incluir_pk:
            continue
        valor = fila.get(columna) if isinstance(fila, dict) else getattr(fila, columna, None)
        valor_normalizado = normalizar_valor_admin(valor, columna)
        if columna == pk and config.get("generated_pk") and not valor_normalizado:
            continue
        datos[columna] = valor_normalizado
    return datos


def cargar_tabla_admin(nombre_tabla, limite=500):
    if not supabase_disponible or supabase is None:
        return False, "El inventario no está conectado correctamente.", pd.DataFrame()
    if not usuario_actual_es_admin():
        return False, "Solo los administradores pueden consultar BBDD.", pd.DataFrame()

    config = ADMIN_TABLAS_BBDD[nombre_tabla]
    columnas = config["columns"]
    try:
        consulta = supabase.table(nombre_tabla).select(",".join(columnas)).limit(int(limite))
        if config.get("order"):
            consulta = consulta.order(config["order"])
        respuesta = consulta.execute()
        df = pd.DataFrame(respuesta.data or [])
        if df.empty:
            return True, "No hay registros.", pd.DataFrame(columns=columnas)
        for col in columnas:
            if col not in df.columns:
                df[col] = None
        return True, "Tabla cargada.", df[columnas].copy()
    except Exception as e:
        return False, f"No se pudo cargar {nombre_tabla}: {e}", pd.DataFrame(columns=columnas)


def guardar_cambios_tabla_admin(nombre_tabla, df_original, editor_state):
    config = ADMIN_TABLAS_BBDD[nombre_tabla]
    pk = config["pk"]
    editados = editor_state.get("edited_rows", {}) if editor_state else {}
    anadidos = editor_state.get("added_rows", []) if editor_state else []
    borrados = editor_state.get("deleted_rows", []) if editor_state else []
    cambios = 0

    for row_idx_str, edits in editados.items():
        row_idx = int(row_idx_str)
        fila_original = df_original.iloc[row_idx].to_dict()
        pk_valor = normalizar_valor_admin(fila_original.get(pk), pk)
        if not pk_valor:
            continue
        fila_actualizada = dict(fila_original)
        fila_actualizada.update(edits)
        datos_actualizados = fila_admin_para_guardar(fila_actualizada, config, incluir_pk=False)
        datos_actualizados.pop(pk, None)
        if datos_actualizados:
            supabase.table(nombre_tabla).update(datos_actualizados).eq(pk, pk_valor).execute()
            cambios += 1

    for nueva_fila in anadidos:
        datos_nuevos = fila_admin_para_guardar(nueva_fila, config, incluir_pk=not config.get("generated_pk"))
        datos_nuevos = {k: v for k, v in datos_nuevos.items() if v is not None}
        if datos_nuevos:
            supabase.table(nombre_tabla).insert(datos_nuevos).execute()
            cambios += 1

    for row_idx in borrados:
        fila_original = df_original.iloc[int(row_idx)].to_dict()
        pk_valor = normalizar_valor_admin(fila_original.get(pk), pk)
        if pk_valor:
            supabase.table(nombre_tabla).delete().eq(pk, pk_valor).execute()
            cambios += 1

    return cambios


def render_pagina_administracion():
    st.subheader("Administración del entorno")
    st.caption("Gestión restringida de inventario, usuarios internos y estado de configuración.")

    if not usuario_actual_es_admin():
        st.warning("Esta página solo está disponible para usuarios administradores.")
        error_perfil = st.session_state.get("usuario_app_error")
        if error_perfil:
            st.caption(error_perfil)
        return

    admin_tab_bbdd, admin_tab_estado = st.tabs(["BBDD", "Configuración"])

    with admin_tab_bbdd:
        tabla_opciones = list(ADMIN_TABLAS_BBDD.keys())
        tabla_seleccionada = st.selectbox(
            "Tabla",
            tabla_opciones,
            format_func=lambda tabla: ADMIN_TABLAS_BBDD[tabla]["label"],
            key="admin_tabla_bbdd"
        )
        limite_registros = st.number_input(
            "Registros a cargar",
            min_value=50,
            max_value=2000,
            value=500,
            step=50,
            key="admin_limite_bbdd"
        )
        config_tabla = ADMIN_TABLAS_BBDD[tabla_seleccionada]
        ok_tabla, mensaje_tabla, tabla_df = cargar_tabla_admin(tabla_seleccionada, limite_registros)

        if not ok_tabla:
            st.warning(mensaje_tabla)
        else:
            st.caption(f"{len(tabla_df)} registros cargados de public.{tabla_seleccionada}")
            column_config = {}
            for columna in config_tabla["columns"]:
                if columna in ADMIN_BOOLEAN_COLUMNS:
                    column_config[columna] = st.column_config.CheckboxColumn(columna)
                elif columna == "rol":
                    column_config[columna] = st.column_config.SelectboxColumn(columna, options=["usuario", "admin"])
                elif columna in ADMIN_NUMERIC_COLUMNS:
                    column_config[columna] = st.column_config.NumberColumn(columna)
                else:
                    column_config[columna] = st.column_config.TextColumn(columna)

            columnas_bloqueadas = [
                col for col in config_tabla["columns"]
                if col in ADMIN_READONLY_COLUMNS
                or (col == config_tabla["pk"] and config_tabla.get("fixed_rows"))
            ]
            editor_key = f"admin_editor_{tabla_seleccionada}"
            st.data_editor(
                tabla_df,
                num_rows="fixed" if config_tabla.get("fixed_rows") else "dynamic",
                use_container_width=True,
                height=520,
                column_config=column_config,
                disabled=columnas_bloqueadas,
                key=editor_key
            )

            acciones_col1, acciones_col2 = st.columns([1, 3])
            with acciones_col1:
                if st.button("Guardar cambios", type="primary", use_container_width=True, key=f"guardar_{tabla_seleccionada}"):
                    editor_state = st.session_state.get(editor_key)
                    try:
                        cambios = guardar_cambios_tabla_admin(tabla_seleccionada, tabla_df, editor_state)
                        if tabla_seleccionada == "usuarios_app":
                            asegurar_usuario_app_supabase()
                        if tabla_seleccionada == "inventario":
                            st.session_state["db_trigger"] += 1
                        st.success(f"Cambios guardados: {cambios}")
                        st.rerun()
                    except Exception as e:
                        st.error(f"No se pudieron guardar los cambios en {tabla_seleccionada}: {e}")
            with acciones_col2:
                st.caption("Las altas en tablas relacionadas necesitan claves foráneas válidas. Los borrados se aplican directamente al guardar.")

    with admin_tab_estado:
        st.subheader("Configuración y estado")
        estado_col1, estado_col2, estado_col3 = st.columns(3)
        with estado_col1:
            st.metric("Inventario", "Conectado" if supabase_disponible else "Desconectado")
        with estado_col2:
            st.metric("CIA", "Activa" if api_key else "Desconectada")
        with estado_col3:
            st.metric("Ingredientes en BBDD", len(inventario_df) if not inventario_df.empty else 0)

        perfil = obtener_perfil_usuario_app()
        if perfil:
            st.write(f"**Usuario actual:** {perfil.get('email', '')}")
            st.write(f"**Rol:** {perfil.get('rol', 'usuario')}")
            st.write(f"**Activo:** {'sí' if perfil.get('activo', True) else 'no'}")

        st.markdown("#### Migraciones necesarias")
        st.code(
            "sql/017_usuarios_app_entorno.sql\n"
            "sql/018_admin_rls_todas_bbdd.sql\n"
            "sql/019_jerarquia_recetas_menus_presupuestos_facturas.sql",
            language="text",
        )
        st.caption(
            "La migración 019 añade la jerarquía de presupuestos y facturas. "
            "Revísala y ejecútala manualmente en Supabase."
        )

pagina_actual = "App"
if usuario_actual_es_admin():
    with st.sidebar:
        st.divider()
        pagina_actual = st.radio(
            "Navegación",
            ["App", "Administración"],
            horizontal=False,
            key="pagina_principal"
        )

if pagina_actual == "Administración":
    render_pagina_administracion()
    st.stop()

main_tab_recetas, main_tab_menus, main_tab_clientes, main_tab_facturas = st.tabs([
    "Recetas",
    "Menús",
    "Clientes",
    "Facturas"
])

with main_tab_recetas:
    sincronizar_inputs_raciones()
    aviso_raciones_ia = st.session_state.pop("aviso_raciones_ia", None)
    if aviso_raciones_ia:
        st.warning(aviso_raciones_ia)
    aviso_receta_antigua = st.session_state.pop("aviso_receta_antigua", None)
    if aviso_receta_antigua:
        st.warning(aviso_receta_antigua)

    st.markdown('<div class="samirarte-form-card">', unsafe_allow_html=True)
    cab_col1, cab_col2, cab_col3 = st.columns([2.3, 1.15, 1.15])
    with cab_col1:
        nombre_plato = st.text_input("Nombre de la receta", key="input_nombre_plato")
    with cab_col2:
        categoria_actual = st.text_input("Categoría", key="input_receta_categoria", placeholder="Categoría")
    with cab_col3:
        tipo_plato_actual = st.text_input("Tipo de plato", key="input_receta_tipo_plato", placeholder="Tipo")
    st.markdown('</div>', unsafe_allow_html=True)
    st.session_state['receta_nombre'] = nombre_plato
    st.session_state['nombre_plato'] = nombre_plato
    st.session_state["receta_categoria"] = categoria_actual
    st.session_state["receta_tipo_plato"] = tipo_plato_actual

    raciones_base_actual = numero_seguro(
        st.session_state.get(
            "receta_raciones_base",
            st.session_state.get("raciones_base", 1.0),
        ),
        1.0,
    )
    if raciones_base_actual <= 0:
        raciones_base_actual = 1.0
    st.session_state['raciones_base'] = raciones_base_actual
    st.session_state['raciones_deseadas'] = raciones_base_actual
    st.session_state['receta_raciones_deseadas'] = raciones_base_actual
    st.session_state['factor_raciones'] = 1.0
    st.session_state['raciones_base_aplicadas'] = raciones_base_actual
    st.session_state['raciones_deseadas_aplicadas'] = raciones_base_actual

    subtotal_cabecera = sum(
        float(ing.get('cantidad_bruta', 0.0)) * float(ing.get('precio_unidad', 0.0))
        for ing in st.session_state.get('ingredientes', [])
    )
    ci_cabecera = float(st.session_state.get('costes_indirectos_pct', 0.0) or 0.0)
    mb_cabecera = float(st.session_state.get('margen_beneficio_pct', 0.0) or 0.0)
    iva_cabecera = float(st.session_state.get('iva_pct', 0.0) or 0.0)
    coste_cabecera = subtotal_cabecera
    coste_para_precio_cabecera = coste_cabecera + (
        coste_cabecera * (ci_cabecera / 100)
    )
    factor_margen_cabecera = 1 - (mb_cabecera / 100)
    pvp_neto_cabecera = (
        coste_para_precio_cabecera / factor_margen_cabecera
        if factor_margen_cabecera > 0 else 0.0
    )
    pvp_final_cabecera = pvp_neto_cabecera + (pvp_neto_cabecera * (iva_cabecera / 100))
    food_cost_cabecera = (coste_cabecera / pvp_final_cabecera * 100) if pvp_final_cabecera > 0 else 0.0

    coste_por_racion_cabecera = calcular_coste_por_racion(
        coste_cabecera,
        raciones_base_actual,
    )
    pvp_por_racion_cabecera = pvp_final_cabecera / raciones_base_actual
    met_col1, met_col2, met_col3, met_col4 = st.columns(4)
    with met_col1:
        if "input_raciones_base_receta" not in st.session_state:
            st.session_state["input_raciones_base_receta"] = float(raciones_base_actual)
        raciones_base_actual = st.number_input(
            "Raciones base receta",
            min_value=0.001,
            step=1.0,
            key="input_raciones_base_receta",
            on_change=normalizar_raciones_desde_campo_receta,
            help="Introduce las raciones de la receta original. Las cantidades se dividirán automáticamente y el valor volverá a 1.",
        )
        st.session_state["receta_raciones_base"] = float(raciones_base_actual)
        st.session_state["raciones_base"] = float(raciones_base_actual)
    met_col2.metric("Coste por ración", f"{coste_por_racion_cabecera:.2f} €")
    met_col3.metric("PVP por ración", f"{pvp_por_racion_cabecera:.2f} €" if pvp_final_cabecera else "-")
    met_col4.metric("Food Cost", f"{food_cost_cabecera:.1f} %" if food_cost_cabecera else "-")

    st.markdown('<div class="samirarte-section-title">Ingredientes</div>', unsafe_allow_html=True)
    # La tabla es el único punto de alta, edición y eliminación de ingredientes.
    # Se muestra incluso vacía para que la primera fila también pueda añadirse aquí.
    columnas_ingrediente = [
        "codigo", "descripcion", "cantidad_bruta",
        "unidad_medida", "merma", "precio_unidad",
    ]
    df_display = pd.DataFrame(st.session_state.get("ingredientes", []))
    for col in columnas_ingrediente:
        if col not in df_display.columns:
            df_display[col] = (
                0.0 if col in ["cantidad_bruta", "merma", "precio_unidad"]
                else ("kg" if col == "unidad_medida" else "")
            )
    df_display["codigo"] = df_display["codigo"].fillna("S/C").astype(str).str.strip().replace("", "S/C")
    df_display["descripcion"] = df_display["descripcion"].fillna("").astype(str)
    df_display["unidad_medida"] = (
        df_display["unidad_medida"].fillna("kg").astype(str).str.strip().replace({"": "kg", "S/C": "kg"})
    )
    for col in ["cantidad_bruta", "merma", "precio_unidad"]:
        df_display[col] = pd.to_numeric(df_display[col], errors="coerce").fillna(0.0)
    df_display["coste_total"] = df_display["cantidad_bruta"] * df_display["precio_unidad"]
    df_display["eliminar"] = False
    df_display = df_display[
        ["eliminar", "descripcion", "cantidad_bruta", "unidad_medida", "merma", "precio_unidad", "coste_total", "codigo"]
    ]
    codigos_editor = ["S/C"] + list(inventario_dict.keys())
    for codigo_existente in df_display["codigo"].dropna().astype(str).str.strip().str.upper():
        if codigo_existente and codigo_existente not in codigos_editor:
            codigos_editor.append(codigo_existente)

    st.caption(
        "Añade filas con el control de la tabla. Puedes editar cualquier ingrediente y marcar «Eliminar» "
        "para quitarlo de la receta, esté o no en inventario."
    )
    ingredientes_editados = st.data_editor(
        df_display,
        num_rows="dynamic",
        use_container_width=True,
        height=560,
        column_config={
            "eliminar": st.column_config.CheckboxColumn("Eliminar", default=False, width="small"),
            "descripcion": st.column_config.TextColumn(
                "Ingrediente",
                help="Nombre editable del ingrediente, exista o no en inventario.",
                width="large",
            ),
            "cantidad_bruta": st.column_config.NumberColumn(
                "Cantidad", format="%.3f", min_value=0.0, default=0.0, width="small"
            ),
            "unidad_medida": st.column_config.SelectboxColumn(
                "Unidad",
                options=["kg", "l", "ud", "sobre", "botella", "lata", "paquete", "caja", "bandeja", "hoja"],
                default="kg",
                width="small",
            ),
            "merma": st.column_config.NumberColumn(
                "Merma %", format="%.2f %%", min_value=0.0, max_value=100.0, default=0.0, width="small"
            ),
            "precio_unidad": st.column_config.NumberColumn(
                "Precio unidad", format="%.2f €", min_value=0.0, default=0.0, width="medium"
            ),
            "coste_total": st.column_config.NumberColumn(
                "Coste", format="%.2f €", min_value=0.0, width="small"
            ),
            "codigo": st.column_config.SelectboxColumn(
                "Código",
                help="Selecciona un ingrediente del inventario o usa S/C para una fila manual.",
                options=codigos_editor,
                format_func=lambda codigo: (
                    f"{codigo} · {inventario_dict[codigo]['descripcion']}"
                    if codigo in inventario_dict
                    else (
                        "S/C · No encontrado en inventario"
                        if codigo == "S/C" else f"{codigo} · No encontrado en inventario"
                    )
                ),
                default="S/C",
                width="small",
            ),
        },
        disabled=["coste_total"],
        key="editor_receta_activa",
    )

    ingredientes_editados = ingredientes_editados[
        ~ingredientes_editados["eliminar"].fillna(False).astype(bool)
    ].copy()
    ingredientes_editados["descripcion"] = (
        ingredientes_editados["descripcion"].fillna("").astype(str).str.strip()
    )
    ingredientes_editados["codigo"] = (
        ingredientes_editados["codigo"].fillna("S/C").astype(str).str.strip().str.upper().replace("", "S/C")
    )
    ingredientes_originales = st.session_state.get("ingredientes", [])
    for idx in ingredientes_editados.index:
        codigo_actual = ingredientes_editados.at[idx, "codigo"]
        try:
            indice_original = int(idx)
        except (TypeError, ValueError):
            indice_original = None
        codigo_anterior = (
            str(ingredientes_originales[indice_original].get("codigo", "S/C")).strip().upper()
            if indice_original is not None and 0 <= indice_original < len(ingredientes_originales)
            else "S/C"
        )
        if codigo_actual in inventario_dict and codigo_actual != codigo_anterior:
            info_bd = inventario_dict[codigo_actual]
            ingredientes_editados.at[idx, "descripcion"] = info_bd.get("descripcion", "")
            ingredientes_editados.at[idx, "unidad_medida"] = info_bd.get("unidad_medida", "kg")
            ingredientes_editados.at[idx, "merma"] = info_bd.get("merma", 0.0)
            ingredientes_editados.at[idx, "precio_unidad"] = info_bd.get("precio_unidad", 0.0)

    # Una fila recién creada no entra en la receta hasta que tenga nombre.
    ingredientes_editados = ingredientes_editados[
        ingredientes_editados["descripcion"] != ""
    ].copy()
    ingredientes_editados["unidad_medida"] = (
        ingredientes_editados["unidad_medida"].fillna("kg").astype(str).str.strip().replace("", "kg")
    )
    for col in ["cantidad_bruta", "merma", "precio_unidad"]:
        ingredientes_editados[col] = pd.to_numeric(
            ingredientes_editados[col], errors="coerce"
        ).fillna(0.0)

    ingredientes_lista = ingredientes_editados.drop(
        columns=["eliminar", "coste_total"],
        errors="ignore",
    ).to_dict(orient="records")
    if ingredientes_lista != st.session_state.get("ingredientes", []):
        st.session_state["ingredientes"] = ingredientes_lista
        marcar_receta_modificada_manualmente()
        st.rerun()

    if st.session_state.get("ingredientes"):
        with st.expander("Acciones secundarias de ingredientes", expanded=False):
            acciones_db_col1, acciones_db_col2, acciones_db_col3 = st.columns([1, 1, 1])
            with acciones_db_col1:
                if st.button("Actualizar BBDD con códigos existentes", use_container_width=True):
                    if not supabase_disponible:
                        st.error("El inventario no está conectado correctamente.")
                    else:
                        filas_validas = [
                            preparar_fila_inventario_desde_ingrediente(ing)
                            for ing in st.session_state['ingredientes']
                            if codigo_ingrediente_valido(ing.get("codigo")) and str(ing.get("codigo", "")).strip().upper() in inventario_dict
                        ]
                        if filas_validas:
                            try:
                                for fila in filas_validas:
                                    supabase.table("inventario").upsert(fila).execute()
                                st.session_state['db_trigger'] += 1
                                st.success(f"{len(filas_validas)} ingredientes actualizados en el inventario.")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Error al actualizar el inventario: {e}")
                        else:
                            st.info("No hay ingredientes con código existente para actualizar.")

            with acciones_db_col2:
                if st.button("Crear nuevos en BBDD", use_container_width=True):
                    if not supabase_disponible:
                        st.error("El inventario no está conectado correctamente.")
                    else:
                        existentes = set(inventario_dict.keys())
                        ingredientes_actualizados = [dict(ing) for ing in st.session_state['ingredientes']]
                        filas_nuevas = []
                        for idx, ing in enumerate(ingredientes_actualizados):
                            codigo_actual = str(ing.get("codigo", "")).strip().upper()
                            if not codigo_ingrediente_valido(codigo_actual) or codigo_actual not in inventario_dict:
                                nuevo_codigo = codigo_actual if codigo_ingrediente_valido(codigo_actual) else generar_codigo_ingrediente_nuevo(ing.get("descripcion", ""), existentes)
                                existentes.add(nuevo_codigo)
                                ing["codigo"] = nuevo_codigo
                                filas_nuevas.append(preparar_fila_inventario_desde_ingrediente(ing, codigo=nuevo_codigo))
                        if filas_nuevas:
                            try:
                                for fila in filas_nuevas:
                                    supabase.table("inventario").upsert(fila).execute()
                                st.session_state['ingredientes'] = ingredientes_actualizados
                                st.session_state['db_trigger'] += 1
                                marcar_receta_modificada_manualmente()
                                st.success(f"{len(filas_nuevas)} ingredientes nuevos creados en el inventario.")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Error al crear ingredientes en el inventario: {e}")
                        else:
                            st.info("No hay ingredientes nuevos pendientes de crear.")

            with acciones_db_col3:
                if st.button("Limpiar toda la receta", use_container_width=True):
                    st.session_state['ingredientes'] = []
                    st.session_state['ingredientes_base_raciones'] = None
                    st.session_state['factor_raciones'] = 1.0
                    st.session_state['receta_id_cargada'] = None
                    st.session_state['codigo_receta_cargada'] = ""
                    st.rerun()

    with st.expander("IA: Generar receta con IA", expanded=False):
        texto_tab, imagen_tab = st.tabs(["Texto", "Imagen"])
        with texto_tab:
            texto_pegado = st.text_area("Texto para analizar", height=120, placeholder="3 kg de pollo asado ING-0027...")
            if st.button("Analizar texto con IA", type="primary"):
                if texto_pegado:
                    with st.spinner("La IA está leyendo y cruzando los datos con tu inventario..."):
                        nuevos = procesar_con_openai(texto_plano=texto_pegado)
                        if incorporar_ingredientes_ia(nuevos):
                            st.rerun()

        with imagen_tab:
            archivo_imagen = st.file_uploader("Foto de receta, factura o albarán (JPG/PNG)", type=['jpg', 'jpeg', 'png'])
            if archivo_imagen:
                if st.button("Escanear imagen con IA Vision", type="primary"):
                    bytes_img = archivo_imagen.read()
                    with st.spinner("Leyendo factura y asociando códigos de inventario..."):
                        nuevos = procesar_con_openai(bytes_imagen=bytes_img, mime_type=archivo_imagen.type)
                        if incorporar_ingredientes_ia(nuevos):
                            st.rerun()


with main_tab_clientes:
    st.subheader("Clientes")

    usuario_id_facturacion = obtener_user_id_actual()
    if not usuario_id_facturacion:
        st.info("Inicia sesión para guardar clientes en tu cuenta")
    else:
        ok_clientes, mensaje_clientes, clientes_df = cargar_clientes_supabase()
        if not ok_clientes:
            st.warning(mensaje_clientes)
        else:
            if clientes_df.empty:
                st.info("Todavía no hay clientes guardados en tu cuenta.")
            else:
                clientes_resumen = clientes_df[[
                    "nombre", "tipo_cliente", "nif_cif", "email", "telefono",
                    "ciudad", "provincia", "pais"
                ]].fillna("")
                st.dataframe(clientes_resumen, use_container_width=True, hide_index=True)

            with st.form("form_cliente_nuevo"):
                st.markdown("#### Nuevo cliente")
                nuevo_col1, nuevo_col2 = st.columns(2)
                with nuevo_col1:
                    nuevo_nombre = st.text_input("Nombre", key="cliente_nuevo_nombre")
                    nuevo_tipo = st.text_input("Tipo cliente", key="cliente_nuevo_tipo")
                    nuevo_nif = st.text_input("NIF/CIF", key="cliente_nuevo_nif")
                    nuevo_email = st.text_input("Email", key="cliente_nuevo_email")
                    nuevo_telefono = st.text_input("Teléfono", key="cliente_nuevo_telefono")
                with nuevo_col2:
                    nuevo_direccion = st.text_input("Dirección", key="cliente_nuevo_direccion")
                    nuevo_cp = st.text_input("Código postal", key="cliente_nuevo_cp")
                    nuevo_ciudad = st.text_input("Ciudad", key="cliente_nuevo_ciudad")
                    nuevo_provincia = st.text_input("Provincia", key="cliente_nuevo_provincia")
                    nuevo_pais = st.text_input("País", value="España", key="cliente_nuevo_pais")
                nuevo_observaciones = st.text_area("Observaciones", key="cliente_nuevo_observaciones", height=90)
                guardar_cliente = st.form_submit_button("Guardar cliente", type="primary", use_container_width=True)

                if guardar_cliente:
                    datos_cliente = datos_cliente_desde_formulario(
                        nuevo_nombre, nuevo_tipo, nuevo_nif, nuevo_email, nuevo_telefono,
                        nuevo_direccion, nuevo_cp, nuevo_ciudad, nuevo_provincia,
                        nuevo_pais, nuevo_observaciones
                    )
                    ok_guardar, mensaje_guardar, _ = guardar_cliente_supabase(datos_cliente)
                    if ok_guardar:
                        st.success(mensaje_guardar)
                        st.rerun()
                    else:
                        st.error(mensaje_guardar)

            if not clientes_df.empty:
                st.divider()
                st.markdown("#### Editar cliente")
                clientes_por_id = {
                    str(row.get("id")): row.to_dict()
                    for _, row in clientes_df.iterrows()
                    if row.get("id")
                }
                opciones_clientes = list(clientes_por_id.keys())
                cliente_id_seleccionado = st.selectbox(
                    "Clientes guardados",
                    opciones_clientes,
                    format_func=lambda cliente_id: clientes_por_id.get(cliente_id, {}).get("nombre", cliente_id),
                    key="selector_cliente_facturacion"
                )
                cliente_seleccionado = clientes_por_id.get(str(cliente_id_seleccionado), {})

                def valor_cliente(campo, defecto=""):
                    valor = cliente_seleccionado.get(campo, defecto)
                    if valor is None:
                        return ""
                    try:
                        if pd.isna(valor):
                            return ""
                    except TypeError:
                        pass
                    return str(valor)

                with st.form("form_cliente_editar"):
                    editar_col1, editar_col2 = st.columns(2)
                    with editar_col1:
                        editar_nombre = st.text_input("Nombre", value=valor_cliente("nombre"), key=f"cliente_editar_nombre_{cliente_id_seleccionado}")
                        editar_tipo = st.text_input("Tipo cliente", value=valor_cliente("tipo_cliente"), key=f"cliente_editar_tipo_{cliente_id_seleccionado}")
                        editar_nif = st.text_input("NIF/CIF", value=valor_cliente("nif_cif"), key=f"cliente_editar_nif_{cliente_id_seleccionado}")
                        editar_email = st.text_input("Email", value=valor_cliente("email"), key=f"cliente_editar_email_{cliente_id_seleccionado}")
                        editar_telefono = st.text_input("Teléfono", value=valor_cliente("telefono"), key=f"cliente_editar_telefono_{cliente_id_seleccionado}")
                    with editar_col2:
                        editar_direccion = st.text_input("Dirección", value=valor_cliente("direccion"), key=f"cliente_editar_direccion_{cliente_id_seleccionado}")
                        editar_cp = st.text_input("Código postal", value=valor_cliente("codigo_postal"), key=f"cliente_editar_cp_{cliente_id_seleccionado}")
                        editar_ciudad = st.text_input("Ciudad", value=valor_cliente("ciudad"), key=f"cliente_editar_ciudad_{cliente_id_seleccionado}")
                        editar_provincia = st.text_input("Provincia", value=valor_cliente("provincia"), key=f"cliente_editar_provincia_{cliente_id_seleccionado}")
                        editar_pais = st.text_input("País", value=valor_cliente("pais", "España") or "España", key=f"cliente_editar_pais_{cliente_id_seleccionado}")
                    editar_observaciones = st.text_area(
                        "Observaciones",
                        value=valor_cliente("observaciones"),
                        key=f"cliente_editar_observaciones_{cliente_id_seleccionado}",
                        height=90
                    )
                    actualizar_cliente = st.form_submit_button("Actualizar cliente", use_container_width=True)

                    if actualizar_cliente:
                        datos_cliente = datos_cliente_desde_formulario(
                            editar_nombre, editar_tipo, editar_nif, editar_email, editar_telefono,
                            editar_direccion, editar_cp, editar_ciudad, editar_provincia,
                            editar_pais, editar_observaciones
                        )
                        ok_actualizar, mensaje_actualizar = actualizar_cliente_supabase(
                            cliente_id_seleccionado,
                            datos_cliente
                        )
                        if ok_actualizar:
                            st.success(mensaje_actualizar)
                            st.rerun()
                        else:
                            st.error(mensaje_actualizar)

                st.markdown("#### Eliminar cliente")
                st.warning("Esta acción no se puede deshacer. No elimines clientes con documentación asociada.")
                confirmar_eliminar_cliente = st.text_input(
                    "Escribe ELIMINAR para confirmar",
                    key=f"confirmar_eliminar_cliente_{cliente_id_seleccionado}"
                )
                if st.button("Eliminar cliente seleccionado", type="secondary", use_container_width=True):
                    if confirmar_eliminar_cliente != "ELIMINAR":
                        st.error("Para eliminar el cliente debes escribir exactamente ELIMINAR.")
                    else:
                        ok_eliminar, mensaje_eliminar = eliminar_cliente_supabase(cliente_id_seleccionado)
                        if ok_eliminar:
                            st.success(mensaje_eliminar)
                            st.rerun()
                        else:
                            st.error(mensaje_eliminar)



with main_tab_facturas:
    st.subheader("Facturas")
    usuario_id_facturas = obtener_user_id_actual()
    if not usuario_id_facturas:
        st.info("Inicia sesión para crear presupuestos y facturas en tu cuenta")
    else:
        ok_clientes_facturas, mensaje_clientes_facturas, clientes_df = cargar_clientes_supabase()
        if not ok_clientes_facturas:
            st.warning(mensaje_clientes_facturas)
        else:
            st.markdown("#### Presupuestos y facturas")
            st.info("Módulo de facturación interna. No usar como sistema fiscal definitivo sin revisión legal/fiscal.")
            feedback_menu_factura = st.session_state.pop("factura_menu_feedback", None)
            if feedback_menu_factura:
                mensaje_menu_factura, aviso_precio_menu_feedback = feedback_menu_factura
                st.success(mensaje_menu_factura)
                if aviso_precio_menu_feedback:
                    st.warning("El menú no tiene precio de venta completo. Se ha usado el coste como referencia.")

            ok_facturas, mensaje_facturas, facturas_df = cargar_facturas_supabase()
            if not ok_facturas:
                st.warning(mensaje_facturas)
            elif clientes_df.empty:
                st.warning("Añade al menos un cliente antes de crear presupuestos o facturas.")
            else:
                clientes_por_id_factura = {
                    str(row.get("id")): row.to_dict()
                    for _, row in clientes_df.iterrows()
                    if row.get("id")
                }
                opciones_clientes_factura = list(clientes_por_id_factura.keys())

                nombres_clientes_factura = {
                    cliente_id: datos.get("nombre", "")
                    for cliente_id, datos in clientes_por_id_factura.items()
                }

                st.markdown("##### Documentos guardados")
                if facturas_df.empty:
                    st.info("Todavía no hay presupuestos ni facturas guardados.")
                else:
                    facturas_selector_df = facturas_df.copy()
                    facturas_selector_df["cliente"] = facturas_selector_df["cliente_id"].apply(
                        lambda cliente_id: nombres_clientes_factura.get(str(cliente_id), "")
                    )
                    facturas_por_id = {
                        str(row.get("id")): row.to_dict()
                        for _, row in facturas_selector_df.iterrows()
                        if row.get("id")
                    }
                    opciones_facturas = list(facturas_por_id.keys())
                    factura_guardada_id = st.selectbox(
                        "Documentos del usuario conectado",
                        opciones_facturas,
                        format_func=lambda factura_id: (
                            f"{facturas_por_id.get(factura_id, {}).get('numero_factura', '')} · "
                            f"{facturas_por_id.get(factura_id, {}).get('tipo_documento', '')} · "
                            f"{facturas_por_id.get(factura_id, {}).get('cliente', '')} · "
                            f"{facturas_por_id.get(factura_id, {}).get('fecha_emision', '')} · "
                            f"{facturas_por_id.get(factura_id, {}).get('estado', '')} · "
                            f"{numero_seguro(facturas_por_id.get(factura_id, {}).get('total'), 0.0):.2f} €"
                        ),
                        key="selector_factura_guardada"
                    )
                    ok_origen_doc, _, _, lineas_origen_doc = cargar_factura_detalle_supabase(factura_guardada_id)
                    if ok_origen_doc and not lineas_origen_doc.empty:
                        menu_id_documento = obtener_menu_id_desde_lineas_factura(lineas_origen_doc.to_dict(orient="records"))
                        if menu_id_documento:
                            nombre_menu_documento = obtener_nombre_menu_por_id(menu_id_documento)
                            etiqueta_menu_documento = (
                                f"Documento creado desde menú: {nombre_menu_documento}"
                                if nombre_menu_documento else
                                "Documento creado desde menú"
                            )
                            st.info(etiqueta_menu_documento)
                    cargar_col, duplicar_col, facturar_col = st.columns(3)
                    with cargar_col:
                        if st.button("Cargar documento", use_container_width=True):
                            ok_detalle, mensaje_detalle, factura_detalle, lineas_detalle = cargar_factura_detalle_supabase(
                                factura_guardada_id
                            )
                            if ok_detalle:
                                cargar_factura_en_sesion(factura_detalle, lineas_detalle)
                                st.session_state["factura_documento_feedback"] = ("success", mensaje_detalle)
                                st.rerun()
                            else:
                                st.error(mensaje_detalle)
                    with duplicar_col:
                        if st.button("Duplicar documento", use_container_width=True):
                            ok_duplicar, mensaje_duplicar, factura_duplicada = duplicar_factura_supabase(
                                factura_guardada_id
                            )
                            if ok_duplicar:
                                st.session_state["factura_documento_feedback"] = (
                                    "success",
                                    f"{mensaje_duplicar} {factura_duplicada.get('numero_factura', '')}"
                                )
                                st.rerun()
                            else:
                                st.error(mensaje_duplicar)
                    with facturar_col:
                        es_presupuesto_seleccionado = (
                            facturas_por_id
                            .get(str(factura_guardada_id), {})
                            .get("tipo_documento") == "presupuesto"
                        )
                        if st.button(
                            "Crear factura desde presupuesto",
                            disabled=not es_presupuesto_seleccionado,
                            use_container_width=True,
                        ):
                            ok_detalle, mensaje_detalle, presupuesto_detalle, lineas_detalle = cargar_factura_detalle_supabase(
                                factura_guardada_id
                            )
                            if ok_detalle and not lineas_detalle.empty:
                                cargar_factura_en_sesion(presupuesto_detalle, lineas_detalle)
                                st.session_state["factura_id_en_edicion"] = None
                                st.session_state["factura_presupuesto_origen_id"] = factura_guardada_id
                                st.session_state["factura_tipo_documento"] = "factura"
                                st.session_state["factura_numero_actual_factura"] = generar_numero_factura("factura")
                                st.session_state["factura_documento_feedback"] = (
                                    "success",
                                    "Factura creada desde el presupuesto con líneas e importes congelados.",
                                )
                                st.rerun()
                            else:
                                st.error(mensaje_detalle or "El presupuesto no contiene menús.")

                    columnas_documentos = [
                        "numero_factura", "tipo_documento", "cliente",
                        "fecha_emision", "estado", "total"
                    ]
                    st.dataframe(
                        facturas_selector_df[columnas_documentos].rename(columns={
                            "numero_factura": "número",
                            "tipo_documento": "tipo",
                            "fecha_emision": "fecha"
                        }).fillna(""),
                        use_container_width=True,
                        hide_index=True
                    )
                    st.caption("La eliminación de documentos se añadirá más adelante con confirmación reforzada.")

                feedback_documento_factura = st.session_state.pop("factura_documento_feedback", None)
                if feedback_documento_factura:
                    tipo_feedback, mensaje_feedback = feedback_documento_factura
                    if tipo_feedback == "success":
                        st.success(mensaje_feedback)
                    else:
                        st.warning(mensaje_feedback)

                if st.session_state.get("factura_cliente_id") not in opciones_clientes_factura:
                    st.session_state["factura_cliente_id"] = opciones_clientes_factura[0]
                opciones_tipo_documento = ["presupuesto", "factura", "factura_rectificativa"]
                opciones_estado_documento = ["borrador", "emitida", "cobrada", "anulada"]
                opciones_estado_cobro = ["pendiente", "parcial", "cobrado"]
                if st.session_state.get("factura_tipo_documento") not in opciones_tipo_documento:
                    st.session_state["factura_tipo_documento"] = "presupuesto"
                if st.session_state.get("factura_estado") not in opciones_estado_documento:
                    st.session_state["factura_estado"] = "borrador"
                if st.session_state.get("factura_estado_cobro") not in opciones_estado_cobro:
                    st.session_state["factura_estado_cobro"] = "pendiente"

                factura_id_en_edicion = st.session_state.get("factura_id_en_edicion")
                st.markdown("##### Documento actual")
                if factura_id_en_edicion:
                    st.info(f"Editando documento guardado `{factura_id_en_edicion}`.")
                else:
                    st.caption("Crea un documento nuevo o carga uno guardado para editarlo.")
                doc_col1, doc_col2, doc_col3 = st.columns(3)
                with doc_col1:
                    factura_cliente_id = st.selectbox(
                        "Cliente",
                        opciones_clientes_factura,
                        format_func=lambda cliente_id: clientes_por_id_factura.get(cliente_id, {}).get("nombre", cliente_id),
                        key="factura_cliente_id"
                    )
                    tipo_documento = st.selectbox(
                        "Tipo de documento",
                        opciones_tipo_documento,
                        key="factura_tipo_documento"
                    )
                    numero_sugerido = generar_numero_factura(tipo_documento)
                    numero_factura = st.text_input(
                        "Número de factura/documento",
                        value=numero_sugerido,
                        key=f"factura_numero_actual_{tipo_documento}"
                    )
                with doc_col2:
                    fecha_emision = st.date_input("Fecha de emisión", key="factura_fecha_emision")
                    fecha_vencimiento = st.date_input("Fecha de vencimiento", key="factura_fecha_vencimiento")
                    estado = st.selectbox(
                        "Estado",
                        opciones_estado_documento,
                        key="factura_estado"
                    )
                with doc_col3:
                    metodo_pago = st.text_input("Método de pago", key="factura_metodo_pago")
                    estado_cobro = st.selectbox(
                        "Estado de cobro",
                        opciones_estado_cobro,
                        key="factura_estado_cobro"
                    )
                    retencion_pct = st.number_input(
                        "Retención (%)",
                        min_value=0.0,
                        max_value=100.0,
                        step=1.0,
                        key="factura_retencion_pct"
                    )

                concepto = st.text_input("Concepto", key="factura_concepto")
                notas = st.text_area("Notas", key="factura_notas", height=80)

                st.markdown("##### Crear documento desde menú")
                if not factura_cliente_id:
                    st.warning("Selecciona un cliente antes de crear líneas desde un menú.")
                else:
                    menus_factura_df = cargar_menus_supabase(obtener_user_id_actual())
                    menus_factura_df = menus_factura_df[
                        menus_factura_df["user_id"].apply(lambda valor: valor_user_id_menu({"user_id": valor}) == obtener_user_id_actual())
                    ].copy() if not menus_factura_df.empty else menus_factura_df
                    if menus_factura_df.empty:
                        st.info("No hay menús guardados para crear un documento.")
                    else:
                        menus_factura_df = menus_factura_df.copy()
                        menus_factura_df["etiqueta_selector"] = menus_factura_df.apply(
                            lambda row: (
                                f"{row.get('nombre', '')} · {row.get('tipo_menu', '')} · "
                                f"{numero_seguro(row.get('coste_total', 0), 0):.2f} € coste"
                            ).strip(" ·"),
                            axis=1
                        )
                        menus_factura_por_id = {
                            str(row.get("id")): row.to_dict()
                            for _, row in menus_factura_df.iterrows()
                            if row.get("id")
                        }
                        opciones_menus_factura = list(menus_factura_por_id.keys())
                        menu_doc_col1, menu_doc_col2, menu_doc_col3 = st.columns([2, 1, 1])
                        with menu_doc_col1:
                            menu_factura_id = st.selectbox(
                                "Menú guardado",
                                opciones_menus_factura,
                                format_func=lambda menu_id: menus_factura_por_id.get(str(menu_id), {}).get("etiqueta_selector", str(menu_id)),
                                key="selector_menu_facturacion"
                            )
                        with menu_doc_col2:
                            cantidad_menu_documento = st.number_input(
                                "Cantidad del menú",
                                min_value=0.001,
                                value=1.0,
                                step=1.0,
                                key="factura_cantidad_menu",
                            )
                        with menu_doc_col3:
                            anadir_menu_documento = st.button(
                                "Añadir menú",
                                type="primary",
                                use_container_width=True,
                            )

                        st.caption("Puedes añadir varios menús; cada uno conserva cantidad, coste y precio.")

                        if anadir_menu_documento:
                            if not factura_cliente_id:
                                st.warning("Selecciona un cliente antes de crear líneas desde un menú.")
                                st.stop()
                            if not menu_factura_id:
                                st.warning("Selecciona un menú antes de crear líneas.")
                                st.stop()
                            cabecera_menu, lineas_menu_df = cargar_menu_detalle_supabase(menu_factura_id)
                            if not cabecera_menu:
                                st.error("No se pudo cargar el menú seleccionado.")
                            elif lineas_menu_df.empty:
                                st.warning("El menú seleccionado no tiene recetas guardadas.")
                            else:
                                lineas_desde_menu, aviso_precio_menu = crear_lineas_factura_desde_menu(
                                    cabecera_menu,
                                    lineas_menu_df,
                                    desglosar=False,
                                    cantidad_menu=cantidad_menu_documento,
                                )
                                if not lineas_desde_menu:
                                    st.warning("No se pudieron crear líneas desde el menú seleccionado.")
                                else:
                                    lineas_existentes = [
                                        linea
                                        for linea in st.session_state["factura_lineas_actual"]
                                        if str(linea.get("descripcion", "") or "").strip()
                                    ]
                                    st.session_state["factura_lineas_actual"] = (
                                        lineas_existentes + lineas_desde_menu
                                    )
                                    st.session_state["factura_menu_feedback"] = (
                                        "Menú añadido con sus importes congelados.",
                                        aviso_precio_menu
                                    )
                                    st.rerun()

                lineas_accion_col1, lineas_accion_col2 = st.columns(2)
                with lineas_accion_col1:
                    if st.button("Quitar último menú", use_container_width=True):
                        if st.session_state["factura_lineas_actual"]:
                            st.session_state["factura_lineas_actual"].pop()
                            st.rerun()
                        else:
                            st.info("No hay menús para quitar.")
                with lineas_accion_col2:
                    if st.button("Limpiar documento", use_container_width=True):
                        st.session_state["factura_lineas_actual"] = []
                        st.session_state["factura_id_en_edicion"] = None
                        st.session_state["factura_presupuesto_origen_id"] = None
                        st.rerun()

                columnas_editor_factura = [
                    "descripcion", "cantidad", "unidad", "precio_unitario",
                    "coste_total_menu", "coste_linea_menu_documento",
                    "base_linea", "iva_linea", "total_linea"
                ]
                lineas_df = pd.DataFrame([
                    calcular_linea_factura(linea)
                    for linea in st.session_state["factura_lineas_actual"]
                ])
                for columna_editor in columnas_editor_factura:
                    if columna_editor not in lineas_df.columns:
                        lineas_df[columna_editor] = None
                lineas_df = lineas_df[columnas_editor_factura]
                lineas_editadas_df = st.data_editor(
                    lineas_df,
                    num_rows="fixed",
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "descripcion": st.column_config.TextColumn("Descripción", required=True, width="large"),
                        "cantidad": st.column_config.NumberColumn("Cantidad menú", min_value=0.001, step=1.0, format="%.3f"),
                        "unidad": st.column_config.TextColumn("Unidad", width="small"),
                        "precio_unitario": st.column_config.NumberColumn("Precio unitario", min_value=0.0, step=1.0, format="%.4f €"),
                        "coste_total_menu": st.column_config.NumberColumn("Coste menú", format="%.2f €"),
                        "coste_linea_menu_documento": st.column_config.NumberColumn("Coste línea", format="%.2f €"),
                        "base_linea": st.column_config.NumberColumn("Base línea", format="%.2f €"),
                        "iva_linea": st.column_config.NumberColumn("IVA línea", format="%.2f €"),
                        "total_linea": st.column_config.NumberColumn("Total línea", format="%.2f €")
                    },
                    disabled=[
                        "descripcion", "unidad", "coste_total_menu",
                        "coste_linea_menu_documento", "base_linea",
                        "iva_linea", "total_linea",
                    ],
                    key="editor_factura_lineas"
                )

                lineas_actualizadas = []
                for idx, linea_editada in enumerate(lineas_editadas_df.to_dict(orient="records")):
                    linea_origen = (
                        st.session_state["factura_lineas_actual"][idx]
                        if idx < len(st.session_state["factura_lineas_actual"])
                        else {}
                    )
                    for campo_origen in [
                        "origen_tipo", "origen_id", "menu_origen_id",
                        "menu_nombre_snapshot", "cantidad_menu",
                        "coste_total_menu", "coste_linea_menu_documento",
                        "precio_total_menu", "precio_linea_menu_documento",
                        "observaciones",
                    ]:
                        if linea_origen.get(campo_origen) is not None:
                            linea_editada[campo_origen] = linea_origen.get(campo_origen)
                    if linea_editada.get("origen_tipo") == "menu":
                        cantidad_editada = numero_seguro(linea_editada.get("cantidad"), 0.0)
                        precio_menu_editado = numero_seguro(
                            linea_editada.get("precio_unitario"),
                            0.0,
                        )
                        linea_editada["cantidad_menu"] = cantidad_editada
                        linea_editada["coste_linea_menu_documento"] = (
                            numero_seguro(linea_editada.get("coste_total_menu"), 0.0)
                            * cantidad_editada
                        )
                        linea_editada["precio_total_menu"] = precio_menu_editado
                        linea_editada["precio_linea_menu_documento"] = (
                            precio_menu_editado * cantidad_editada
                        )
                    lineas_actualizadas.append(calcular_linea_factura(linea_editada))
                if lineas_actualizadas != st.session_state["factura_lineas_actual"]:
                    st.session_state["factura_lineas_actual"] = lineas_actualizadas
                    st.rerun()

                totales_factura = calcular_totales_factura(
                    st.session_state["factura_lineas_actual"],
                    retencion_pct=retencion_pct
                )
                lineas_menu_documento = extraer_lineas_menu_documento(
                    st.session_state["factura_lineas_actual"]
                )
                coste_documento = (
                    calcular_coste_presupuesto(lineas_menu_documento)
                    if tipo_documento == "presupuesto" and lineas_menu_documento
                    else calcular_coste_factura(lineas_menu_documento)
                    if lineas_menu_documento
                    else 0.0
                )
                margen_documento = totales_factura["base_imponible"] - coste_documento
                food_cost_documento = (
                    coste_documento / totales_factura["base_imponible"] * 100
                    if totales_factura["base_imponible"] > 0 else 0.0
                )
                total_col1, total_col2, total_col3, total_col4, total_col5, total_col6 = st.columns(6)
                with total_col1:
                    st.metric("Coste", f"{coste_documento:.2f} €")
                with total_col2:
                    st.metric("Venta neta", f"{totales_factura['base_imponible']:.2f} €")
                with total_col3:
                    st.metric("Margen", f"{margen_documento:.2f} €")
                with total_col4:
                    st.metric("Food Cost", f"{food_cost_documento:.1f} %")
                with total_col5:
                    st.metric("IVA", f"{totales_factura['iva_importe']:.2f} €")
                with total_col6:
                    st.metric("Total", f"{totales_factura['total']:.2f} €")

                factura_actual = {
                    "cliente_id": factura_cliente_id,
                    "numero_factura": str(numero_factura or "").strip(),
                    "tipo_documento": tipo_documento,
                    "estado": estado,
                    "fecha_emision": fecha_emision.isoformat() if fecha_emision else None,
                    "fecha_vencimiento": fecha_vencimiento.isoformat() if fecha_vencimiento else None,
                    "concepto": str(concepto or "").strip() or None,
                    "iva_pct": 21.0,
                    "retencion_pct": float(retencion_pct),
                    "metodo_pago": str(metodo_pago or "").strip() or None,
                    "estado_cobro": estado_cobro,
                    "notas": str(notas or "").strip() or None
                }
                if tipo_documento != "presupuesto" and st.session_state.get("factura_presupuesto_origen_id"):
                    factura_actual["presupuesto_id"] = st.session_state["factura_presupuesto_origen_id"]
                cliente_actual_factura = clientes_por_id_factura.get(str(factura_cliente_id), {})
                lineas_validas_descarga = [
                    calcular_linea_factura(linea)
                    for linea in st.session_state["factura_lineas_actual"]
                    if str(linea.get("descripcion", "") or "").strip()
                ]
                menu_id_descarga = obtener_menu_id_desde_lineas_factura(lineas_validas_descarga)
                if menu_id_descarga:
                    factura_actual["menu_nombre_asociado"] = obtener_nombre_menu_por_id(menu_id_descarga)

                descarga_col, guardar_col, actualizar_col = st.columns(3)
                with descarga_col:
                    if not cliente_actual_factura or not lineas_validas_descarga:
                        st.warning("Selecciona un cliente y añade al menos una línea para descargar el documento.")
                        st.button("Descargar documento", disabled=True, use_container_width=True)
                    else:
                        documento_descarga = generar_pdf_factura(
                            factura_actual,
                            lineas_validas_descarga,
                            cliente_actual_factura
                        )
                        st.download_button(
                            documento_descarga["label"],
                            data=documento_descarga["data"],
                            file_name=documento_descarga["file_name"],
                            mime=documento_descarga["mime"],
                            use_container_width=True
                        )

                with guardar_col:
                    guardar_documento = st.button("Guardar documento", type="primary", use_container_width=True)

                with actualizar_col:
                    actualizar_documento = st.button(
                        "Actualizar documento seleccionado",
                        disabled=not bool(st.session_state.get("factura_id_en_edicion")),
                        use_container_width=True
                    )

                if guardar_documento:
                    ok_guardar_doc, mensaje_guardar_doc, factura_guardada = guardar_factura_supabase(
                        factura_actual,
                        st.session_state["factura_lineas_actual"]
                    )
                    if ok_guardar_doc:
                        st.success(f"{mensaje_guardar_doc} {factura_guardada.get('numero_factura', numero_factura)}")
                        st.session_state["factura_lineas_actual"] = []
                        st.session_state["factura_id_en_edicion"] = None
                        st.session_state["factura_presupuesto_origen_id"] = None
                        st.rerun()
                    else:
                        st.error(mensaje_guardar_doc)

                if actualizar_documento:
                    ok_actualizar_doc, mensaje_actualizar_doc, factura_actualizada = actualizar_factura_supabase(
                        st.session_state.get("factura_id_en_edicion"),
                        factura_actual,
                        st.session_state["factura_lineas_actual"]
                    )
                    if ok_actualizar_doc:
                        st.success(f"{mensaje_actualizar_doc} {factura_actualizada.get('numero_factura', numero_factura)}")
                        st.rerun()
                    else:
                        st.error(mensaje_actualizar_doc)

                st.markdown("##### Últimos documentos")
                if facturas_df.empty:
                    st.info("Todavía no hay presupuestos ni facturas guardados.")
                else:
                    facturas_resumen = facturas_df.copy()
                    facturas_resumen["cliente"] = facturas_resumen["cliente_id"].apply(
                        lambda cliente_id: nombres_clientes_factura.get(str(cliente_id), "")
                    )
                    facturas_resumen = facturas_resumen[[
                        "numero_factura", "tipo_documento", "cliente",
                        "fecha_emision", "estado", "total"
                    ]].rename(columns={
                        "numero_factura": "número",
                        "tipo_documento": "tipo",
                        "fecha_emision": "fecha"
                    })
                    st.dataframe(facturas_resumen.fillna(""), use_container_width=True, hide_index=True)


with main_tab_recetas:
    with st.expander("Parámetros avanzados", expanded=False):
        c1, c2, c3 = st.columns(3)
        with c1:
            costes_indirectos_pct = st.number_input("Costes Indirectos (%)", min_value=0.0, key="costes_indirectos_pct")
        with c2:
            margen_beneficio_pct = st.number_input("Margen Beneficio (%)", min_value=0.0, max_value=99.9, key="margen_beneficio_pct")
        with c3:
            iva_pct = st.number_input("IVA Evento (%)", min_value=0.0, max_value=100.0, step=1.0, key="iva_pct")

    if st.session_state['ingredientes']:
        subtotal_ing = sum(float(ing.get('cantidad_bruta', 0.0)) * float(ing.get('precio_unidad', 0.0)) for ing in st.session_state['ingredientes'])

        ci_val = costes_indirectos_pct if costes_indirectos_pct is not None else 0.0
        mb_val = margen_beneficio_pct if margen_beneficio_pct is not None else 0.0
        iva_val = iva_pct if iva_pct is not None else 0.0

        costes_ind = subtotal_ing * (ci_val / 100)
        coste_total = subtotal_ing
        coste_para_precio = coste_total + costes_ind

        factor_margen = (1 - (mb_val / 100))
        pvp_neto = coste_para_precio / factor_margen if factor_margen > 0 else 0.0
        monto_iva = pvp_neto * (iva_val / 100)
        pvp_final = pvp_neto + monto_iva
        raciones_deseadas_metricas = pd.to_numeric(st.session_state.get('raciones_deseadas', 0.0), errors='coerce')
        raciones_deseadas_metricas = 0.0 if pd.isna(raciones_deseadas_metricas) else float(raciones_deseadas_metricas)
        pvp_por_racion = pvp_final / raciones_deseadas_metricas if raciones_deseadas_metricas > 0 else 0.0

        excel_virtual = generar_excel(
            nombre_plato,
            st.session_state['ingredientes'],
            ci_val,
            mb_val,
            iva_val,
            st.session_state['raciones_base'],
            st.session_state['raciones_deseadas'],
            st.session_state['factor_raciones']
        )
        st.download_button(
            label=f"Descargar ficha Excel",
            data=excel_virtual,
            file_name=f"Ficha_{nombre_plato.replace(' ', '_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True
        )
    else:
        st.info("Anade ingredientes para ver costes y exportar la ficha.")

with main_tab_recetas:
    st.markdown("##### Recetas guardadas")
    st.session_state.pop("mensaje_receta_cargada", None)

    opciones_recetas = []
    recetas_por_id = {}
    receta_id_seleccionada = None
    if st.session_state.pop("limpiar_selector_receta_guardada", False):
        st.session_state.pop("selector_receta_guardada", None)

    if not supabase_disponible:
        st.warning("El inventario no está disponible. No se pueden cargar recetas guardadas ahora.")
    else:
        recetas_guardadas_df = cargar_recetas_supabase(obtener_user_id_actual())
        if recetas_guardadas_df.empty:
            st.info("Todavía no hay recetas guardadas en el inventario.")
        else:
            recetas_guardadas_df = recetas_guardadas_df.copy()
            recetas_sin_propietario = recetas_guardadas_df["user_id"].apply(
                lambda valor: valor_user_id_receta({"user_id": valor}) is None
            ).sum()
            if obtener_user_id_actual() and recetas_sin_propietario:
                st.info(
                    f"Hay {recetas_sin_propietario} recetas antiguas sin propietario. "
                    "Puedes cargarlas para consultarlas o duplicarlas en tu cuenta."
                )
            recetas_guardadas_df["etiqueta_selector"] = recetas_guardadas_df.apply(
                lambda row: (
                    f"{row.get('codigo_receta', '')} · {row.get('nombre', '')}"
                    f"{' · sin propietario' if valor_user_id_receta(row.to_dict()) is None else ''}"
                ).strip(" ·"),
                axis=1
            )
            opciones_recetas = recetas_guardadas_df["id"].dropna().astype(str).tolist()
            recetas_por_id = {
                str(row.get("id")): row.to_dict()
                for _, row in recetas_guardadas_df.iterrows()
                if row.get("id")
            }

            if not opciones_recetas:
                st.info("Hay recetas guardadas, pero no se pudo leer su identificador.")
            else:
                receta_pendiente_selector = st.session_state.pop("selector_receta_guardada_pendiente", None)
                if receta_pendiente_selector and str(receta_pendiente_selector) in opciones_recetas:
                    st.session_state["selector_receta_guardada"] = str(receta_pendiente_selector)

                cargar_col1, cargar_col2 = st.columns([4, 1])
                with cargar_col1:
                    receta_id_seleccionada = st.selectbox(
                        "Recetas guardadas",
                        opciones_recetas,
                        format_func=lambda receta_id: recetas_por_id.get(str(receta_id), {}).get("etiqueta_selector", str(receta_id)),
                        key="selector_receta_guardada"
                    )
                with cargar_col2:
                    cargar_receta_btn = st.button("Cargar receta", use_container_width=True)

                if cargar_receta_btn:
                    receta_seleccionada = recetas_por_id.get(str(receta_id_seleccionada), {})
                    if receta_seleccionada and not receta_es_propia_o_antigua(receta_seleccionada):
                        st.error("No puedes modificar una receta de otro usuario")
                    else:
                        cabecera, ingredientes_df = cargar_receta_detalle_supabase(receta_id_seleccionada)
                        ingredientes_cargados = preparar_ingredientes_receta_para_sesion(ingredientes_df)
                        if not cabecera:
                            st.error("No se pudo cargar la cabecera de la receta seleccionada.")
                        elif not ingredientes_cargados:
                            st.warning("La receta seleccionada no tiene ingredientes guardados.")
                        else:
                            raciones_base_raw = pd.to_numeric(cabecera.get("raciones_base"), errors="coerce")
                            raciones_cargadas = 0.0 if pd.isna(raciones_base_raw) else float(raciones_base_raw)
                            if raciones_cargadas <= 0:
                                raciones_cargadas = 1.0
                                st.session_state["aviso_receta_antigua"] = (
                                    "La receta tenía una base de raciones no válida; "
                                    "se muestra con 1 ración para poder corregirla."
                                )
                            st.session_state["receta_raciones_base_cargada"] = float(raciones_cargadas)
                            st.session_state["ingredientes"] = ingredientes_cargados
                            st.session_state["ingredientes_base_raciones"] = [dict(ing) for ing in ingredientes_cargados]
                            st.session_state["factor_raciones"] = 1.0
                            st.session_state["raciones_base"] = raciones_cargadas
                            st.session_state["raciones_deseadas"] = raciones_cargadas
                            st.session_state["receta_raciones_base"] = raciones_cargadas
                            st.session_state["input_raciones_base_receta_pendiente"] = raciones_cargadas
                            st.session_state["receta_raciones_deseadas"] = raciones_cargadas
                            st.session_state["raciones_base_aplicadas"] = raciones_cargadas
                            st.session_state["raciones_deseadas_aplicadas"] = raciones_cargadas
                            st.session_state["sincronizar_inputs_raciones"] = True
                            st.session_state["receta_nombre"] = str(cabecera.get("nombre", "") or "Mi Receta")
                            st.session_state["receta_categoria"] = str(cabecera.get("categoria", "") or "")
                            st.session_state["receta_tipo_plato"] = str(cabecera.get("tipo_plato", "") or "")
                            st.session_state["receta_observaciones"] = str(
                                cabecera.get("observaciones") or cabecera.get("descripcion") or ""
                            )
                            st.session_state["sincronizar_campos_receta"] = True
                            codigo_cargado = str(cabecera.get("codigo_receta", "") or "").strip()
                            nombre_cargado = str(cabecera.get("nombre", "") or "receta").strip()
                            st.session_state["receta_id_cargada"] = str(cabecera.get("id", receta_id_seleccionada))
                            st.session_state["codigo_receta_cargada"] = codigo_cargado
                            st.session_state["mensaje_receta_cargada"] = f"Receta cargada: {nombre_cargado} ({codigo_cargado})."
                            st.rerun()


    if st.session_state['ingredientes']:
        subtotal_ing_guardado = sum(float(ing.get('cantidad_bruta', 0.0)) * float(ing.get('precio_unidad', 0.0)) for ing in st.session_state['ingredientes'])
        ci_val = float(st.session_state.get('costes_indirectos_pct', 0.0))
        mb_val = float(st.session_state.get('margen_beneficio_pct', 0.0))
        iva_val = float(st.session_state.get('iva_pct', 0.0))
        costes_ind_guardado = subtotal_ing_guardado * (ci_val / 100)
        coste_total = subtotal_ing_guardado
        coste_para_precio_guardado = coste_total + costes_ind_guardado
        factor_margen_guardado = (1 - (mb_val / 100))
        pvp_neto = (
            coste_para_precio_guardado / factor_margen_guardado
            if factor_margen_guardado > 0 else 0.0
        )
        pvp_final = pvp_neto + (pvp_neto * (iva_val / 100))
    else:
        ci_val = float(st.session_state.get('costes_indirectos_pct', 0.0))
        mb_val = float(st.session_state.get('margen_beneficio_pct', 0.0))
        iva_val = float(st.session_state.get('iva_pct', 0.0))
        coste_total = 0.0
        pvp_neto = 0.0
        pvp_final = 0.0
    st.markdown("##### Guardar")
    if not supabase_disponible:
        st.warning("El inventario no está disponible. No se puede guardar la receta ahora.")
    elif not obtener_user_id_actual():
        pass
    with st.form("form_guardar_receta_nueva"):
        nombre_receta_guardar = st.text_input("Nombre de receta", value=st.session_state.get("receta_nombre", nombre_plato))
        categoria_receta = str(st.session_state.get("receta_categoria", "") or "")
        tipo_plato_receta = str(st.session_state.get("receta_tipo_plato", "") or "")
        raciones_base_receta = float(st.session_state.get('receta_raciones_base', st.session_state.get('raciones_base', 1.0)))
        with st.expander("Elaboración y observaciones", expanded=False):
            observaciones_receta = st.text_area("Elaboración / observaciones", key="input_receta_observaciones", height=80)

        guardar_col, actualizar_col, duplicar_col = st.columns(3)
        with guardar_col:
            guardar_nueva = st.form_submit_button("Guardar receta", type="primary", use_container_width=True)
        with actualizar_col:
            actualizar_existente = st.form_submit_button("Actualizar receta", use_container_width=True)
        with duplicar_col:
            duplicar_existente = st.form_submit_button("Duplicar receta", use_container_width=True)

        if guardar_nueva or actualizar_existente or duplicar_existente:
            st.session_state["receta_nombre"] = str(nombre_receta_guardar or "").strip() or st.session_state.get("receta_nombre", "Mi Receta")
            st.session_state["receta_categoria"] = categoria_receta
            st.session_state["receta_tipo_plato"] = tipo_plato_receta
            st.session_state["receta_observaciones"] = observaciones_receta
            nombre_limpio = nombre_receta_guardar.strip()
            receta_id_cargada = st.session_state.get("receta_id_cargada")
            codigo_receta_cargada = st.session_state.get("codigo_receta_cargada", "")
            if not supabase_disponible or supabase is None:
                st.error("El inventario no está conectado correctamente.")
            elif not nombre_limpio:
                st.error("Indica un nombre de receta antes de guardar.")
            elif not st.session_state.get("ingredientes"):
                st.error("Añade al menos un ingrediente antes de guardar la receta.")
            elif actualizar_existente and not receta_id_cargada:
                st.warning("Carga una receta guardada antes de actualizarla.")
            elif duplicar_existente and not receta_id_cargada:
                st.warning("Carga una receta guardada antes de duplicarla.")
            else:
                codigo_receta = codigo_receta_cargada if actualizar_existente else generar_codigo_receta()
                datos_receta = {
                    "user_id": obtener_user_id_actual(),
                    "codigo_receta": codigo_receta,
                    "nombre": nombre_limpio,
                    "categoria": categoria_receta.strip(),
                    "tipo_plato": tipo_plato_receta.strip(),
                    # La capa de persistencia usa este valor como divisor y
                    # guarda finalmente la cabecera con raciones_base = 1.
                    "raciones_base": float(raciones_base_receta),
                    "unidad_servicio": "ración",
                    "descripcion": observaciones_receta.strip(),
                    "observaciones": observaciones_receta.strip(),
                    "costes_indirectos_pct": float(ci_val),
                    "margen_beneficio_pct": float(mb_val),
                    "iva_pct": float(iva_val),
                    "coste_total": float(coste_total),
                    "precio_venta_sin_iva": float(pvp_neto),
                    "precio_venta_con_iva": float(pvp_final),
                    "activa": True
                }
                if actualizar_existente:
                    ok, mensaje = actualizar_receta_supabase(
                        receta_id_cargada,
                        datos_receta,
                        st.session_state["ingredientes"]
                    )
                    if ok:
                        st.session_state["codigo_receta_cargada"] = codigo_receta
                        st.session_state["receta_raciones_base"] = 1.0
                        st.session_state["receta_raciones_base_cargada"] = 1.0
                        st.session_state["input_raciones_base_receta_pendiente"] = 1.0
                        limpiar_cache_recetas_guardadas()
                        st.session_state["selector_receta_guardada_pendiente"] = receta_id_cargada
                        st.session_state["mensaje_receta_cargada"] = (
                            f"Receta actualizada correctamente con código {codigo_receta}. "
                            "Listado de recetas actualizado."
                        )
                        st.rerun()
                    else:
                        st.error(mensaje)
                elif duplicar_existente:
                    ok, mensaje, receta_guardada = duplicar_receta_supabase(
                        datos_receta,
                        st.session_state["ingredientes"]
                    )
                    if ok:
                        nuevo_codigo = receta_guardada.get("codigo_receta", "")
                        nuevo_nombre = receta_guardada.get("nombre", nombre_limpio)
                        st.session_state["receta_id_cargada"] = receta_guardada.get("id")
                        st.session_state["codigo_receta_cargada"] = nuevo_codigo
                        st.session_state["receta_nombre"] = nuevo_nombre
                        st.session_state["receta_raciones_base"] = 1.0
                        st.session_state["receta_raciones_base_cargada"] = 1.0
                        st.session_state["input_raciones_base_receta_pendiente"] = 1.0
                        st.session_state["sincronizar_campos_receta"] = True
                        limpiar_cache_recetas_guardadas()
                        st.session_state["selector_receta_guardada_pendiente"] = receta_guardada.get("id")
                        st.session_state["mensaje_receta_cargada"] = (
                            f"{'Receta duplicada en tu cuenta' if obtener_user_id_actual() else 'Receta duplicada correctamente'} "
                            f"con código {nuevo_codigo}. "
                            "Listado de recetas actualizado."
                        )
                        st.rerun()
                    else:
                        st.error(mensaje)
                else:
                    ok, mensaje, receta_guardada = guardar_receta_nueva_supabase(
                        datos_receta,
                        st.session_state["ingredientes"]
                    )
                    if ok:
                        codigo_mostrado = receta_guardada.get("codigo_receta", codigo_receta)
                        st.session_state["receta_id_cargada"] = receta_guardada.get("id")
                        st.session_state["codigo_receta_cargada"] = codigo_mostrado
                        st.session_state["receta_raciones_base"] = 1.0
                        st.session_state["receta_raciones_base_cargada"] = 1.0
                        st.session_state["input_raciones_base_receta_pendiente"] = 1.0
                        limpiar_cache_recetas_guardadas()
                        st.session_state["selector_receta_guardada_pendiente"] = receta_guardada.get("id")
                        st.session_state["mensaje_receta_cargada"] = (
                            f"{mensaje} con código {codigo_mostrado}. "
                            "Listado de recetas actualizado."
                        )
                        st.rerun()
                    else:
                        st.error(mensaje)

    with st.expander("Mantenimiento", expanded=False):
        # TODO: mover eliminación de recetas al entorno Administración/Mantenimiento.
        receta_id_para_eliminar = (
            str(receta_id_seleccionada)
            if receta_id_seleccionada
            else str(st.session_state.get("receta_id_cargada") or "")
        )
        receta_para_eliminar = recetas_por_id.get(receta_id_para_eliminar, {})

        if not supabase_disponible:
            st.warning("El inventario no está disponible. No se puede eliminar la receta ahora.")
        elif not receta_id_para_eliminar:
            st.info("Selecciona o carga una receta antes de intentar eliminarla.")
        elif receta_para_eliminar and not receta_es_modificable(receta_para_eliminar):
            st.warning("No puedes modificar una receta de otro usuario")
            if valor_user_id_receta(receta_para_eliminar) is None:
                st.info("Esta receta antigua no tiene propietario. Cierra sesión para eliminarla o duplícala en tu cuenta.")
        else:
            codigo_eliminar = str(
                receta_para_eliminar.get("codigo_receta")
                or st.session_state.get("codigo_receta_cargada")
                or "sin código"
            )
            nombre_eliminar = str(
                receta_para_eliminar.get("nombre")
                or st.session_state.get("receta_nombre")
                or "receta seleccionada"
            )
            st.warning(
                "Esta acción no se puede deshacer. Se borrará la cabecera de la receta y "
                "sus ingredientes asociados. No se borrarán ingredientes del inventario."
            )
            st.write(f"**Receta seleccionada:** {codigo_eliminar} · {nombre_eliminar}")
            confirmacion_eliminar = st.text_input(
                "Escribe ELIMINAR para confirmar",
                key=f"confirmar_eliminar_receta_{receta_id_para_eliminar}"
            )
            if st.button("Borrar receta seleccionada", type="secondary", use_container_width=True):
                if confirmacion_eliminar != "ELIMINAR":
                    st.error("Para borrar la receta debes escribir exactamente ELIMINAR.")
                else:
                    ok, mensaje = eliminar_receta_supabase(receta_id_para_eliminar)
                    if ok:
                        if str(st.session_state.get("receta_id_cargada") or "") == receta_id_para_eliminar:
                            st.session_state["receta_id_cargada"] = None
                            st.session_state["codigo_receta_cargada"] = ""
                            st.session_state["receta_raciones_base_cargada"] = None
                        limpiar_cache_recetas_guardadas()
                        st.session_state["limpiar_selector_receta_guardada"] = True
                        st.session_state["mensaje_receta_cargada"] = "Receta eliminada correctamente. Listado de recetas actualizado."
                        st.rerun()
                    else:
                        st.error(mensaje)

with main_tab_menus:
    st.markdown("### Menús")
    st.caption("Construye un menú básico combinando recetas guardadas.")
    sincronizar_widgets_menu()

    mensaje_menu_cargado = st.session_state.pop("mensaje_menu_cargado", None)
    if mensaje_menu_cargado:
        st.success(mensaje_menu_cargado)
    mensaje_menu_aviso = st.session_state.pop("mensaje_menu_aviso", None)
    if mensaje_menu_aviso:
        st.warning(mensaje_menu_aviso)

    st.markdown("##### Cargar menú guardado")
    if not supabase_disponible:
        st.warning("El inventario no está disponible. No se pueden cargar menús guardados ahora.")
    else:
        menus_guardados_df = cargar_menus_supabase(obtener_user_id_actual())
        if menus_guardados_df.empty:
            st.info("Todavía no hay menús guardados en el inventario.")
        else:
            menus_guardados_df = menus_guardados_df.copy()
            menus_sin_propietario = menus_guardados_df["user_id"].apply(
                lambda valor: valor_user_id_menu({"user_id": valor}) is None
            ).sum()
            if obtener_user_id_actual() and menus_sin_propietario:
                st.info(
                    f"Hay {menus_sin_propietario} menús antiguos sin propietario. "
                    "Puedes cargarlos para consultarlos o duplicarlos en tu cuenta."
                )
            menus_guardados_df["etiqueta_selector"] = menus_guardados_df.apply(
                lambda row: (
                    f"{row.get('nombre', '')} · {row.get('tipo_menu', '')}"
                    f"{' · sin propietario' if valor_user_id_menu(row.to_dict()) is None else ''}"
                ).strip(" ·"),
                axis=1
            )
            opciones_menus = menus_guardados_df["id"].dropna().astype(str).tolist()
            menus_por_id = {
                str(row.get("id")): row.to_dict()
                for _, row in menus_guardados_df.iterrows()
                if row.get("id")
            }

            if opciones_menus:
                cargar_menu_col1, cargar_menu_col2 = st.columns([3, 1])
                with cargar_menu_col1:
                    menu_id_seleccionado = st.selectbox(
                        "Menús guardados",
                        opciones_menus,
                        format_func=lambda menu_id: menus_por_id.get(str(menu_id), {}).get("etiqueta_selector", str(menu_id)),
                        key="selector_menu_guardado"
                    )
                with cargar_menu_col2:
                    cargar_menu_btn = st.button("Cargar menú", use_container_width=True)

                if cargar_menu_btn:
                    menu_seleccionado = menus_por_id.get(str(menu_id_seleccionado), {})
                    if menu_seleccionado and not menu_es_propio_o_antiguo(menu_seleccionado):
                        st.error("No puedes modificar un menú de otro usuario")
                    else:
                        cabecera_menu, lineas_menu_df = cargar_menu_detalle_supabase(menu_id_seleccionado)
                        if not cabecera_menu:
                            st.error("No se pudo cargar la cabecera del menú seleccionado.")
                        elif lineas_menu_df.empty:
                            st.warning("El menú seleccionado no tiene recetas guardadas.")
                        else:
                            lineas_reconstruidas = []
                            for orden, linea in enumerate(lineas_menu_df.to_dict(orient="records"), start=1):
                                precio_receta = linea.get("precio_receta_sin_iva", None)
                                if precio_receta is None or pd.isna(precio_receta):
                                    precio_receta = linea.get("precio_receta_con_iva", None)

                                hay_precio_receta = precio_receta is not None and not pd.isna(precio_receta)
                                lineas_reconstruidas.append(recalcular_linea_menu({
                                    "receta_id": str(linea.get("receta_id", "") or ""),
                                    "codigo_receta": str(linea.get("codigo_receta", "") or ""),
                                    "nombre_receta": str(linea.get("nombre_receta", "") or "Receta sin nombre"),
                                    "raciones_receta_en_menu": max(
                                        float(numero_seguro(linea.get("raciones", 1.0), 1.0)),
                                        0.001,
                                    ),
                                    "raciones_base_receta": max(
                                        float(numero_seguro(linea.get("raciones_base", 1.0), 1.0)),
                                        0.001,
                                    ),
                                    "coste_receta": float(numero_seguro(linea.get("coste_receta", 0.0), 0.0)),
                                    "precio_receta": float(numero_seguro(precio_receta, 0.0)) if hay_precio_receta else None,
                                    "orden": int(numero_seguro(linea.get("orden", orden), orden)),
                                    "seccion": str(linea.get("seccion", "") or "")
                                }))

                            st.session_state["menu_actual"] = normalizar_lineas_menu(lineas_reconstruidas)
                            st.session_state["menu_id"] = str(cabecera_menu.get("id", menu_id_seleccionado))
                            st.session_state["menu_nombre"] = str(cabecera_menu.get("nombre", "") or "")
                            st.session_state["menu_tipo"] = str(cabecera_menu.get("tipo_menu", "") or "Otro")
                            st.session_state["menu_numero_comensales"] = int(numero_seguro(cabecera_menu.get("numero_comensales", 1), 1))
                            st.session_state["mensaje_menu_cargado"] = f"Menú cargado: {st.session_state['menu_nombre']}."
                            st.rerun()

    menu_col1, menu_col2, menu_col3 = st.columns([2.1, 1, 1])
    with menu_col1:
        nombre_menu = st.text_input("Nombre del menú", key="menu_nombre")
    with menu_col2:
        opciones_tipo_menu = ["Cóctel", "Buffet", "Menú sentado", "Catering", "Otro"]
        tipo_menu_actual = st.session_state.get("menu_tipo", "")
        if tipo_menu_actual and tipo_menu_actual not in opciones_tipo_menu:
            opciones_tipo_menu.append(tipo_menu_actual)
        tipo_menu = st.selectbox(
            "Tipo de menú",
            opciones_tipo_menu,
            key="menu_tipo"
        )
    with menu_col3:
        numero_comensales = st.number_input(
            "Pax de referencia (informativo)",
            min_value=1,
            step=1,
            value=int(st.session_state.get("menu_numero_comensales", 1) or 1),
            help="Dato informativo. No cambia las raciones asignadas a cada receta.",
            key="menu_numero_comensales"
        )

    if st.session_state.get("menu_id"):
        st.caption(f"Menú cargado para actualizar: {st.session_state['menu_id']}")
    if supabase_disponible and not obtener_user_id_actual():
        st.caption("Inicia sesión para guardar menús en tu cuenta")

    st.markdown("##### Añadir recetas")

    recetas_menu_df = pd.DataFrame()
    recetas_menu_por_id = {}
    opciones_recetas_menu = []

    if not supabase_disponible:
        st.warning("El inventario no está disponible. No se pueden cargar recetas guardadas ahora.")
    else:
        recetas_menu_df = cargar_recetas_supabase(obtener_user_id_actual())
        if recetas_menu_df.empty:
            st.info("Todavía no hay recetas guardadas para crear un menú.")
        else:
            recetas_menu_df = recetas_menu_df.copy()
            recetas_menu_df["etiqueta_selector"] = recetas_menu_df.apply(
                lambda row: f"{row.get('codigo_receta', '')} · {row.get('nombre', '')}".strip(" ·"),
                axis=1
            )
            opciones_recetas_menu = recetas_menu_df["id"].dropna().astype(str).tolist()
            recetas_menu_por_id = {
                str(row.get("id")): row.to_dict()
                for _, row in recetas_menu_df.iterrows()
                if row.get("id")
            }

    if opciones_recetas_menu:
        add_col1, add_col2, add_col3, add_col4 = st.columns([3, 1, 1, 1])
        with add_col1:
            receta_menu_id = st.selectbox(
                "Recetas guardadas",
                opciones_recetas_menu,
                format_func=lambda receta_id: recetas_menu_por_id.get(str(receta_id), {}).get("etiqueta_selector", str(receta_id)),
                key="selector_receta_menu"
            )
        with add_col2:
            raciones_linea_menu = st.number_input(
                "Raciones de esta receta",
                min_value=0.001,
                step=1.0,
                value=1.0,
                key="menu_raciones_receta"
            )
        with add_col3:
            seccion_linea_menu = st.selectbox(
                "Sección",
                SECCIONES_MENU,
                format_func=lambda valor: valor or "Sin sección",
                key="menu_seccion_receta"
            )
        with add_col4:
            anadir_receta_menu = st.button("Añadir receta al menú", use_container_width=True)

        if anadir_receta_menu:
            receta_seleccionada = recetas_menu_por_id.get(str(receta_menu_id), {})
            if raciones_linea_menu <= 0:
                st.error("Las raciones de la receta deben ser mayores que 0.")
            elif any(str(linea.get("receta_id")) == str(receta_menu_id) for linea in st.session_state["menu_actual"]):
                st.session_state["mensaje_menu_aviso"] = (
                    "Esta receta ya está en el menú. Puedes mantenerla duplicada si representa otro pase "
                    "o quitar una de las líneas."
                )
            if raciones_linea_menu > 0:
                try:
                    linea_menu = crear_linea_menu_desde_receta(
                        receta_menu_id,
                        receta_seleccionada,
                        raciones_linea_menu,
                        len(st.session_state["menu_actual"]) + 1,
                        seccion_linea_menu
                    )
                    st.session_state["menu_actual"].append(linea_menu)
                    st.session_state["menu_actual"] = normalizar_lineas_menu(st.session_state["menu_actual"])
                    st.success(f"Receta añadida al menú: {linea_menu['nombre_receta']}.")
                    st.rerun()
                except ValueError as exc:
                    st.error(str(exc))

    st.markdown("##### Menú actual")

    lineas_menu_actual = normalizar_lineas_menu(st.session_state.get("menu_actual", []))
    if lineas_menu_actual != st.session_state.get("menu_actual", []):
        st.session_state["menu_actual"] = lineas_menu_actual

    if lineas_menu_actual:
        menu_editor_df = pd.DataFrame([
            {
                "orden": linea.get("orden", idx + 1),
                "seccion": linea.get("seccion", ""),
                "nombre_receta": linea.get("nombre_receta", ""),
                "coste_por_racion": linea.get("coste_por_racion_receta", 0.0),
                "raciones_receta_en_menu": linea.get("raciones_receta_en_menu", 1.0),
                "coste_total_linea": linea.get("coste_total_linea", 0.0),
                "precio_por_racion": (
                    numero_seguro(linea.get("precio_receta", 0.0), 0.0) / max(numero_seguro(linea.get("raciones_base", 1.0), 1.0), 1.0)
                    if linea.get("precio_receta") is not None and not pd.isna(linea.get("precio_receta"))
                    else None
                ),
                "precio_total_linea": linea.get("precio_total_linea", None),
            }
            for idx, linea in enumerate(lineas_menu_actual)
        ])
        menu_editado_df = st.data_editor(
            menu_editor_df,
            num_rows="fixed",
            use_container_width=True,
            hide_index=True,
            column_config={
                "orden": st.column_config.NumberColumn("Orden", min_value=1, step=1, width="small"),
                "seccion": st.column_config.SelectboxColumn(
                    "Sección",
                    options=SECCIONES_MENU,
                    help="Se guarda en menu_recetas.seccion si la columna existe."
                ),
                "nombre_receta": st.column_config.TextColumn("Receta", width="medium"),
                "raciones_receta_en_menu": st.column_config.NumberColumn(
                    "Raciones asignadas",
                    min_value=0.001,
                    step=1.0,
                    format="%.2f",
                ),
                "coste_por_racion": st.column_config.NumberColumn("Coste por ración", format="%.2f €"),
                "coste_total_linea": st.column_config.NumberColumn("Coste línea", format="%.2f €"),
                "precio_por_racion": st.column_config.NumberColumn("Precio por ración", format="%.2f €"),
                "precio_total_linea": st.column_config.NumberColumn("Precio línea", format="%.2f €"),
            },
            disabled=[
                "nombre_receta",
                "coste_por_racion",
                "coste_total_linea",
                "precio_por_racion",
                "precio_total_linea"
            ],
            key="editor_lineas_menu"
        )

        lineas_editadas = []
        for idx, fila in menu_editado_df.iterrows():
            linea_actualizada = dict(lineas_menu_actual[idx])
            linea_actualizada["orden"] = int(numero_seguro(fila.get("orden", idx + 1), idx + 1))
            linea_actualizada["raciones_receta_en_menu"] = numero_seguro(
                fila.get("raciones_receta_en_menu", 1.0),
                1.0,
            )
            linea_actualizada["seccion"] = str(fila.get("seccion", "") or "").strip()
            lineas_editadas.append(recalcular_linea_menu(linea_actualizada))
        lineas_editadas = normalizar_lineas_menu(lineas_editadas)
        if lineas_editadas != st.session_state.get("menu_actual", []):
            st.session_state["menu_actual"] = lineas_editadas
            lineas_menu_actual = lineas_editadas

        st.markdown("#### Acciones por plato")
        for idx, linea in enumerate(lineas_menu_actual):
            fila_col1, fila_col2, fila_col3, fila_col4 = st.columns([4, 1, 1, 1])
            with fila_col1:
                st.write(f"{linea.get('orden', idx + 1)}. {linea.get('codigo_receta', '')} · {linea.get('nombre_receta', '')}")
            with fila_col2:
                if st.button("Subir", key=f"subir_menu_{idx}", disabled=idx == 0, use_container_width=True):
                    lineas_reordenadas = list(lineas_menu_actual)
                    lineas_reordenadas[idx - 1], lineas_reordenadas[idx] = lineas_reordenadas[idx], lineas_reordenadas[idx - 1]
                    for orden, item in enumerate(lineas_reordenadas, start=1):
                        item["orden"] = orden
                    st.session_state["menu_actual"] = normalizar_lineas_menu(lineas_reordenadas)
                    st.rerun()
            with fila_col3:
                if st.button("Bajar", key=f"bajar_menu_{idx}", disabled=idx == len(lineas_menu_actual) - 1, use_container_width=True):
                    lineas_reordenadas = list(lineas_menu_actual)
                    lineas_reordenadas[idx + 1], lineas_reordenadas[idx] = lineas_reordenadas[idx], lineas_reordenadas[idx + 1]
                    for orden, item in enumerate(lineas_reordenadas, start=1):
                        item["orden"] = orden
                    st.session_state["menu_actual"] = normalizar_lineas_menu(lineas_reordenadas)
                    st.rerun()
            with fila_col4:
                if st.button("Quitar del menú", key=f"quitar_menu_{idx}", use_container_width=True):
                    lineas_reordenadas = list(lineas_menu_actual)
                    lineas_reordenadas.pop(idx)
                    for orden, item in enumerate(lineas_reordenadas, start=1):
                        item["orden"] = orden
                    st.session_state["menu_actual"] = normalizar_lineas_menu(lineas_reordenadas)
                    st.rerun()
    else:
        st.info("Añade recetas guardadas para construir el menú actual.")

    total_coste_menu = (
        calcular_coste_menu(lineas_menu_actual)
        if lineas_menu_actual else 0.0
    )
    lineas_con_precio = [
        linea for linea in lineas_menu_actual
        if linea.get("precio_total_linea") is not None and not pd.isna(linea.get("precio_total_linea"))
    ]
    hay_precio_menu = bool(lineas_con_precio)
    total_precio_menu = sum(numero_seguro(linea.get("precio_total_linea", 0.0), 0.0) for linea in lineas_con_precio)
    resumen_col1, resumen_col2, resumen_col3 = st.columns(3)
    with resumen_col1:
        st.metric("Recetas / platos", f"{len(lineas_menu_actual)}")
        st.metric("Pax de referencia", f"{int(numero_comensales)}")
    with resumen_col2:
        st.metric("Coste total del menú", f"{total_coste_menu:.2f} €")
    with resumen_col3:
        if hay_precio_menu:
            st.metric("Precio total del menú", f"{total_precio_menu:.2f} €")
        else:
            st.metric("Precio total del menú", "Sin datos")

    accion_menu_col1, accion_menu_col2, accion_menu_col3, accion_menu_col4 = st.columns(4)
    with accion_menu_col1:
        if st.button("Guardar menú", type="primary", use_container_width=True):
            nombre_menu_limpio = str(nombre_menu or "").strip()
            if not supabase_disponible or supabase is None:
                st.error("El inventario no está conectado correctamente.")
            elif not nombre_menu_limpio:
                st.error("Indica un nombre de menú antes de guardar.")
            elif not lineas_menu_actual:
                st.error("Añade al menos una receta antes de guardar el menú.")
            else:
                datos_menu = {
                    "user_id": obtener_user_id_actual(),
                    "nombre": nombre_menu_limpio,
                    "tipo_menu": tipo_menu,
                    "descripcion": "",
                    "numero_comensales": int(numero_comensales),
                    "coste_total": float(total_coste_menu),
                    "precio_total": float(total_precio_menu) if hay_precio_menu else None
                }
                ok, mensaje, menu_guardado = guardar_menu_supabase(datos_menu, lineas_menu_actual)
                if ok:
                    st.session_state["menu_id"] = menu_guardado.get("id")
                    st.success(f"{mensaje}: {menu_guardado.get('nombre', nombre_menu_limpio)}.")
                else:
                    st.error(mensaje)
    with accion_menu_col2:
        if st.button("Actualizar menú seleccionado", use_container_width=True):
            menu_id_actual = st.session_state.get("menu_id")
            nombre_menu_limpio = str(nombre_menu or "").strip()
            if not supabase_disponible or supabase is None:
                st.error("El inventario no está conectado correctamente.")
            elif not menu_id_actual:
                st.warning("Carga un menú guardado antes de actualizarlo.")
            elif not nombre_menu_limpio:
                st.error("Indica un nombre de menú antes de actualizar.")
            elif not lineas_menu_actual:
                st.warning("Añade al menos una receta antes de actualizar el menú.")
            else:
                datos_menu = {
                    "nombre": nombre_menu_limpio,
                    "tipo_menu": tipo_menu,
                    "descripcion": "",
                    "numero_comensales": int(numero_comensales),
                    "coste_total": float(total_coste_menu),
                    "precio_total": float(total_precio_menu) if hay_precio_menu else None
                }
                ok, mensaje = actualizar_menu_supabase(menu_id_actual, datos_menu, lineas_menu_actual)
                if ok:
                    st.success("Menú actualizado correctamente.")
                else:
                    st.error(mensaje)
    with accion_menu_col3:
        if st.button("Duplicar menú seleccionado", use_container_width=True):
            menu_id_actual = st.session_state.get("menu_id")
            nombre_menu_limpio = str(nombre_menu or "").strip()
            if not supabase_disponible or supabase is None:
                st.error("El inventario no está conectado correctamente.")
            elif not menu_id_actual:
                st.warning("Carga un menú guardado antes de duplicarlo.")
            elif not nombre_menu_limpio:
                st.error("Indica un nombre de menú antes de duplicarlo.")
            elif not lineas_menu_actual:
                st.warning("Añade al menos una receta antes de duplicar el menú.")
            else:
                nuevo_nombre_menu = generar_nombre_copia_menu(nombre_menu_limpio)
                datos_menu = {
                    "user_id": obtener_user_id_actual(),
                    "nombre": nuevo_nombre_menu,
                    "tipo_menu": tipo_menu,
                    "descripcion": "",
                    "numero_comensales": int(numero_comensales),
                    "coste_total": float(total_coste_menu),
                    "precio_total": float(total_precio_menu) if hay_precio_menu else None
                }
                ok, mensaje, menu_guardado = guardar_menu_supabase(datos_menu, lineas_menu_actual)
                if ok:
                    st.session_state["menu_id"] = menu_guardado.get("id")
                    st.session_state["menu_nombre_pendiente"] = menu_guardado.get("nombre", nuevo_nombre_menu)
                    st.session_state["menu_tipo_pendiente"] = tipo_menu
                    st.session_state["menu_numero_comensales_pendiente"] = int(numero_comensales)
                    st.session_state["sincronizar_campos_menu"] = True
                    st.session_state["mensaje_menu_cargado"] = (
                        f"{'Menú duplicado en tu cuenta' if obtener_user_id_actual() else 'Menú duplicado correctamente'}: "
                        f"{st.session_state['menu_nombre_pendiente']}."
                    )
                    st.rerun()
                else:
                    st.error(mensaje)
    with accion_menu_col4:
        if st.button("Limpiar menú actual", use_container_width=True):
            st.session_state["menu_actual"] = []
            st.session_state["menu_id"] = None
            st.success("Menú actual limpiado.")
            st.rerun()

